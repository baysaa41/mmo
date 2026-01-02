#!/usr/bin/env python
"""
Province Manager Template Generator

Аймаг/дүүргүүдийн боловсролын мэргэжилтнүүдийг бүртгэх Excel template үүсгэнэ.
Аймаг, дүүрэг бүрт тусдаа файл үүсгэнэ.
"""

import os
import sys
import django
import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Django setup
sys.path.insert(0, '/home/deploy/django/mmo')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mmo.settings')
django.setup()

from accounts.models import Province


def create_instruction_sheet(wb, blank=False):
    """Зааварчилгаа sheet үүсгэх"""
    ws_info = wb.create_sheet('ЗААВАРЧИЛГАА', 0)

    instructions = [
        ['АЙМАГ/ДҮҮРГИЙН МЭРГЭЖИЛТЭН БҮРТГЭХ TEMPLATE - ЗААВАРЧИЛГАА'],
        [''],
        ['1. ЕРӨНХИЙ МЭДЭЭЛЭЛ'],
        ['   Энэ файл нь аймаг/дүүргүүдийн боловсролын мэргэжилтнүүдийг бүртгэх зориулалттай.'],
        ['   Аймаг/Дүүрэг бүрт 3 мөр байгаа:'],
        ['   - 1 мөр: Холбоо барих хүн (үндсэн) - ЗААВАЛ бөглөх (ногоон өнгөөр тэмдэглэгдсэн)'],
        ['   - 2 мөр: Нэмэлт админ хүмүүс - Заавал биш'],
        [''],
    ]

    if not blank:
        instructions.extend([
            ['   ⚠️ ОДОО БАЙГАА ХОЛБОО БАРИХ ХҮМҮҮСИЙН МЭДЭЭЛЭЛ АЛЬ ХЭДИЙН БӨГЛӨГДСӨН БАЙНА'],
            ['   ⚠️ Хэрэв өөрчлөх шаардлагатай бол зөвхөн холбогдох мөрийг засаж болно'],
            [''],
        ])

    instructions.extend([
        ['2. БАГАНЫН ТАЙЛБАР'],
        ['   № (A)                      - Дугаар (автоматаар үүссэн)'],
        ['   Аймаг/Дүүргийн ID (B)      - Аймгийн ID (ӨӨРЧЛӨХГҮЙ)'],
        ['   Аймаг/Дүүргийн нэр (C)     - Аймгийн нэр (ӨӨРЧЛӨХГҮЙ)'],
        ['   Хэрэглэгчийн ID (D)        - User.id (одоо байгаа хэрэглэгчийн ID, заавал биш)'],
        ['   Овог (E)                   - Овог (ЗААВАЛ - холбоо барих хүн)'],
        ['   Нэр (F)                    - Нэр (ЗААВАЛ - холбоо барих хүн)'],
        ['   Имэйл (G)                  - Имэйл хаяг (ЗААВАЛ, уникал байх ёстой)'],
        ['   Үүрэг (H)                  - Үүрэг (ӨӨРЧЛӨХГҮЙ: contact эсвэл admin)'],
        ['   Тэмдэглэл (I)              - Тэмдэглэл (заавал биш)'],
        [''],
        ['3. БӨГЛӨХ ЗААВАР'],
        ['   А) Холбоо барих хүн (ногоон өнгөөр тэмдэглэгдсэн мөрүүд):'],
    ])

    if blank:
        instructions.append(['      - Овог, Нэр, Имэйл ЗААВАЛ бөглөх'])
    else:
        instructions.extend([
            ['      - Одоо байгаа мэдээлэл БӨГЛӨГДСӨН байна'],
            ['      - Засах шаардлагатай бол Овог, Нэр, Имэйл засаж болно'],
        ])

    instructions.extend([
        ['      - Хэрэв шинэ хүн оруулах бол Хэрэглэгчийн ID-г хоослож, шинэ мэдээлэл оруулна'],
        [''],
        ['   Б) Нэмэлт админ хүмүүс (цагаан мөрүүд):'],
        ['      - ЗААВАЛ БИШИ, хэрэв нэмэлт хүн байвал бөглөнө'],
        ['      - Овог, Нэр, Имэйл бөглөх'],
        ['      - Хэрэв хэрэглэгч аль хэдийн системд байвал Хэрэглэгчийн ID бичих'],
        [''],
        ['4. АНХААРУУЛГА'],
        ['   ⚠️ Имэйл хаяг нь УНИКАЛ байх ЁСТОЙ'],
        ['   ⚠️ Холбоо барих хүн нь аймаг бүрт ЗААВАЛ байх ёстой'],
        ['   ⚠️ №, Аймаг/Дүүргийн ID, Аймаг/Дүүргийн нэр, Үүрэг баганыг ӨӨРЧЛӨХГҮЙ'],
        ['   ⚠️ Хэрэв нэмэлт админ нэмэхгүй бол мөрийг ХООСОН үлдээнэ (устгахгүй)'],
        [''],
        ['5. ЖИШЭЭ - Шинэ нэмэлт админ нэмэх'],
        ['   № | Аймаг/Дүүргийн ID | Аймаг/Дүүргийн нэр | Хэрэглэгчийн ID | Овог    | Нэр  | Имэйл              | Үүрэг  | Тэмдэглэл'],
        ['   2 | 1                 | Архангай аймаг     |                 | Доржийн | Баяр | bayar@example.com  | admin  | Нэмэлт админ #1'],
        [''],
        ['6. ИМПОРТ ХИЙХ КОМАНД'],
        ['   # Эхлээд шалгах (dry-run):'],
        ['   python manage.py import_province_managers <файлын_нэр>.xlsx --dry-run'],
        [''],
        ['   # Бодит импорт:'],
        ['   python manage.py import_province_managers <файлын_нэр>.xlsx'],
    ])

    for row_idx, row_data in enumerate(instructions, 1):
        ws_info.cell(row=row_idx, column=1, value=row_data[0])
        if row_idx == 1:
            ws_info.cell(row=row_idx, column=1).font = Font(bold=True, size=14, color='366092')
        elif row_data[0].startswith(('1.', '2.', '3.', '4.', '5.', '6.')):
            ws_info.cell(row=row_idx, column=1).font = Font(bold=True, size=12)

    ws_info.column_dimensions['A'].width = 120


def create_template(blank=False):
    """Excel template үүсгэх - province бүрт тусдаа файл"""

    output_dir = 'province_managers_export'

    # Хавтас үүсгэх
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f'📁 Хавтас үүслээ: {output_dir}')

    # Province-уудыг авах
    provinces = Province.objects.all().order_by('id')

    files_created = []
    total_rows = 0

    # Province бүрээр тусдаа файл үүсгэх
    for p in provinces:
        # Excel workbook үүсгэх
        wb = Workbook()
        ws = wb.active
        ws.title = 'Province Managers'

        # Header - "Холбоо барих хүний ID" хасав
        headers = ['№', 'Аймаг/Дүүргийн ID', 'Аймаг/Дүүргийн нэр',
                   'Хэрэглэгчийн ID', 'Овог', 'Нэр', 'Имэйл', 'Үүрэг', 'Тэмдэглэл']

        # Header style
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header мөр бичих
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Contact person мэдээлэл авах (blank бол хоосон)
        contact_user_id = ''
        contact_last_name = ''
        contact_first_name = ''
        contact_email = ''

        if not blank and p.contact_person:
            contact_user_id = p.contact_person.id
            contact_last_name = p.contact_person.last_name
            contact_first_name = p.contact_person.first_name
            contact_email = p.contact_person.email or ''

        # Contact person мөр (ногоон өнгөөр)
        contact_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

        row_data = [
            [1, p.id, p.name, contact_user_id,
             contact_last_name, contact_first_name, contact_email,
             'contact', 'Үндсэн холбоо барих хүн'],
            [2, p.id, p.name, '', '', '', '', 'admin', 'Нэмэлт админ #1 (заавал биш)'],
            [3, p.id, p.name, '', '', '', '', 'admin', 'Нэмэлт админ #2 (заавал биш)'],
        ]

        for row_idx, data in enumerate(row_data, 2):
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

                # Contact person мөрийг ногоон өнгөөр
                if row_idx == 2:
                    cell.fill = contact_fill

        # Баганын өргөнийг тохируулах
        ws.column_dimensions['A'].width = 6   # №
        ws.column_dimensions['B'].width = 18  # Аймаг/Дүүргийн ID
        ws.column_dimensions['C'].width = 35  # Аймаг/Дүүргийн нэр
        ws.column_dimensions['D'].width = 18  # Хэрэглэгчийн ID
        ws.column_dimensions['E'].width = 18  # Овог
        ws.column_dimensions['F'].width = 18  # Нэр
        ws.column_dimensions['G'].width = 35  # Имэйл
        ws.column_dimensions['H'].width = 10  # Үүрэг
        ws.column_dimensions['I'].width = 35  # Тэмдэглэл

        # Зааварчилгаа sheet нэмэх
        create_instruction_sheet(wb, blank=blank)

        # Файлын нэр үүсгэх (зайгүй, underscore-ээр, _admin залгаартай)
        safe_province_name = p.name.replace('/', '_').replace('\\', '_').replace(',', '').replace(' ', '_')
        suffix = '_admin_blank' if blank else '_admin'
        output_file = os.path.join(output_dir, f'{safe_province_name}{suffix}.xlsx')

        # Файл хадгалах
        wb.save(output_file)
        files_created.append(output_file)
        total_rows += 3

        print(f'✅ {p.name}: {output_file}')

    # Эцсийн тайлан
    print('')
    print('=' * 80)
    print('📊 ЭЦСИЙН ТАЙЛАН')
    print('=' * 80)
    print(f'📁 Хавтас: {os.path.abspath(output_dir)}')
    print(f'📄 Нийт файл: {len(files_created)}')
    print(f'🏛️  Нийт province: {len(provinces)}')
    print(f'📝 Нийт мөр: {total_rows} (province бүрт 3 мөр)')
    print(f'👤 Contact person мөрүүд: {len(provinces)} (ногоон өнгөөр)')
    print(f'👥 Admin мөрүүд: {len(provinces) * 2} (заавал биш)')
    if blank:
        print(f'⚪ Горим: BLANK (бүх мэдээлэл хоосон)')
    else:
        print(f'✏️  Горим: PRE-FILLED (одоо байгаа contact person мэдээлэл бөглөгдсөн)')
    print('=' * 80)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Province Managers Template Generator')
    parser.add_argument('--blank', action='store_true',
                        help='Бүх мэдээлэл хоосон template үүсгэх (одоо байгаа contact person мэдээллийг бөглөхгүй)')
    args = parser.parse_args()

    create_template(blank=args.blank)
