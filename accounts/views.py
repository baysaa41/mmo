import math

from django.shortcuts import render, redirect, HttpResponse
from accounts.forms import UserForm, UserMetaForm, EmailForm
from django.contrib.auth.models import User, Group
from django.contrib.auth.views import PasswordResetView
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from .forms import LoginForm, CustomPasswordResetForm
from .models import UserMeta, Province, Level, Grade, UserMails, Zone
from django.urls import reverse
from django.core.mail import send_mass_mail
from olympiad.models import Article, SchoolYear, Result
from datetime import datetime, timezone



from django.db import connection
from django.template.loader import render_to_string
from django.utils.html import strip_tags

from django.core.mail import get_connection, EmailMultiAlternatives, EmailMessage

from django_pandas.io import read_frame

import os
import pandas as pd
import numpy as np
import string
import random
import re
import openpyxl


# Create your views here.

def index(request):
    # shine hereglegchiin burtgel
    if request.user.is_authenticated:
        meta = UserMeta.objects.get_or_create(user=request.user)
        if not meta[0].is_valid:
            return redirect('user_profile')

    now = datetime.now(timezone.utc)
    mode = request.GET.get('mode', 0)

    school_year = SchoolYear.objects.filter(start__lt=now, end__gt=now).first()
    if school_year:
        id = request.GET.get('year', school_year.id)
        year = SchoolYear.objects.filter(pk=id).first()
    else:
        year = SchoolYear.objects.create()
        year.id = 0
        school_year = SchoolYear.objects.create()
        school_year.id = 0

    prev = SchoolYear.objects.filter(pk=year.id - 1).first()
    next = SchoolYear.objects.filter(pk=year.id + 1).first()

    articles = Article.objects.filter(year_id=year.id, IsShow=True).exclude(EndDate__lt=now).order_by('-IsSpec',
                                                                                                      '-StartDate')

    context = {'articles': articles, 'mode': mode, 'year': year, 'prev': prev, 'next': next}

    return render(request, 'accounts/site_home.html', context=context)


def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    form = LoginForm(request.POST or None)
    if form.is_valid():
        username = form.cleaned_data.get('username')
        password = form.cleaned_data.get('password')
        user = authenticate(request, username=username, password=password)
        if user != None:
            login(request, user)
            is_valid, meta = UserMeta.objects.get_or_create(user=user)
            if is_valid:
                next = request.GET.get('next', '/')
            else:
                next = reverse('user_profile')
            return redirect(next)
        else:
            request.session['invalid_user'] = 1
    return render(request, 'accounts/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return render(request, 'accounts/logout.html')


@login_required(login_url='/accounts/login/')
def profile(request):
    if request.POST:
        form1 = UserForm(request.POST, request.FILES, instance=request.user)
        form1.save()
        instance = UserMeta.objects.get(user_id=request.user.id)
        form2 = UserMetaForm(request.POST, instance=instance)
        form2.save()
        return redirect('user_profile_done')
    form1 = UserForm(instance=request.user)
    user_meta = UserMeta.objects.get_or_create(user_id=request.user.id)
    form2 = UserMetaForm(instance=user_meta[0])
    return render(request, 'accounts/register.html', {'form1': form1, 'form2': form2})


def profile_ready(request):
    return render(request, 'accounts/profile_ready.html')


class CustomPasswordResetView(PasswordResetView):
    form_class = CustomPasswordResetForm
    email_template_name = 'registration/password_reset_email.html'
    success_url = '/password_reset/done/'  # customize as needed

def users(request):
    # create_users()
    provinces = Province.objects.all()
    levels = Level.objects.all()
    p = request.GET.get('p', 0)
    l = request.GET.get('l', 0)
    if p == 0 and l == 0:
        return render(request, 'accounts/users-home.html', {'provinces': provinces, 'levels': levels})
    elif p != 0 and l != 0:
        users = UserMeta.objects.filter(province_id=int(p), level_id=int(l), is_valid=1).order_by('level_id')
        province = Province.objects.get(pk=p)
        level = Level.objects.get(pk=l)
        title = province.name + ', ' + level.name
        return render(request, 'accounts/users.html', {'users': users, 'title': title})
    elif p != 0:
        users = UserMeta.objects.filter(province_id=int(p), is_valid=1).order_by('level_id', 'grade_id')
        province = Province.objects.get(pk=p)
        return render(request, 'accounts/users.html', {'users': users, 'title': province.name})
    else:
        users = UserMeta.objects.filter(level_id=int(l), is_valid=1).order_by('grade_id', 'province_id')
        level = Level.objects.get(pk=l)
        return render(request, 'accounts/users.html', {'users': users, 'title': level.name})


def special_members(request):
    if request.user.is_staff:
        users = UserMeta.objects.filter(is_valid=1).filter(grade_id__gt=12).order_by('grade', 'province')
        return render(request, 'accounts/staff.html', {'users': users})
    else:
        return redirect('account_users')


def staff(request):
    users = User.objects.filter(is_staff=1).order_by('data__province__zone')
    return render(request, 'accounts/staff.html', {'users': users})


def group_users(request, group_id):
    try:
        group = Group.objects.get(pk=group_id)
    except:
        return render(request, 'errors/error.html', {'message': 'Ийм бүлэг байхгүй.'})
    pd.options.display.float_format = '{:,.0f}'.format

    pid = int(request.GET.get('p', 0))
    zid = int(request.GET.get('z', 0))

    users = group.user_set.all()
    if pid:
        users = users.filter(data__province__id=pid)
    elif zid:
        users =users.filter(data__province__zone__id=zid)

    users = group.user_set.all().order_by('data__province__zone', 'data__province', 'data__school')

    z_id = request.GET.get('z', False)
    if z_id:
        try:
            zone = Zone.objects.filter(pk=z_id).first()
            title = group.name + ', ' + zone.name
            users = users.filter(data__province__zone_id=z_id)
        except Exception as e:
            print(str(e))

    p_id = request.GET.get('p', False)
    if p_id:
        try:
            province = Province.objects.filter(pk=p_id).first()
            title = group.name + ', ' + province.name
            users = users.filter(data__province_id=p_id)
        except Exception as e:
            print(str(e))

    g_id = request.GET.get('g', False)
    if g_id:
        try:
            grade = Grade.objects.filter(pk=g_id).first()
            title = title + ', ' + grade.name
            users = users.filter(data__grade_id=g_id)
        except Exception as e:
            print(str(e))
    if request.user.is_staff:
        users_df = read_frame(users,
                              fieldnames=['id', 'username', 'last_name', 'first_name', 'data__province__name',
                                          'data__school', 'data__grade__name', 'data__reg_num','data__mobile', 'email'],
                              verbose=False)
        users_df['data__mobile'] = users_df['data__mobile'].astype(pd.Int64Dtype())
        users_df.rename(columns={
            'id': 'ID',
            'first_name': 'Нэр',
            'last_name': 'Овог',
            'username': 'Хэрэглэгчийн нэр',
            'data__province__name': 'Аймаг/Дүүрэг',
            'data__school': 'Cургууль',
            'data__grade__name': 'Анги',
            'data__reg_num': 'Регистрын дугаар',
            'data__mobile': 'Гар утас',
            'email': 'И-мэйл'
        }, inplace=True)
    else:
        users_df = read_frame(users,
                              fieldnames=['id', 'username', 'last_name', 'first_name', 'data__province__name',
                                          'data__school', 'data__grade__name'],
                              verbose=False)

        users_df.rename(columns={
            'id': 'ID',
            'first_name': 'Нэр',
            'last_name': 'Овог',
            'username': 'Хэрэглэгчийн нэр',
            'data__province__name': 'Аймаг/Дүүрэг',
            'data__school': 'Cургууль',
            'data__grade__name': 'Анги'
        }, inplace=True)

    users_df.index = np.arange(1, users_df.__len__() + 1)

    styled_df = (users_df.style.set_table_attributes('classes="table table-bordered table-hover"').set_table_styles([
        {
            'selector': 'th',
            'props': [('text-align', 'center')],
        },
        {
            'selector': 'td, th',
            'props': [('border', '1px solid #ccc')],
        },
        {
            'selector': 'td, th',
            'props': [('padding', '3px 5px 3px 5px')],
        },
    ]))

    last_hack = styled_df.to_html(classes='table table-bordered table-hover',na_rep="",escape=False)
    last_hack = re.sub('<th class="blank level0" >&nbsp;</th>','<th class="blank level0" >№</th>', last_hack)

    context = {
        'title': group.name,
        'pivot': last_hack,
    }
    return render(request, 'accounts/pd-users.html', context)

def get_active_emails():
    with connection.cursor() as cursor:
        cursor.execute("select distinct email from auth_user where is_active=1 and email<>''")
        emails = cursor.fetchall()

    return emails


def create_diploms_mail(request):
    num = 0
    subject = 'Охидын багыг дэмжих олммпиад батламж'
    from_email = 'ММОХ <no-reply@mmo.mn>'
    ids = [100, 101, 102, 103, 104]
    num = 0
    for id in ids:
        with connection.cursor() as cursor:
            cursor.execute("select distinct contestant_id from  olympiad_result where olympiad_id=%s", [id])
            results = cursor.fetchall()
            for result in results:
                contestant = User.objects.get(pk=result[0])
                text_html = render_to_string("newsletter/certificate_email.html",
                                             {'quiz_id': id, 'contestant_id': result[0]})
                text = strip_tags(text_html)
                UserMails.objects.create(subject=subject, text=text, text_html=text_html, from_email=from_email,
                                         to_email=contestant.email)
                num = num + 1
    return HttpResponse(num)


def create_emails(request):
    num = 0
    subject = 'ММОХ, Мэдээллийн товхимол №1'
    text_html = render_to_string("newsletter/202301.html")
    text = strip_tags(text_html)
    from_email = 'ММОХ <newsletter@mmo.mn>'

    emails = get_active_emails()
    for email in emails:
        UserMails.objects.create(subject=subject, text=text, text_html=text_html, from_email=from_email,
                                 to_email=email[0])
        num = num + 1
    return HttpResponse(num)


def create_mails(request):
    num = 0
    users = User.objects.all()
    subject = 'I шатны шалгаруулалт'
    text = '''
Та бүхэнд 2021-2022 оны хичээлийн жилийн мэндийг хүргэе!

2021-2022 оны хичээлийн жилээс эхлэн ММОХ I шатны сонгон шалгаруулалт хийхээр болж байна.

Шалгаруулалтын талаарх мэдээллийг www.mmo.mn сайтын мэдээллийн хэсгээс харж болно.

Шалгаруулалтад оролцохын тулд өөрийн бүртгэлийн мэдээллийг шинэчилсэн байх шаардлагатай.

Таны нэвтрэх нэр: {}

Хэрвээ та нууц үгээ мартсан бол нэвтрэх хуудасны Нууц үгээ сэргээх линк ашиглаарай.

Хэрвээ танд ямар нэг асуулт байвал baysa.edu@gmail.com хаягаар холбогдож асуугаарай 
(Ажлын ачааллаас шалтгаалж хариулт саатаж очихыг анхаарна уу!)

ММОХ
    '''
    from_email = 'Монголын Математикийн Олимпиадын Хороо <no-reply@mmo.mn>'

    for user in users:
        if user.is_active:
            UserMails.objects.create(subject=subject, text=text.format(user.username), from_email=from_email,
                                     to_email=user.email)
            num = num + 1
    return HttpResponse(num)


def send_user_mails(request):
    uemails = UserMails.objects.filter(is_sent=False)[0:50]
    emails = ()
    for email in uemails:
        new = (email.subject, email.text, email.from_email, [email.to_email])
        email.is_sent = True
        email.save()
        emails = (*emails, new)
    return HttpResponse(send_mass_mail(emails))


def send_mass_html_mail(request):
    """
    Given a datatuple of (subject, text_content, html_content, from_email,
    recipient_list), sends each message to each recipient list. Returns the
    number of emails sent.

    If from_email is None, the DEFAULT_FROM_EMAIL setting is used.
    If auth_user and auth_password are set, they're used to log in.
    If auth_user is None, the EMAIL_HOST_USER setting is used.
    If auth_password is None, the EMAIL_HOST_PASSWORD setting is used.

    """
    connection = get_connection(fail_silently=False)
    messages = []
    uemails = UserMails.objects.filter(is_sent=False)[0:50]
    for uemail in uemails:
        message = EmailMultiAlternatives(uemail.subject, uemail.text, uemail.from_email, [uemail.to_email])
        message.attach_alternative(uemail.text_html, 'text/html')
        messages.append(message)
        uemail.is_sent = True
        uemail.save()
    return HttpResponse(connection.send_messages(messages))

#import from sisi
def import_users():

    dir = '/home/deploy/results/students'

    list = os.listdir(dir)
    print(list)
    num = 0
    for f in list:
        name = dir + '/' + f
        print(name)
        filename, file_extension = os.path.splitext(name)
        print(file_extension)
        if file_extension.lower() in ['.xls', '.xlsx']:
            try:
                df = pd.read_excel(name, 'Sheet1', engine='openpyxl')
                for item in df.iterrows():
                    ind, row = item
                    user, created = User.objects.get_or_create(username=row[row.keys()[3]])
                    if created:
                        user.firstname = row[row.keys()[2]]
                        user.set_password(row[row.keys()[3]])
                        user.save()
                        m, c = UserMeta.objects.get_or_create(user=user)
                        m.is_valid = True
                        m.save()
                        num = num + 1
            except:
                print('error')
                pass

    return num


def random_salt(n=8):
    characterList = string.ascii_letters + string.digits
    salt = ['s']
    for i in range(n):
        randomchar = random.choice(characterList)
        salt.append(randomchar)
    return "".join(salt)


# Define a function for
# for validating an Email
def check(email):
    regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
    # pass the regular expression
    # and the string into the fullmatch() method
    if (re.fullmatch(regex, email)):
        return True
    else:
        return False

# obsolete
def register_sheet(name, sheet, teacher):
    num = 0
    users = []
    if sheet == 'C (5-6)':
        pgroup = Group.objects.get(pk=26)
        mgroup = Group.objects.get(pk=30)
    elif sheet == 'D (7-8)':
        pgroup = Group.objects.get(pk=27)
        mgroup = Group.objects.get(pk=31)
    elif sheet == 'E (9-10)':
        pgroup = Group.objects.get(pk=28)
        mgroup = Group.objects.get(pk=32)
    else:
        pgroup = Group.objects.get(pk=29)
        mgroup = Group.objects.get(pk=33)
    text = '''{}
    Шинээр бүртгүүлсэн:
    {}
    
    Өмн
    '''
    text_rows = ''
    html = '''
    <h4>{}</h4>
    <table border=1 cellspacing=0>
    <tr><th>Овог</th><th>Нэр</th><th>ID</th><th>Хэрэглэгчийн нэр</th><th>Нууц үг</th></tr>
    {}
    </table>'''
    html_rows = ''
    error = sheet
    error_html = '''
    <h4>{}</h4>
    <table border=1 cellspacing=0>
    <tr><th>Овог</th><th>Нэр</th><th>ID</th><th>Хэрэглэгчийн нэр</th><th>Тайлбар</th></tr>{}
    </table>'''
    error_rows = ''
    try:
        df = pd.read_excel(name, sheet, engine='openpyxl')
        df['ID'] = pd.to_numeric(df['ID'], errors='coerce').fillna(0).astype(int)
        df['Овог'] = df['Овог'].fillna('')
        df['Нэр'] = df['Нэр'].fillna('')
        df['Утасны дугаар'] = pd.to_numeric(df['Утасны дугаар'], errors='coerce').fillna(0).astype(int)
        df['E-mail'] = df['E-mail'].fillna('')
        df['Аймаг, Дүүргийн ID код'] = pd.to_numeric(df['Аймаг, Дүүргийн ID код'], errors='coerce').fillna(30).astype(int)
        df['Сургууль'] = df['Сургууль'].fillna('')
        df['Анги'] = pd.to_numeric(df['Анги'], errors='coerce').fillna(12).astype(int)
        df['Бэлтгэл даваанд оролцох эсэх'] = pd.to_numeric(df['Бэлтгэл даваанд оролцох эсэх'], errors='coerce').fillna(0).astype(int)
    except Exception as e:
        print("df aldaa")
        print(e)

    if sheet == 'C (5-6)':
        level_id = 2
    elif sheet == 'D (7-8)':
        level_id = 3
    elif sheet == 'E (9-10)':
        level_id = 4
    elif sheet == 'F (11-12)':
        level_id = 5
    else:
        level_id = 8

    print(name, sheet)
    for item in df.iterrows():
        ind, row = item
        # print(row)
        salt = random_salt(8)
        # user_id = 0
        try:
            # print(row['ID'])
            user_id = row['ID']
            # print("user_id:", user_id)
            user = User.objects.get(pk=user_id)
            text_rows = text_rows + '\n' + ', '.join(
                [row['Овог'], row['Нэр'], str(user.id), user.username,'Бүртгэлтэй имэйл: ' + user.email])
            error_rows = (error_rows + "<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>"
                .format(row['Овог'], row['Нэр'], user.id, user.username,'Бүртгэлтэй имэйл: ' + user.email))
        except User.DoesNotExist:
            meta = UserMeta.objects.filter(reg_num=row['Регистрийн дугаар']).first()
            if meta:
                user = meta.user
                text_rows = text_rows + '\n' + ', '.join(
                    [row['Овог'], row['Нэр'], str(user.id), user.username, 'Бүртгэлтэй имэйл: ' + user.email])
                error_rows = (error_rows + "<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>"
                              .format(row['Овог'], row['Нэр'], user.id, user.username,
                                      'Бүртгэлтэй имэйл: ' + user.email))
            else:
                email = row['E-mail'].strip()
                if not check(email):
                    email = teacher.email
                if not check(email):
                    email = 'mmo60official@gmail.com'

                user = User.objects.create_user(
                    username=salt,
                    first_name=row['Нэр'],
                    last_name=row['Овог'],
                    email=email
                )

                user.username = 'u' + str(user.id)
                user.set_password(salt)
                try:
                    user.save()
                    text_rows = text_rows + "\n" + ",".join([user.last_name, user.first_name, user.username, salt])
                    html_rows = (html_rows + "<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>"
                                 .format(user.last_name, user.first_name, user.id, user.username, salt))
                except:
                    user.username = 'u' + str(user.id) + random_salt(3)
                    user.set_password(salt)
                    user.save()
                    text_rows = text_rows + "\n" + ",".join([user.last_name, user.first_name, user.username, salt])
                    html_rows = (html_rows + "<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>"
                                 .format(user.last_name, user.first_name, user.id, user.username, salt))
                num = num + 1
                if user.first_name == '' and user.last_name == '':
                    user.first_name = row['Нэр']
                    user.last_name = row['Овог']
                    user.save()
        except Exception as e:
            print(e)

        m, c = UserMeta.objects.get_or_create(user=user)
        if c:
            m.mobile = int(float(row['Утасны дугаар']))
            m.grade_id = int(row['Анги'])
            m.level_id = level_id
            m.province_id = int(row['Аймаг, Дүүргийн ID код'])
            m.school = row['Сургууль']
            try:
                m.reg_num = row['Регистрийн дугаар'].strip()
            except:
                m.reg_num = row['Регистрийн дугаар']
            m.is_valid = True
            m.save()
        else:
            m.mobile = int(float(row['Утасны дугаар']))
            m.grade_id = int(row['Анги'])
            m.level_id = level_id
            m.province_id = int(row['Аймаг, Дүүргийн ID код'])
            m.school = row['Сургууль']
            m.save()

        if row['Бэлтгэл даваанд оролцох эсэх']:
            pgroup.user_set.add(user)
        mgroup.user_set.add(user)
        users.append(user.id)

    return num, text.format(sheet, text_rows), html.format(sheet, html_rows), error_html.format(sheet, error_rows)


def register_students():
    dir = '/home/deploy/registration'

    list = os.listdir(dir)

    subject = "ММОХ бүртгэл"
    body = '''
    Та 2023-2024 оны хичээлийн жилд ММОХ-ноос зохион байгуулах олимпиадад оролцох сурагчдаа бүртгүүллээ. 

    Бүртгэлтэй холбоотой мэдээллийг илгээж байна:
    
    {}
    
    ММОХ
        '''
    body_html = '''
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <title>Мэдээллийн товхимол №1</title>
    </head>
    <body>
        <p>Та 2023-2024 оны хичээлийн жилд ММОХ-ноос зохион байгуулах олимпиадад оролцох сурагчдаа бүртгүүллээ.</p>
    
        <p>Бүртгэлтэй холбоотой мэдээллийг илгээж байна:</p>
        
        <p>Шинэээр бүртгүүлсэн:</p>
        {}
        <p>Бусад (Өмнө бүртгүүлсэн сурагчдыг регистрийн дугаар, ID-аар шүүв. Хэрвээ нууц үгээ мартсан бол бүртгэлтэй имэйл хаягаар сэргээнэ):</p>
        {}
    </body>
    </html>
    '''


    from_email = 'ММОХ <no-reply@mmo.mn>'

    num = 0
    for f in list:
        message = ''
        html_message = ''
        errors = ''
        html_errors = ''

        name = dir + '/' + f
        if os.path.isfile(name):
            filename, file_extension = os.path.splitext(name)
            if file_extension.lower() in ['.xls', '.xlsx']:
                print(name)
                province_id = pd.read_excel(name, "school",
                                           usecols="B",
                                           engine='openpyxl',
                                           skiprows=1,
                                           nrows=1,
                                           header=None,
                                           names=["Value"]).iloc[0]["Value"]
                school_name = pd.read_excel(name, "school",
                                            usecols="B",
                                            engine='openpyxl',
                                            skiprows=2,
                                            nrows=1,
                                            header=None,
                                            names=["Value"]).iloc[0]["Value"]
                school_name = str(school_name)
                teacher_id = pd.read_excel(name, "school",
                                           usecols="B",
                                           engine='openpyxl',
                                           skiprows=3,
                                           nrows=1,
                                           header=None,
                                           names=["Value"]).iloc[0]["Value"]

                try:
                    teacher_id = int(teacher_id)
                    teacher = User.objects.get(pk=teacher_id)
                except:
                    teacher = User.objects.get(pk=1)

                group = Group.objects.get(pk=34)
                group.user_set.add(teacher)

                new, text, html, error_html = register_sheet(name, 'C (5-6)', teacher)
                num = num + new
                message = message + '\n' + text
                html_message = html_message + html
                html_errors = html_errors + error_html
                new, text, html, error_html = register_sheet(name, 'D (7-8)', teacher)
                num = num + new
                message = message + '\n' + text
                html_message = html_message + html
                html_errors = html_errors + error_html
                new, text, html, error_html = register_sheet(name, 'E (9-10)', teacher)
                num = num + new
                message = message + '\n' + text
                html_message = html_message + html
                html_errors = html_errors + error_html
                new, text, html, error_html = register_sheet(name, 'F (11-12)', teacher)
                num = num + new
                message = message + '\n' + text
                html_message = html_message + html
                html_errors = html_errors + error_html

                UserMails.objects.create(subject=subject, text=body.format(message),
                             text_html=body_html.format(html_message, html_errors), from_email=from_email,
                             to_email=teacher.email)

                source = name
                dest = dir + '/processed/' + str(province_id)
                if not os.path.isdir(dest):
                    os.mkdir(dest)

                try:
                    os.rename(source, dest + '/' + str(teacher_id) + '-' + school_name.strip().replace(" ","") + '-' + random_salt(3) + file_extension)
                except Exception as e:
                    print(e)

    return num


def check_row(row):
    if row[0].value is not None:
        try:
            id = int(float(row[0].value))
            user = User.objects.get(pk=id)
        except ValueError as e:
            return False, 'ID натурал тоо байх ёстой'
        except TypeError as e:
            return False, 'ID натурал тоо байх ёстой'
        except User.DoesNotExist:
            return False, 'ID буруу.'
    else:
        try:
            reg_num = str(row[3].value).strip()
            user = UserMeta.objects.filter(reg_num=reg_num).first()
        except UserMeta.DoesNotExist:
            if row[1].value is None and row[2].value is None:
                return False, 'Хэрэглэгч үүсгэх боломжгүй. Мэдээлэл дутуу. Нэмэлт мөрийг устгах.'
        except ValueError as e:
            return False, 'Регистрийн дугаар буруу'
        except TypeError as e:
            return False, 'Регистрийн дугаар буруу'

    try:
        mobile = int(float(row[4].value))
    except ValueError as e:
        return False, 'Утасны дугаар буруу. Хэрвээ хоосон мөр бол уг мөрийг арилгаарай.'
    except TypeError as e:
        return False, 'Утасны дугаар буруу. Хэрвээ хоосон мөр бол уг мөрийг арилгаарай.'
    try:
        email = str(row[5].value)
        if email is None:
            if not check(email.strip()):
                return False, 'Email буруу: ' + email
    except Exception as e:
        return False, 'Email алдаатай.'

    try:
        province_id = int(row[6].value)
        province = Province.objects.get(pk=province_id)
    except Province.DoesNotExist:
        return False, 'Аймаг, дүүргийн код буруу. Reference sheet-ээс хар.'
    except Exception as e:
        return False, str(e) + ' Аймаг, дүүргийн код Reference sheet-ээс хар. Тоо байна.'
    try:
        school = str(row[7].value)
        if school == 'None':
            return False, 'Сургуулийн нэр оруулаагүй.'
    except Exception as e:
        return False, str(e) + ' Сургуулийн нэр'
    try:
        grade = int(float(row[8].value))
    except Exception as e:
        return False, str(e)

    return True, ''

def check_error(wb):
    aldaa = False
    messages = list()
    try:
        sheet = wb["school"]
    except Exception as e:
        aldaa = True
        messages.append('school sheet байхгүй байна.')
    try:
        province_id = int(float(sheet['B2'].value))
        province = Province.objects.get(pk=province_id)
    except Exception as e:
        aldaa = True
        messages.append(str(e) + ': B2-д дүүргийн ID-г тоогоор бичнэ. Reference sheet-ээс хар.')

    try:
        school_name = str(sheet['B3'].value).strip()
        if school_name == 'None':
            aldaa = True
            messages.append('B3-д сургуулийн нэр бичнэ.')
    except Exception as e:
        aldaa = True
        messages.append(str(e) + ': Сургуулийн нэр алдаатай.')

    try:
        teacher_id = int(float(sheet['B4'].value))
        teacher = User.objects.get(pk=teacher_id)
        #if not teacher.groups.filter(pk=34) and not teacher.is_staff:
        #    aldaa = True
        #    messages.append('Сэдэв хүлээн авах хүн заавал урьдчилж бүртгүүлсэн байна.')
    except User.DoesNotExist:
        aldaa = True
        messages.append('Сэдэв хүлээн авах хүний ID буруу.')
    except Exception as e:
        aldaa = True
        messages.append(str(e) + ': B4-д сэдэв хүлээн авах хүний ID-г тоогоор бичнэ.')

    if 'C (5-6)' not in wb.sheetnames:
        aldaa = True
        messages.append('C (5-6) sheet байхгүй байна. Хариултын хуудаст засвар оруулж болохгүй.')
    elif 'D (7-8)' not in wb.sheetnames:
        aldaa = True
        messages.append('D (7-8) sheet байхгүй байна. Хариултын хуудаст засвар оруулж болохгүй.')
    elif 'E (9-10)' not in wb.sheetnames:
        aldaa = True
        messages.append('E (9-10) sheet байхгүй байна. Хариултын хуудаст засвар оруулж болохгүй.')
    elif 'F (11-12)' not in wb.sheetnames:
        aldaa = True
        messages.append('F (11-12) sheet байхгүй байна. Хариултын хуудаст засвар оруулж болохгүй.')
    else:
        title_hash = ['ID', 'Овог', 'Нэр', 'Регистрийн дугаар', 'Утасны дугаар', 'E-mail', 'Аймаг, Дүүргийн ID код', 'Сургууль', 'Анги']
        sheet_c = wb['C (5-6)']
        title_c = [sheet_c.cell(row=1,column=i).value for i in range(1, 10)]
        if title_c != title_hash:
            aldaa = True
            messages.append('C (5-6) sheet-ийн толгой өөрчлөгдсөн байна. Хариултын хуудаст засвар оруулж болохгүй.')

        sheet_d = wb['D (7-8)']
        title_d = [sheet_c.cell(row=1,column=i).value for i in range(1, 10)]
        if title_d != title_hash:
            aldaa = True
            messages.append('D (7-8) sheet-ийн толгой өөрчлөгдсөн байна. Хариултын хуудаст засвар оруулж болохгүй.')

        sheet_e = wb['E (9-10)']
        title_e = [sheet_c.cell(row=1,column=i).value for i in range(1, 10)]
        if title_e != title_hash:
            aldaa = True
            messages.append('E (9-10) sheet-ийн толгой өөрчлөгдсөн байна. Хариултын хуудаст засвар оруулж болохгүй.')

        sheet_f = wb['F (11-12)']
        title_f = [sheet_c.cell(row=1,column=i).value for i in range(1, 10)]
        if title_f != title_hash:
            aldaa = True
            messages.append('F (11-12) sheet-ийн толгой өөрчлөгдсөн байна. Хариултын хуудаст засвар оруулж болохгүй.')

        ind = 0
        messages.append('C (5-6)')
        for row in sheet_c:
            if ind > 0:
                result, message = check_row(row)
                message = message + ': Мөрийн дугаар ' + str(ind+1)
                if not result:
                    aldaa = True
                    messages.append(message)
            ind = ind + 1

        ind = 0
        messages.append('D (7-8)')
        for row in sheet_d:
            if ind > 0:
                result, message = check_row(row)
                message = message + ': Мөрийн дугаар: ' + str(ind+1)
                if not result:
                    aldaa = True
                    messages.append(message)
            ind = ind + 1

        ind = 0
        messages.append('E (9-10)')
        for row in sheet_e:
            if ind > 0:
                result, message = check_row(row)
                message = message + ': Мөрийн дугаар ' + str(ind+1)
                if not result:
                    aldaa = True
                    messages.append(message)
            ind = ind + 1

        ind = 0
        messages.append('F (11-12)')
        for row in sheet_f:
            if ind > 0:
                result, message = check_row(row)
                message = message + ': Мөрийн дугаар: ' + str(ind+1)
                if not result:
                    aldaa = True
                    messages.append(message)
            ind = ind + 1

    return aldaa, messages

def import_row(row, level_id):
    message = ''
    if row[0].value is not None:
        id = int(float(row[0].value))
        try:
            user = User.objects.get(pk=id)
            message = '{}, {}, {}, {} хэрэглэгч бүртгэлтэй. Бүртгэлтэй имэйл хаяг: {}'.format(
               user.id, user.username, user.first_name, user.last_name, user.email)
        except User.DoesNotExist:
            return 'Ийм ID-тай хэрэглэгч байхгүй байна.'
    else:
        try:
            reg_num = str(row[3].value).strip()
            m = UserMeta.objects.get(reg_num=reg_num)
            user = m.user
            message = '{}, {}, {}, {} хэрэглэгч бүртгэлтэй. Бүртгэлтэй имэйл хаяг: {}'.format(
                user.id, user.username, user.first_name, user.last_name, user.email)
        except UserMeta.MultipleObjectsReturned:
            reg_num = str(row[3].value).strip()
            m = UserMeta.objects.filter(reg_num=reg_num).first()
            user = m.user
            message = '{}, {}, {}, {} хэрэглэгч бүртгэлтэй. Бүртгэлтэй имэйл хаяг: {}'.format(
                user.id, user.username, user.first_name, user.last_name, user.email)
        except UserMeta.DoesNotExist:
            username = random_salt(8)
            if str(row[1].value) !='None' or str(row[2].value) !='None':
                user = User.objects.create(username=username,
                                           last_name=str(row[1].value),
                                           first_name=str(row[2].value),
                                           email=str(row[5].value),
                                           is_active=1)
                user.set_password(username)
                try:
                    user.username = 'u' + str(user.id)
                except:
                    user.username = 'u' + str(user.id)+'-'+random_salt(3)
                user.save()
                message = '{}, {}, {} хэрэглэгч шинээр бүртгүүллээ. Хэрэглэгчийн нэр {}, Нууц үг {}. Бүртгүүлсэн имэйл'.format(
                    user.id, user.last_name, user.first_name, user.username, username, user.email)
            else:
                return 'Мэдээлэл дутуу.'
        except Exception as e:
            return str(e)

    m, c = UserMeta.objects.get_or_create(user=user)
    if c:
        m.reg_num = str(row[3].value).upper()
        m.mobile = int(float(row[4].value))
        m.province_id = int(float(row[6].value))
        m.school = str(row[7].value)
        m.grade_id = int(float(row[8].value))
        m.level_id = level_id
    else:
        m.province_id = int(float(row[6].value))
        m.school = str(row[7].value)
        m.grade_id = int(float(row[8].value))
        m.level_id = level_id
    print(user.first_name, level_id)
    for item in row[:7]:
        print(item.value)
    m.save()

    if level_id == 2:
        group = Group.objects.get(pk=30)
    elif level_id == 3:
        group = Group.objects.get(pk=31)
    elif level_id == 4:
        group = Group.objects.get(pk=32)
    elif level_id == 5:
        group = Group.objects.get(pk=33)
    group.user_set.add(user)

    return message

def import_checked_users(wb,user):

    messages = list()
    try:
        sheet = wb["school"]
        teacher_id = int(float(sheet['B4'].value))
        teacher = User.objects.get(pk=teacher_id)
        group = Group.objects.get(pk=34)
        group.user_set.add(teacher)
        group.user_set.add(user)
    except Exception as e:
        messages.append(str(e))

    ind = 0
    messages.append('C (5-6)')
    sheet_c = wb['C (5-6)']
    for row in sheet_c:
        if ind > 0:
            message = import_row(row, 2)
            messages.append(message)
        ind = ind + 1

    ind = 0
    sheet_d = wb['D (7-8)']
    messages.append('D (7-8)')
    for row in sheet_d:
        if ind > 0:
            message = import_row(row,3)
            messages.append(message)
        ind = ind + 1

    ind = 0
    sheet_e = wb['E (9-10)']
    messages.append('E (9-10)')
    for row in sheet_e:
        if ind > 0:
            message = import_row(row,4)
            messages.append(message)
        ind = ind + 1

    ind = 0
    sheet_f = wb['F (11-12)']
    messages.append('F (11-12)')
    for row in sheet_f:
        if ind > 0:
            message = import_row(row,5)
            messages.append(message)
        ind = ind + 1

    return messages, teacher_id

@login_required()
def import_file(request):
    if "GET" == request.method:
        context = {'error': 0}
        return render(request, 'accounts/upload_file.html', context)
    else:
        error = False
        messages = list()

        excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(excel_file, data_only=True)

        error, messages = check_error(wb)

        if not error:
            messages, teacher_id = import_checked_users(wb,request.user)
            context = {'error': 1, 'messages': messages}
            text = '\n'.join(messages)
            text_html = '<html><meta charset="utf-8">{}<body></body></html>'
            html_content = '<p><strong>Бүртгэлийн мэдээлэл илгээж байна. Хэрвээ сэдэв хүлээн авах багшийн нэр <a href="{}">энд</a> байхгүй бол \
            mmo60official@gmail.com уруу мэдэгдээрэй.</strong></p>'.format('https://www.mmo.mn/accounts/users/34/')
            teacher = User.objects.filter(pk=teacher_id).first()
            for message in messages:
                html_content = html_content + "<p>{}</p>".format(message)
            text_html = text_html.format(html_content)
            email = UserMails.objects.create(subject = 'ММОХ бүртгэлийн мэдээлэл',
                    text = text,
                    text_html = text_html,
                    from_email = 'no-reply@mmo.mn',
                    to_email = teacher.email,
                    is_sent = False)
            if teacher_id != request.user.id:
                teacher_email = UserMails.objects.create(subject='ММОХ бүртгэлийн мэдээлэл',
                                 text=text,
                                 text_html=text_html,
                                 from_email='no-reply@mmo.mn',
                                 to_email=request.user.email,
                                 is_sent=False)
        else:
            context = {'error': error, 'messages': messages}

        return render(request, 'accounts/upload_file.html', context)

@login_required(login_url='/accounts/login/')
def send_email_with_attachments(request):
    if request.method == 'POST':
        form = EmailForm(request.POST, request.FILES)
        if form.is_valid():
            # group_name = 'my_group'  # Replace with your group name
            # group = Group.objects.get(pk=34)
            group = Group.objects.get(pk=35)
            users_in_group = group.user_set.all()

            subject = form.cleaned_data['subject']
            message = form.cleaned_data['message']
            from_email = 'MMO <no-reply@mmo.mn>'

            # Batch recipients and send emails
            recipients = [user.email for user in users_in_group]
            batch_size = 50  # Adjust as needed
            for i in range(0, len(recipients), batch_size):
                batch_recipients = recipients[i:i+batch_size]

                email = EmailMessage(subject, message, from_email, batch_recipients)
                email.content_subtype = "html"

                # Attach files
                attachments = request.FILES.getlist('attachments')
                for attachment in attachments:
                    attachment.seek(0)
                    email.attach(attachment.name, attachment.read(), attachment.content_type)

                # Send the email
                email.send()

            return render(request, 'accounts/success.html', {'message': 'Email sent successfully.'})

    else:
        form = EmailForm()

    return render(request, 'accounts/send_email.html', {'form': form})
