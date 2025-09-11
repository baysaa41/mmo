# accounts/services.py

import pandas as pd
import numpy as np
import string
import random
import re
import openpyxl

from django.db import transaction
from django.contrib.auth.models import User, Group
from django_pandas.io import read_frame

from .models import UserMeta, Province, Level, Grade

# --- Utility Functions ---
def random_salt(n=8):
    characterList = string.ascii_letters + string.digits
    salt = ['s']
    for i in range(n):
        randomchar = random.choice(characterList)
        salt.append(randomchar)
    return "".join(salt)

def check_email(email):
    regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
    if email and re.fullmatch(regex, email.strip()):
        return True
    else:
        return False

# --- Pandas Dataframe Service ---

def generate_styled_user_dataframe_html(users_queryset, is_staff=False):
    """
    Given a queryset of users, generates a styled HTML table using Pandas.
    This keeps the complex Pandas logic out of the view.
    """
    if is_staff:
        fieldnames = ['id', 'username', 'last_name', 'first_name', 'data__province__name',
                      'data__school', 'data__grade__name', 'data__reg_num', 'data__mobile', 'email']
        rename_map = {
            'id': 'ID', 'first_name': 'Нэр', 'last_name': 'Овог', 'username': 'Хэрэглэгчийн нэр',
            'data__province__name': 'Аймаг/Дүүрэг', 'data__school': 'Cургууль', 'data__grade__name': 'Анги',
            'data__reg_num': 'Регистрын дугаар', 'data__mobile': 'Гар утас', 'email': 'И-мэйл'
        }
    else:
        fieldnames = ['id', 'username', 'last_name', 'first_name', 'data__province__name',
                      'data__school', 'data__grade__name']
        rename_map = {
            'id': 'ID', 'first_name': 'Нэр', 'last_name': 'Овог', 'username': 'Хэрэглэгчийн нэр',
            'data__province__name': 'Аймаг/Дүүрэг', 'data__school': 'Cургууль', 'data__grade__name': 'Анги'
        }

    pd.options.display.float_format = '{:,.0f}'.format
    users_df = read_frame(users_queryset, fieldnames=fieldnames, verbose=False)
    if 'data__mobile' in users_df.columns:
        users_df['data__mobile'] = users_df['data__mobile'].astype(pd.Int64Dtype())

    users_df.rename(columns=rename_map, inplace=True)
    users_df.index = np.arange(1, len(users_df) + 1)

    styled_df = users_df.style.set_table_attributes('class="table table-bordered table-hover"').set_table_styles([
        {'selector': 'th', 'props': [('text-align', 'center')]},
        {'selector': 'td, th', 'props': [('border', '1px solid #ccc'), ('padding', '3px 5px')]},
    ])

    html_table = styled_df.to_html(na_rep="", escape=False)
    return re.sub('<th class="blank level0" >&nbsp;</th>', '<th class="blank level0" >№</th>', html_table)

# --- Excel Import Service ---

def process_excel_import(excel_file, user_who_uploaded):
    """
    Main service function to orchestrate the entire Excel import process.
    It validates the file and imports the data within a database transaction.
    """
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        is_invalid, validation_messages = _check_excel_file_errors(wb)

        if is_invalid:
            return False, validation_messages, None

        with transaction.atomic():
            # If validation passes, run the import within a transaction.
            # If any error occurs here, all database changes will be rolled back.
            import_messages, teacher = _import_users_from_workbook(wb, user_who_uploaded)
            return True, import_messages, teacher

    except Exception as e:
        # Catch any other unexpected errors during file processing.
        return False, [f"Файл боловсруулахад системийн алдаа гарлаа: {e}"], None


def _check_excel_file_errors(wb):
    """Internal function to validate the structure and content of the Excel workbook."""
    # This function combines the logic from the old `check_error` and `check_row` functions.
    # ... (The full logic of `check_error` and `check_row` goes here) ...
    # For brevity, I'm showing a simplified structure.
    messages = []
    has_error = False

    # Example check:
    if "school" not in wb.sheetnames:
        messages.append('school sheet байхгүй байна.')
        has_error = True

    # ... Add all other validation checks from your original `check_error` function ...

    if has_error:
        return True, messages
    return False, ["Файл шалгалтыг амжилттай давлаа."]


def _import_users_from_workbook(wb, user_who_uploaded):
    """
    Internal function that performs the actual user creation from the validated workbook.
    MUST be called within a transaction.
    """
    # This function combines the logic from `import_checked_users`, `import_row` etc.
    # ... (The full logic of those functions goes here) ...
    messages = []

    sheet = wb["school"]
    teacher_id = int(float(sheet['B4'].value))
    teacher = User.objects.get(pk=teacher_id)

    # Example of processing one sheet:
    sheet_c = wb['C (5-6)']
    messages.append('C (5-6) ангилал:')
    for index, row in enumerate(sheet_c.iter_rows(min_row=2)): # Skip header
        message = _import_single_row(row, level_id=2)
        messages.append(f"  Мөр {index+2}: {message}")

    # ... Process other sheets (D, E, F) similarly ...

    return messages, teacher


def _import_single_row(row, level_id):
    """Internal function to process a single row from an Excel sheet."""
    ## FIXME: АЮУЛГҮЙ БАЙДЛЫН ЭРСДЭЛ!
    ## Энэ хэсэгт хэрэглэгчийн нууц үгийг үүсгээд, энгийн текстээр буцааж,
    ## и-мэйлээр илгээж байна. Үүнийг яаралтай засаж, нууц үг сэргээх
    ## холбоос илгээдэг болгох шаардлагатай.

    # ... (The logic from your original `import_row` function goes here) ...
    # This is a simplified version for demonstration.
    try:
        if row[0].value: # Existing user by ID
            user = User.objects.get(pk=int(float(row[0].value)))
            # update UserMeta...
            return f"{user.username} хэрэглэгчийн мэдээллийг шинэчиллээ."
        else: # New user
            username = random_salt(8)
            user = User.objects.create_user(
                username=username,
                #... other fields
            )
            user.set_password(username) # <-- Security risk
            user.username = 'u' + str(user.id)
            user.save()
            # create UserMeta...
            return f"Шинэ хэрэглэгч {user.username} үүслээ. Нууц үг: {username}" # <-- Security risk
    except Exception as e:
        return f"АЛДАА: {e}"