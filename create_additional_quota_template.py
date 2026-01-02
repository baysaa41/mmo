#!/usr/bin/env python
"""
Нэмэлт эрх олгох тохиргооны Excel файл үүсгэх
Шинэ формат: Мэдээлэл хэсэг + Аймгууд хэсэг
"""
import os
import sys
import django
import argparse

# Django тохиргоо
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mmo.settings')
django.setup()

from accounts.models import Province
from olympiad.models import Olympiad, ScoreSheet
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Аргументууд
parser = argparse.ArgumentParser(description='Нэмэлт эрх олгох тохиргооны Excel файл үүсгэх')
parser.add_argument('--olympiad-ids', type=str, help='Олимпиадын ID-ууд (таслалаар: C:1,D:2,E:3,F:4)')
parser.add_argument('--categories', type=str, help='Ангиллалын нэрүүд (таслалаар: C,D,E,F)')
parser.add_argument('-o', '--output', type=str, default='additional_quota.xlsx', help='Гарах файлын нэр')
args = parser.parse_args()

# Ангиллал болон олимпиадын ID-г тодорхойлох
category_olympiad_map = {}

if args.olympiad_ids:
    # Формат: C:1,D:2,E:3,F:4
    for pair in args.olympiad_ids.split(','):
        if ':' in pair:
            cat, oid = pair.split(':')
            category_olympiad_map[cat.strip()] = int(oid.strip())
elif args.categories:
    # Зөвхөн ангиллалын нэр, олимпиадын ID хоосон
    for cat in args.categories.split(','):
        category_olympiad_map[cat.strip()] = None
else:
    # Үндсэн ангиллалууд
    category_olympiad_map = {'C': None, 'D': None, 'E': None, 'F': None}
    print('Үндсэн ангиллалуудыг ашиглаж байна: C, D, E, F')

categories = list(category_olympiad_map.keys())

# Бүх аймаг, дүүргийг татах
provinces = Province.objects.all().order_by('id')

# Excel workbook үүсгэх
wb = Workbook()
ws = wb.active
ws.title = "Нэмэлт эрх"

# === МЭДЭЭЛЭЛ ХЭСЭГ ===
row_idx = 1
ws.cell(row=row_idx, column=1, value='Мэдээлэл')
row_idx += 1

ws.cell(row=row_idx, column=1, value='Ангилал')
ws.cell(row=row_idx, column=2, value='Олимпиадын ID')
row_idx += 1

for category, olympiad_id in category_olympiad_map.items():
    ws.cell(row=row_idx, column=1, value=category)
    ws.cell(row=row_idx, column=2, value=olympiad_id if olympiad_id else '')
    row_idx += 1

# Хоосон мөр
row_idx += 1

# === АЙМГУУД ХЭСЭГ ===
ws.cell(row=row_idx, column=1, value='Аймгууд')
row_idx += 1

# Header
ws.cell(row=row_idx, column=1, value='ID')
ws.cell(row=row_idx, column=2, value='Нэр')
for col_idx, category in enumerate(categories, start=3):
    ws.cell(row=row_idx, column=col_idx, value=category)
row_idx += 1

# Аймаг бүрийн мөр
for province in provinces:
    ws.cell(row=row_idx, column=1, value=province.id)
    ws.cell(row=row_idx, column=2, value=province.name)
    # Ангиллалын багануудыг хоосон үлдээнэ
    row_idx += 1

# Файл хадгалах
wb.save(args.output)

print(f'\n✓ {args.output} файл амжилттай үүсгэгдлээ')
print(f'  Нийт {provinces.count()} аймаг/дүүрэг')
print(f'  Ангиллалын багана: {", ".join(categories)}')
print(f'\nФайлын бүтэц:')
print(f'  1. Мэдээлэл хэсэг: Ангилал бүрийн олимпиадын ID')
print(f'  2. Аймгууд хэсэг: Аймаг бүрийн босго оноо')
print(f'\nЗааварчилгаа:')
print(f'  - Мэдээлэл хэсэгт ангилал бүрийн олимпиадын ID-г оруулна')
print(f'  - Аймгууд хэсэгт ангиллалын баганад босго оноог оруулна')
print(f'  - Хоосон нүд = нэмэлт эрх олгохгүй')
