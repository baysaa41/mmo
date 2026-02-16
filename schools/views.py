from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Count, Q
from django.utils import timezone

from .email_service import SchoolEmailService
import random
import string
from django.urls import reverse

from django.utils.crypto import get_random_string

import numpy as np
import logging
# Logger үүсгэх
logger = logging.getLogger(__name__)

# Шаардлагатай import-ууд
import io
import re
from django.http import HttpResponse
from datetime import date, timedelta
from openpyxl.styles import Border, Side, Font, Alignment

from .models import School
from .forms import UserSearchForm, AddUserForm, UserForm, UserMetaForm, UploadExcelForm
from accounts.models import UserMeta, Level, Province
from olympiad.models import Olympiad, SchoolYear, Problem, Result

import pandas as pd
from django.db import transaction
from .forms import UploadExcelForm

from django.contrib.admin.views.decorators import staff_member_required
from .forms import SchoolModeratorChangeForm, EditSchoolInfoForm

@login_required
def my_managed_schools_view(request):
    """
    Миний удардаж байгаа сургуулиуд - Shows schools where user is moderator or manager
    """
    schools_moderating = request.user.moderating.all()
    schools_managing = request.user.managing.all()
    my_schools = (schools_moderating | schools_managing).select_related('user', 'manager', 'group', 'province').distinct().order_by('province__name', 'name')

    context = {
        'my_schools': my_schools,
    }
    return render(request, 'schools/my_managed_schools.html', context)


@login_required
def all_schools_registry_view(request):
    """
    Сургуулийн бүртгэл - Shows all schools in the system with search/filter
    Staff болон аймгийн contact/moderator нар хандах эрхтэй.
    """
    # Эрх шалгах
    is_staff = request.user.is_staff
    pid = request.GET.get('p')
    zid = request.GET.get('z')
    name_query = request.GET.get('q', '')

    # Аймгийн manager эсэхийг шалгах
    managed_province = None
    if pid:
        try:
            province = Province.objects.get(id=pid)
            # Аймгийн contact эсвэл Province_{id}_Managers group-д байгаа эсэхийг шалгах
            if province.contact_person == request.user or request.user.groups.filter(name=f"Province_{province.id}_Managers").exists():
                managed_province = province
            elif not is_staff:
                messages.error(request, 'Та энэ аймгийн сургуулиудыг удирдах эрхгүй.')
                return redirect('my_managed_schools')
        except Province.DoesNotExist:
            if not is_staff:
                messages.error(request, 'Аймаг олдсонгүй.')
                return redirect('my_managed_schools')
    elif not is_staff:
        # Province ID байхгүй бөгөөд staff биш бол эрхгүй
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_schools')

    schools_qs = School.objects.select_related('user', 'manager', 'group', 'province')

    # Аймгийн manager бол зөвхөн өөрийн аймгийн сургуулиудыг харуулах
    if managed_province:
        schools_qs = schools_qs.filter(province=managed_province)

    # Apply filters
    if name_query:
        schools_qs = schools_qs.filter(
            Q(name__icontains=name_query) | Q(alias__icontains=name_query)
        )
    # Province шүүлтүүр - зөвхөн staff хэрэглэгчдэд
    if pid and is_staff and not managed_province:
        schools_qs = schools_qs.filter(province_id=pid)
    elif zid and is_staff and not managed_province:
        schools_qs = schools_qs.filter(province__zone_id=zid)

    schools = schools_qs.order_by('province__name', 'name')

    context = {
        'schools': schools,
        'search_query': name_query,
        'managed_province': managed_province,
        'is_staff': is_staff,
    }
    return render(request, 'schools/all_schools_registry.html', context)


@login_required
def school_moderators_view(request):
    """
    DEPRECATED: This view is kept for backwards compatibility
    Redirects to my_managed_schools_view
    """
    return redirect('my_managed_schools')

@login_required
def school_dashboard(request, school_id):
    """
    Сургуулийн удирдлагын самбар. Сурагчдын тоо болон боломжит
    олимпиадуудын жагсаалтыг харуулна.
    """
    school = get_object_or_404(School, id=school_id)
    if not school.user_has_access(request.user):
        messages.error(request, 'Та энэ сургуулийг удирдах эрхгүй.')
        return render(request, 'error.html', {'message': 'Хандах эрхгүй.'})

    # Check if current user is manager
    is_manager = (school.manager == request.user) or request.user.is_staff

    # Сурагчдын тоог ангиллаар гаргах
    student_counts = Level.objects.annotate(
        student_count=Count('usermeta__user', filter=Q(usermeta__user__groups=school.group))
    ).order_by('name')

    uncategorized_count = school.group.user_set.filter(data__level__isnull=True).count()

    # Одоогийн хичээлийн жилийн, 1-р давааны тестүүдийг шүүх
    today = date.today()
    current_school_year = SchoolYear.objects.filter(start__lte=today, end__gte=today).first()

    olympiads = Olympiad.objects.none()
    if current_school_year:
        olympiads = Olympiad.objects.filter(
            round=1,
            school_year=current_school_year
        ).order_by('level__name')

    # --- ШИНЭЭР НЭМЭГДСЭН ЛОГИК ---
    # Тухайн сургуулийг сонгосон боловч группт нь ороогүй сурагчдыг олох
    pending_students = User.objects.filter(
        data__school=school
    ).exclude(
        groups=school.group
    )

    context = {
        'school': school,
        'is_manager': is_manager,  # <-- manager эсэхийг context-д нэмэх
        'student_counts_by_level': student_counts,
        'uncategorized_count': uncategorized_count,
        'pending_students': pending_students,
        'olympiads': olympiads,
    }
    return render(request, 'schools/school_dashboard.html', context)

@login_required
def manage_school_by_level(request, school_id, level_id):
    """
    Сонгогдсон нэг ангиллын (эсвэл ангилалгүй) сурагчдыг удирдах хуудас.
    """
    school = get_object_or_404(School, id=school_id)
    if not school.user_has_access(request.user):
        messages.error(request, 'Та энэ сургуулийг удирдах эрхгүй.')
        return render(request, 'error.html', {'message': 'Хандах эрхгүй.'})

    group = school.group

    # --- ШИНЭЧИЛСЭН ЛОГИК ---
    if level_id == 100: # 100 нь "Бүх сурагчид" гэсэн утгатай
        selected_level = {'id': 100, 'name': 'Бүх сурагчид'}
        users_in_level = group.user_set.all().select_related('data__grade')
        # Сургуульд бүртгэлгүй боловч өөрийгөө тухайн сургууль гэж бүртгүүлсэн сурагчид
        pending_users = User.objects.filter(
            data__school=school
        ).exclude(
            groups=group
        ).select_related('data__grade')
        # Нэг жагсаалт болгож нэгтгэх
        registered_ids = set(group.user_set.values_list('id', flat=True))
        all_school_users = []
        for u in users_in_level:
            all_school_users.append({'user': u, 'is_registered': True})
        for u in pending_users:
            all_school_users.append({'user': u, 'is_registered': False})
        all_school_users.sort(key=lambda x: (x['user'].last_name or '', x['user'].first_name or ''))
    elif level_id == 0:
        selected_level = {'id': 0, 'name': 'Ангилалгүй'}
        users_in_level = group.user_set.filter(data__level__isnull=True).select_related('data__grade')
    else:
        selected_level = get_object_or_404(Level, id=level_id)
        users_in_level = group.user_set.filter(data__level=selected_level).select_related('data__grade')

    search_results = None
    if request.method == 'POST':
        if 'search_users' in request.POST:
            search_form = UserSearchForm(request.POST)
            add_user_form = AddUserForm()
            if search_form.is_valid():
                search_results = search_form.search_users()

        elif 'add_user' in request.POST:
            add_user_form = AddUserForm(request.POST)
            search_form = UserSearchForm()
            if add_user_form.is_valid():
                new_user = None
                try:
                    with transaction.atomic():
                        new_user = add_user_form.save(commit=False)
                        new_user.username = ''.join(random.choice(string.ascii_letters) for _ in range(32))
                        # temp_password = User.objects.make_random_password(length=20)
                        temp_password = get_random_string(20)
                        new_user.set_password(temp_password)
                        new_user.save()
                        new_user.username = f'u{new_user.id}'
                        new_user.save()

                        meta_data = {
                            'user': new_user,
                            'school': request.user.data.school,
                            'province': request.user.data.province,
                        }
                        if level_id != 0:
                            meta_data['level'] = selected_level

                        if level_id == 100:
                            selected_level = Level.objects.get(pk=8)
                            meta_data['level'] = selected_level

                        UserMeta.objects.create(**meta_data)
                        new_user.groups.add(group)

                    # Тавтай морилно уу имэйл + password reset link илгээх
                    school_name = request.user.data.school.name if request.user.data.school else "ММОХ"
                    success, error = SchoolEmailService.send_new_user_welcome_with_reset_link(
                        new_user,
                        school_name,
                        request
                    )

                    level_name = selected_level['name'] if level_id == 0 else selected_level.name

                    if success:
                        messages.success(
                            request,
                            f"'{new_user.get_full_name()}' хэрэглэгчийг '{level_name}' хэсэгт "
                            f"амжилттай нэмж, нэвтрэх мэдээллийг и-мэйлээр илгээлээ."
                        )
                    else:
                        messages.warning(
                            request,
                            f"Хэрэглэгч үүссэн ч и-мэйл илгээхэд алдаа гарлаа. "
                            f"Та дахин илгээх эсвэл password reset хийлгэнэ үү."
                        )

                except Exception as e:
                    messages.error(request, f"Хэрэглэгч үүсгэхэд алдаа гарлаа: {e}")
                    logger.error(f"User creation failed: {e}", exc_info=True)

                return redirect('manage_school_by_level', school_id=school_id, level_id=level_id)

        elif 'add_existing_user' in request.POST:
            search_form = UserSearchForm()
            add_user_form = AddUserForm()
            user_id = request.POST.get('user_id')
            user_to_add = get_object_or_404(User, id=user_id)

            user_meta = getattr(user_to_add, "data", None)

            if not user_meta:
                messages.error(request, "Энэ хэрэглэгчид профайл (UserMeta) бүртгэл алга байна.")
                return redirect('manage_school_by_level', school_id=school_id, level_id=level_id)

            # зөвхөн staff л staff-ийг нэмэх эрхтэй
            if not request.user.is_staff and user_to_add.is_staff:
                messages.error(request, "Танд энэ хэрэглэгчийг нэмэх эрх байхгүй.")
                return redirect('manage_school_by_level', school_id=school_id, level_id=level_id)

            # админыг нэмж болохгүй
            if user_to_add.is_superuser:
                messages.error(request, "Танд супер хэрэглэгчийг нэмэх эрх байхгүй.")
                return redirect('manage_school_by_level', school_id=school_id, level_id=level_id)

            # Өөр сургуулийн бүртгэгч багш, менежерийг нэмж болохгүй
            if not request.user.is_staff:
                # Бүртгэгч багш эсэхийг шалгах (өөр сургуулийн)
                other_moderating = user_to_add.moderating.exclude(pk=school.pk)
                if other_moderating.exists():
                    other_school = other_moderating.first()
                    messages.error(
                        request,
                        f"'{user_to_add.get_full_name()}' хэрэглэгч '{other_school.name}' сургуулийн бүртгэгч багш тул нэмэх боломжгүй."
                    )
                    return redirect('manage_school_by_level', school_id=school_id, level_id=level_id)

                # Менежер эсэхийг шалгах (өөр сургуулийн)
                other_managing = user_to_add.managing.exclude(pk=school.pk)
                if other_managing.exists():
                    other_school = other_managing.first()
                    messages.error(
                        request,
                        f"'{user_to_add.get_full_name()}' хэрэглэгч '{other_school.name}' сургуулийн менежер тул нэмэх боломжгүй."
                    )
                    return redirect('manage_school_by_level', school_id=school_id, level_id=level_id)

            is_in_old_group = False
            if user_meta.school and user_meta.school.group:
                is_in_old_group = user_to_add.groups.filter(pk=user_meta.school.group.pk).exists()

            # Хуучин сургууль нь өөр байвал зөвхөн staff эрхтэй хүн сольж чадна
            if user_meta.school and user_meta.school != school and is_in_old_group and not request.user.is_staff:
                messages.error(
                    request,
                    f"'{user_to_add.get_full_name()}' хэрэглэгч '{user_meta.school.name}' сургуульд бүртгэлтэй тул нэмэх боломжгүй. "
                    f"Зөвхөн системийн админ шилжүүлэг хийж чадна."
                )
            else:
                # Хуучин сургуульд байсан group-оос хасах (staff үед)
                if request.user.is_staff and user_meta.school and user_meta.school != school:
                    old_group = user_meta.school.group
                    if old_group:
                        old_group.user_set.remove(user_to_add)

                # Шинэ сургуульд оноох
                user_meta.school = school
                # Сургуулийн дүүргийн мэдээллийг хэрэглэгчийн дүүрэг болгож оноох
                if school.province:
                    user_meta.province = school.province
                if level_id != 0:
                    user_meta.level = selected_level if isinstance(selected_level, Level) else None
                user_meta.save()

                group.user_set.add(user_to_add)
                messages.success(
                    request,
                    f"'{user_to_add.get_full_name() or user_to_add.username}' хэрэглэгчийг '{school.name}' сургуульд амжилттай нэмлээ."
                )

            if user_to_add.data.level_id != level_id:
                level_id = 100

            return redirect('manage_school_by_level', school_id=school_id, level_id=level_id)

        elif 'approve_user' in request.POST:
            search_form = UserSearchForm()
            add_user_form = AddUserForm()
            user_id = request.POST.get('user_id')
            user_to_approve = get_object_or_404(User, id=user_id)
            if hasattr(user_to_approve, 'data') and user_to_approve.data.school == school:
                group.user_set.add(user_to_approve)
                messages.success(request, f"'{user_to_approve.last_name} {user_to_approve.first_name}' бүртгэл нийлүүлэгдлээ.")
            return redirect('manage_school_by_level', school_id=school_id, level_id=level_id)

        elif 'approve_all_users' in request.POST:
            search_form = UserSearchForm()
            add_user_form = AddUserForm()
            pending = User.objects.filter(data__school=school).exclude(groups=group)
            count = pending.count()
            for u in pending:
                group.user_set.add(u)
            messages.success(request, f"{count} сурагчийн бүртгэл нийлүүлэгдлээ.")
            return redirect('manage_school_by_level', school_id=school_id, level_id=level_id)

        elif 'remove_user' in request.POST:
            search_form = UserSearchForm()
            add_user_form = AddUserForm()
            user_id = request.POST.get('user_id')
            user_to_remove = get_object_or_404(User, id=user_id)
            group.user_set.remove(user_to_remove)
            messages.info(request, f"'{user_to_remove.get_full_name() or user_to_remove.username}' хэрэглэгчийг сургуулиас хаслаа.")
            return redirect('manage_school_by_level', school_id=school_id, level_id=level_id)

    else: # GET request
        search_form = UserSearchForm()
        add_user_form = AddUserForm()

    context = {
        'school': school,
        'group': group,
        'selected_level': selected_level,
        'users_in_level': users_in_level,
        'search_form': search_form,
        'add_user_form': add_user_form,
        'search_results': search_results,
        'all_school_users': all_school_users if level_id == 100 else None,
    }
    return render(request, 'schools/manage_school_users_level.html', context)

@login_required
def school_olympiad_view(request, school_id, olympiad_id):
    school = get_object_or_404(School, id=school_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    selected_level = olympiad.level if olympiad.level else {'id': 0, 'name': 'Ангилалгүй'}

    upload_form = UploadExcelForm()
    context = {
        'school': school,
        'olympiad': olympiad,
        'selected_level': selected_level,
        'upload_form': upload_form,
    }
    return render(request, 'schools/school_level_olympiad_detail.html', context)


@login_required
def generate_school_answer_sheet(request, school_id, olympiad_id):
    """Сонгосон сургууль, олимпиадын хариултын хуудсыг үүсгэж, шууд download хийлгэнэ."""
    school = get_object_or_404(School, pk=school_id)
    olympiad = get_object_or_404(Olympiad, pk=olympiad_id)

    # Эрхийн шалгалт
    if not school.user_has_access(request.user):
        messages.error(request, 'Та энэ үйлдлийг хийх эрхгүй.')
        return redirect('school_dashboard', school_id=school_id)

    # Дата бэлдэх
    level = olympiad.level
    contestants = User.objects.filter(groups=school.group, data__level=level).order_by('last_name', 'first_name')
    problems = olympiad.problem_set.all().order_by('order')

    # 1-р Sheet: Хариулт
    answers_data = []
    for c in contestants:
        answers_data.append({
            'ID': c.id,
            'Овог': c.last_name,
            'Нэр': c.first_name,
            **{f'№{p.order}': '' for p in problems}
        })
    header = ['ID', 'Овог', 'Нэр'] + [f'№{p.order}' for p in problems]
    answers_df = pd.DataFrame(answers_data, columns=header)

    # 2-р Sheet: Мэдээлэл
    info_data = {
        'Түлхүүр': ['olympiad_id', 'olympiad_name', 'school_id', 'school_name', 'level_id', 'level_name'],
        'Утга': [olympiad.id, olympiad.name, school.id, school.name, level.id, level.name]
    }
    info_df = pd.DataFrame(info_data)

    # Excel файлыг санах ойд үүсгэх
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        answers_df.to_excel(writer, sheet_name='Хариулт', index=False)
        info_df.to_excel(writer, sheet_name='Мэдээлэл', index=False)

        # Style-г тохируулах workbook болон worksheet-г авах
        workbook = writer.book
        worksheet = writer.sheets['Хариулт']

        # Style-уудыг тодорхойлох
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')

        # Header буюу эхний мөрийг загварчлах
        for cell in worksheet["1:1"]:
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_align

        # Үндсэн өгөгдлийн нүднүүдэд хүрээ нэмэх
        for row in worksheet.iter_rows(min_row=2,
                                       max_row=worksheet.max_row,
                                       max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border

        # Баганын өргөнийг тохируулах
        for i, column_cells in enumerate(worksheet.columns):
            column_letter = column_cells[0].column_letter
            if i < 3: # ID, Овог, Нэр баганууд
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_letter].width = length + 2
            else: # Хариултын баганууд
                worksheet.column_dimensions[column_letter].width = 5

    output.seek(0)

    # Хэрэглэгчид файл болгож буцаах
    filename = f"answer_sheet_{olympiad.id}_{school.id}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# schools/views.py файлын дээд хэсэгт "time" модулийг импортлоно
import time
# ... (бусад import-ууд хэвээрээ) ...

@login_required
def import_school_answer_sheet(request, school_id, olympiad_id):
    """Excel-ээс бөглөсөн хариултын хуудсыг уншиж, Result руу импортлоно."""
    school = get_object_or_404(School, pk=school_id)
    olympiad = get_object_or_404(Olympiad, pk=olympiad_id)
    level_id = olympiad.level_id

    # Эрхийн шалгалт
    if not school.user_has_access(request.user):
        messages.error(request, "Та энэ үйлдлийг хийх эрхгүй.")
        return redirect('school_dashboard', school_id=school_id)

    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_excel = form.save(commit=False)
            uploaded_excel.uploaded_by = request.user
            uploaded_excel.save()
            excel_file = uploaded_excel.file.path

            # --- Хугацаа хэмжиж эхлэх ---
            start_time = time.time()

            try:
                # --- МЭДЭЭЛЭЛ sheet унших ---
                df_info = pd.read_excel(excel_file, sheet_name='Мэдээлэл')
                info_dict = pd.Series(df_info.Утга.values, index=df_info.Түлхүүр).to_dict()

                # Зөв сургууль, зөв олимпиад эсэхийг шалгах
                if int(info_dict.get('olympiad_id')) != olympiad.id or int(info_dict.get('school_id')) != school.id:
                    messages.error(request, "Файл өөр олимпиад/сургуульд зориулагдсан байна.")
                    return redirect('school_olympiad_view', school_id=school_id, olympiad_id=olympiad_id)

                # --- Хариулт sheet унших ---
                df = pd.read_excel(excel_file, sheet_name='Хариулт')
                problems_map = {p.order: p for p in Problem.objects.filter(olympiad=olympiad)}

                if 'ID' not in df.columns:
                    messages.error(request, "Excel-ийн 'Хариулт' sheet-д 'ID' багана байхгүй байна.")
                    return redirect('school_dashboard', school_id=school_id)

                updated_count = created_count = skipped_rows = invalid_format_count = 0

                for index, row in df.iterrows():
                    user_id = row.get('ID')
                    if pd.isna(user_id):
                        skipped_rows += 1
                        continue

                    try:
                        user = User.objects.get(pk=int(user_id))
                    except (User.DoesNotExist, ValueError, TypeError):
                        skipped_rows += 1
                        continue

                    # Сурагч зөв сургуулийн group-д багтсан эсэх
                    if not user.groups.filter(pk=school.group.id).exists():
                        skipped_rows += 1
                        continue

                    with transaction.atomic():
                        for order, problem in problems_map.items():
                            col = f'№{order}'
                            if col not in df.columns:
                                continue

                            answer = row[col]
                            db_value = None
                            valid = False

                            if pd.notna(answer) and str(answer).strip() != '':
                                try:
                                    fa = float(answer)
                                    if fa > 0 and fa.is_integer():
                                        db_value = int(fa)
                                        valid = True
                                except (ValueError, TypeError):
                                    pass

                            if not valid and pd.notna(answer) and str(answer).strip() != '':
                                invalid_format_count += 1

                            obj, created = Result.objects.update_or_create(
                                contestant=user, olympiad=olympiad, problem=problem,
                                defaults={'answer': db_value}
                            )
                            if created:
                                created_count += 1
                            else:
                                updated_count += 1


                # --- Хугацаа хэмжиж дуусгах ---
                elapsed_time = round(time.time() - start_time, 2)

                messages.success(
                    request,
                    f"Excel боловсрууллаа ({elapsed_time} секунд). "
                    f"Үүссэн: {created_count}, "
                    f"Шинэчлэгдсэн: {updated_count}, "
                    f"Буруу форматтай: {invalid_format_count}, "
                    f"Алгассан: {skipped_rows}."
                )

            except Exception as e:
                messages.error(request, f"Файл боловсруулахад алдаа: {e}")
                return redirect('school_dashboard', school_id=school_id)

            # Амжилтын дараа дүнгийн хуудас руу шилжих
            target_url = reverse('olympiad_answer_view', args=[olympiad_id])
            final_url = f"{target_url}?s={school_id}&p={school.province.id}"
            return redirect(final_url)

        else:
            messages.error(request, "Файл хуулахад алдаа гарлаа.")

    return redirect('school_dashboard', school_id=school_id)


@login_required
def view_school_olympiad_results(request, school_id, olympiad_id):
    """
    Сонгосон сургууль, олимпиадын хувьд сурагчдын авсан оноог харуулна.
    Энэ хувилбарт аймгийн шүүлтүүр хасагдаж, нийт онооны багана нэмэгдсэн.
    """
    sid = school_id
    context_data = ''

    try:
        # Олимпиад болон сургуулийн мэдээллийг татах
        olympiad = Olympiad.objects.get(pk=olympiad_id)
        school = School.objects.get(pk=sid)
    except (Olympiad.DoesNotExist, School.DoesNotExist):
        return render(request, 'olympiad/results/no_olympiad.html', {'message': 'Олимпиад эсвэл сургууль олдсонгүй.'})

    # Тухайн олимпиад, сургуульд хамаарах бүх дүнг татах
    results = Result.objects.filter(olympiad_id=olympiad_id, contestant__data__school_id=sid)

    if results.exists():
        # Дүнгүүдээс pandas DataFrame үүсгэх (хариултын оронд оноог ашиглана)
        rows = list(results.values_list('contestant_id', 'problem_id', 'score'))
        data = pd.DataFrame(rows, columns=['contestant_id', 'problem_id', 'score'])

        # Хүснэгтийг эргүүлж, сурагчдыг мөр, бодлогуудыг багана болгох
        # Хэрэв тухайн бодлогод оноо аваагүй бол 0 гэж тооцно
        results_df = pd.pivot_table(data, index='contestant_id', columns='problem_id', values='score', aggfunc='sum', fill_value=0)

        # Баганын нэрийг бодлогын дугаараар солих
        problem_ids = results_df.columns.values
        problem_orders = {p.id: f'№{p.order}' for p in Problem.objects.filter(id__in=problem_ids)}
        results_df.columns = [problem_orders.get(col, 'Unknown') for col in results_df.columns]

        # "Нийт" онооны багана нэмэх
        results_df['Нийт'] = results_df.sum(axis=1)

        # Багануудыг эрэмбэлэх (Бодлогууд -> Нийт)
        problem_cols = sorted([col for col in results_df.columns if col != 'Нийт'])
        results_df = results_df[problem_cols + ['Нийт']]

        # Сурагчдын мэдээллийг авах
        contestant_ids = list(results_df.index)
        contestants_data = User.objects.filter(pk__in=contestant_ids).values('id', 'last_name', 'first_name')
        user_df = pd.DataFrame(list(contestants_data))
        user_df.columns = ['ID', 'Овог', 'Нэр']

        # Сурагчдын мэдээллийг онооны хүснэгттэй нэгтгэх
        user_results_df = pd.merge(user_df, results_df, left_on='ID', right_index=True, how='left')

        # Нийт оноогоор буурахаар, дараа нь нэрээр өсөхөөр эрэмбэлэх
        sorted_df = user_results_df.sort_values(
            by=['Нийт', 'Овог', 'Нэр'],
            ascending=[False, True, True]
        ).drop(columns=['ID'])

        # Индексийг 1-ээс эхлүүлэх
        sorted_df.index = np.arange(1, len(sorted_df) + 1)

        # Онооны багануудыг тодорхойлох
        numeric_columns = [col for col in sorted_df.columns if str(col).startswith('№') or str(col) == 'Нийт']

        # Оноог нэг орны нарийвчлалтай форматлах
        formatter = '{:.1f}'.format

        # HTML гаргалт руу хөрвүүлэхдээ style тохируулах
        styled_df = (sorted_df.style
                              .format(formatter, subset=numeric_columns, na_rep="-")
                              .set_table_attributes('class="table table-bordered table-hover"'))
        context_data = styled_df.to_html()

    # Template-д дамжуулах мэдээллийг бэлтгэх
    title = f"{school.name} - {olympiad.name}"
    name = f"{olympiad.name}, {olympiad.level.name} ангилал"

    context = {
        'title': f"{title} - Дүн",
        'name': name,
        'data': context_data,
        'school': school,
        'selected_sid': sid,
        'olympiad_id': olympiad_id,
    }

    return render(request, 'olympiad/results/school_results.html', context)

@login_required
def edit_user_in_group(request, user_id):
    """
    Handles editing a student's profile information by a school moderator.
    """
    target_user = get_object_or_404(User, id=user_id)

    if request.user.is_staff:
        school = School.objects.filter(group__user=target_user).first()
    else:
        school = School.objects.filter(group__user=target_user, user=request.user).first()
        if not school:
            messages.error(request, 'Та энэ хэрэглэгчийг засах эрхгүй.')
            return render(request, 'error.html', {'message': 'Хандах эрхгүй.'})

    user_meta = get_object_or_404(UserMeta, user=target_user)

    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=target_user)
        user_meta_form = UserMetaForm(request.POST, request.FILES, instance=user_meta)

        if user_form.is_valid() and user_meta_form.is_valid():
            user_form.save()
            user_meta_form.save()
            messages.success(request, 'Профайлыг амжилттай шинэчиллээ!')
            return redirect('edit_user_in_group', user_id=user_id)
        else:
            messages.error(request, 'Доорх алдааг засна уу.')
    else:
        user_form = UserForm(instance=target_user)
        user_meta_form = UserMetaForm(instance=user_meta)

    context = {
        'user_form': user_form,
        'user_meta_form': user_meta_form,
        'target_user': target_user,
        'school': school,
        'level': user_meta.level,
    }
    return render(request, 'schools/edit_user_in_group.html', context)

@login_required
def edit_profile(request):
    """This view is for users editing their own profile."""
    user = request.user
    user_meta, created = UserMeta.objects.get_or_create(user=user)

    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=user)
        user_meta_form = UserMetaForm(request.POST, request.FILES, instance=user_meta)
        if user_form.is_valid() and user_meta_form.is_valid():
            user_form.save()
            user_meta_form.save()
            messages.success(request, 'Таны мэдээлэл амжилттай шинэчлэгдлээ.')
            return redirect('edit_profile')
    else:
        user_form = UserForm(instance=user)
        user_meta_form = UserMetaForm(instance=user_meta)

    context = {
        'user_form': user_form,
        'user_meta_form': user_meta_form,
    }
    return render(request, 'schools/edit_profile.html', context)

@login_required
def add_student_to_group_view(request, school_id, user_id):
    school = get_object_or_404(School, id=school_id)
    student = get_object_or_404(User, id=user_id)

    # Эрхийн шалгалт
    if not school.user_has_access(request.user):
        messages.error(request, 'Та энэ үйлдлийг хийх эрхгүй.')
        return redirect('school_dashboard', school_id=school.id)

    # Сурагчийг сургуулийн группт нэмэх
    if school.group:
        # Шалгалт: Сурагч үнэхээр энэ сургуулийг сонгосон эсэх
        if student.data.school == school:
            school.group.user_set.add(student)
            messages.success(request, f"'{student.get_full_name()}' сурагчийг сургуулийн бүлэгт амжилттай нэмлээ.")
        else:
            messages.warning(request, f"'{student.get_full_name()}' сурагч энэ сургуулийг сонгоогүй байна.")
    else:
        messages.error(request, "Энэ сургуульд групп оноогоогүй тул сурагч нэмэх боломжгүй.")

    return redirect('school_dashboard', school_id=school.id)

# ... (import-ууд) ...
from .forms import SchoolAdminPasswordChangeForm

@login_required
def change_student_password_view(request, user_id):
    """
    Сургуулийн модератор нь сурагчийн нууц үгийг солих хуудас.
    """
    target_user = get_object_or_404(User, id=user_id)
    moderator = request.user

    # --- АЮУЛГҮЙ БАЙДЛЫН ШАЛГАЛТУУД ---
    # 1. Засварлах гэж буй хэрэглэгч нь staff/superuser биш байх ёстой.
    if (target_user.is_staff or target_user.is_superuser) and not request.user.is_superuser:
        messages.error(request, "Та staff эрхтэй хэрэглэгчийн нууц үгийг солих боломжгүй.")
        return redirect('school_dashboard', school_id=moderator.data.school.id) # Өөрийнх нь dashboard руу буцаах

    # 2. Модератор нь тухайн сурагчийн сургуульд хамааралтай эсэхийг шалгах.
    try:
        # Сурагчийн сургууль
        student_school = target_user.data.school
        if not student_school:
            student_school = moderator.data.school
        # Модератор/менежерын удирддаг сургуулиудыг олох
        moderator_schools = School.objects.filter(Q(user=moderator) | Q(manager=moderator))

        if (not student_school or not moderator_schools.filter(pk=student_school.pk).exists()) and not request.user.is_staff:
            messages.error(request, "Та энэ сурагчийн нууц үгийг солих эрхгүй.")
            return redirect('school_dashboard', school_id=moderator.data.school.id)
    except UserMeta.DoesNotExist:
        messages.error(request, "Сурагчийн профайлын мэдээлэл олдсонгүй.")
        return redirect('school_dashboard', school_id=moderator.data.school.id)


    if request.method == 'POST':
        form = SchoolAdminPasswordChangeForm(user=target_user, data=request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f"'{target_user.get_full_name()}' хэрэглэгчийн нууц үгийг амжилттай солилоо.")
            # Буцах замыг зөв тодорхойлох
            return redirect('school_all_users', school_id=student_school.id)
    else:
        form = SchoolAdminPasswordChangeForm(user=target_user)

    context = {
        'form': form,
        'target_user': target_user,
    }
    return render(request, 'schools/change_password.html', context)


@login_required
def manage_all_schools_view(request):
    """
    Бүх сургуулийн жагсаалтыг дэлгэрэнгүй мэдээлэлтэй харуулж,
    модератор солих үйлдэл хийдэг хуудас.
    Staff болон аймгийн contact/moderator нар хандах эрхтэй.
    """
    # Эрх шалгах
    is_staff = request.user.is_staff
    province_id = request.GET.get('p', '')

    # Аймгийн manager эсэхийг шалгах
    managed_province = None
    if province_id:
        try:
            province = Province.objects.get(id=province_id)
            # Аймгийн contact эсвэл Province_{id}_Managers group-д байгаа эсэхийг шалгах
            if province.contact_person == request.user or request.user.groups.filter(name=f"Province_{province.id}_Managers").exists():
                managed_province = province
            elif not is_staff:
                messages.error(request, 'Та энэ аймгийн сургуулиудыг удирдах эрхгүй.')
                return redirect('my_managed_schools')
        except Province.DoesNotExist:
            messages.error(request, 'Аймаг олдсонгүй.')
            return redirect('my_managed_schools')
    elif not is_staff:
        # Province ID байхгүй бөгөөд staff биш бол эрхгүй
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_schools')

    if request.method == 'POST':
        school_id = request.POST.get('school_id')
        school_to_change = get_object_or_404(School, id=school_id)

        # Province manager бол зөвхөн өөрийн аймгийн сургуулийг засах эрхтэй
        if managed_province and school_to_change.province != managed_province:
            messages.error(request, 'Та энэ сургуулийг засах эрхгүй.')
            return redirect('manage_all_schools')

        form = SchoolModeratorChangeForm(request.POST)

        if form.is_valid():
            new_moderator = form.cleaned_data['user']
            school_to_change.user = new_moderator
            school_to_change.save()
            messages.success(request, f"'{school_to_change.name}' сургуулийн модераторыг амжилттай солилоо.")
        else:
            messages.error(request, "Модератор солиход алдаа гарлаа.")

        return redirect('manage_all_schools')

    # Сургуулиудын жагсаалтыг бүх мэдээлэлтэй нь авах
    all_schools = School.objects.select_related(
        'province', 'user', 'user__data'
    ).annotate(
        student_count=Count('group__user')
    ).order_by('province__name', 'name')

    # Аймгийн manager бол зөвхөн өөрийн аймгийн сургуулиудыг харуулах
    if managed_province:
        all_schools = all_schools.filter(province=managed_province)

    # URL-аас шүүлтүүрийн утгуудыг авах
    name_query = request.GET.get('q', '')
    zone_id = request.GET.get('z', '')
    inactive_filter = request.GET.get('inactive', '')
    active_filter = request.GET.get('active', '')

    # Шүүлтүүрүүдийг хийх
    if name_query:
        all_schools = all_schools.filter(name__icontains=name_query)
    # Province шүүлтүүр - зөвхөн staff хэрэглэгчдэд (province manager-т аль хэдийн шүүгдсэн)
    if province_id and is_staff and not managed_province:
        all_schools = all_schools.filter(province_id=province_id)
    if zone_id and is_staff and not managed_province:
        all_schools = all_schools.filter(province__zone_id=zone_id)

    if inactive_filter == '1': # Хэрэв URL-д ?inactive=1 гэж ирвэл
        seven_days_ago = timezone.now() - timedelta(days=7)
        # 7 хоногоос өмнө нэвтэрсэн ЭСВЭЛ огт нэвтрээгүй (last_login is NULL) хэрэглэгчдийг шүүнэ.
        all_schools = all_schools.filter(
            Q(user__last_login__lt=seven_days_ago) | Q(user__last_login__isnull=True)
        )

    if active_filter == '1': # Хэрэв URL-д ?active=1 гэж ирвэл
        seven_days_ago = timezone.now() - timedelta(days=7)
        # 7 хоногоос дотор нэвтэрсэн хэрэглэгчдийг шүүнэ.
        all_schools = all_schools.filter(
            Q(user__last_login__gt=seven_days_ago)
        )

    # Эрэмбэлэх
    all_schools = all_schools.order_by('province__name', 'name')

    change_form = SchoolModeratorChangeForm()

    context = {
        'schools': all_schools,
        'change_form': change_form,
        'managed_province': managed_province,
        'is_staff': is_staff,
    }
    return render(request, 'schools/manage_all_schools.html', context)

@staff_member_required
def change_school_admin_view(request, school_id):
    """
    Сургуулийн админыг хайж олоод солих хуудас.
    """
    school = get_object_or_404(School, id=school_id)
    search_results = None

    if request.method == 'POST':
        # Хэрэв "assign_admin" үйлдэл хийгдэж байвал
        if 'assign_admin' in request.POST:
            user_id = request.POST.get('user_id')
            new_admin = get_object_or_404(User, id=user_id)
            school.user = new_admin
            school.save()
            messages.success(request, f"'{school.name}' сургуулийн админыг '{new_admin.get_full_name()}' хэрэглэгчээр амжилттай солилоо.")
            return redirect('manage_all_schools')

        # Хэрэв "search_users" үйлдэл хийгдэж байвал
        search_form = UserSearchForm(request.POST)
        if search_form.is_valid():
            search_results = search_form.search_users()

    else:
        search_form = UserSearchForm()

    context = {
        'school': school,
        'search_form': search_form,
        'search_results': search_results,
    }
    return render(request, 'schools/change_school_admin.html', context)


@staff_member_required
def edit_school_admin_view(request, user_id):
    """Сургуулийн админы профайлыг засах хуудас."""
    target_user = get_object_or_404(User, id=user_id)
    # Админ хэрэглэгчид UserMeta байхгүй бол үүсгэх
    user_meta, created = UserMeta.objects.get_or_create(user=target_user)

    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=target_user)
        user_meta_form = UserMetaForm(request.POST, request.FILES, instance=user_meta)
        if user_form.is_valid() and user_meta_form.is_valid():
            user_form.save()
            user_meta_form.save()
            messages.success(request, f"'{target_user.get_full_name()}' хэрэглэгчийн мэдээллийг амжилттай шинэчиллээ.")
            return redirect('manage_all_schools')
    else:
        user_form = UserForm(instance=target_user)
        user_meta_form = UserMetaForm(instance=user_meta)

    context = {
        'user_form': user_form,
        'user_meta_form': user_meta_form,
        'target_user': target_user,
    }
    return render(request, 'schools/edit_school_admin.html', context)


@staff_member_required
def change_school_admin_password_view(request, user_id):
    """Сургуулийн админы нууц үгийг солих хуудас."""
    target_user = get_object_or_404(User, id=user_id)

    # Аюулгүй байдлын шалгалт: staff хэрэглэгч өөр staff-ийн нууц үгийг солихыг хориглох
    if target_user.is_staff or target_user.is_superuser:
        # Өөрийнхөөс бусад staff-ийн нууц үгийг солихгүй
        if target_user != request.user:
            messages.error(request, "Та өөр staff эрхтэй хэрэглэгчийн нууц үгийг солих боломжгүй.")
            return redirect('manage_all_schools')

    if request.method == 'POST':
        form = SchoolAdminPasswordChangeForm(user=target_user, data=request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f"'{target_user.get_full_name()}' хэрэглэгчийн нууц үгийг амжилттай солилоо.")
            return redirect('manage_all_schools')
    else:
        form = SchoolAdminPasswordChangeForm(user=target_user)

    context = {
        'form': form,
        'target_user': target_user,
    }
    return render(request, 'schools/change_password.html', context) # Өмнөх template-г дахин ашиглаж болно


def school_list_view(request):
    # Textarea-аас орж ирсэн түүхий текстийг авах
    query_string = request.GET.get('q', '')

    # Орж ирсэн текстийг зай, таслал, эсвэл мөрийн төгсгөлөөр нь салгаж цэвэрлэх
    # Жишээ: "a@a.com, b@b.com\nc@c.com" -> ['a@a.com', 'b@b.com', 'c@c.com']
    emails = [email for email in re.split(r'[\s,;]+', query_string) if email]

    # Үндсэн шүүлт хийгээгүй үеийн бүх сургуулийн жагсаалт
    schools = School.objects.select_related('user', 'province').order_by('name')

    # Хэрэв имэйл хаягууд орж ирсэн бол шүүлт хийх
    if emails:
        # __in шүүлтүүр нь жагсаалтад байгаа ямар нэг утгатай таарч байвал шүүнэ
        schools = schools.filter(user__email__in=emails)

    context = {
        'schools': schools,
        'search_query': query_string,  # Хэрэглэгчийн оруулсан утгыг буцааж харуулах
    }
    return render(request, 'schools/school_list.html', context)

@login_required
def edit_school_info_view(request, school_id):
    """
    Сургуулийн нэр, аймаг/дүүрэг засах болон модератор солих хоёр үйлдлийг нэг хуудсанд.
    Staff болон аймгийн manager нар хандах эрхтэй.
    """
    school = get_object_or_404(School, id=school_id)

    # Эрх шалгах
    is_staff = request.user.is_staff
    can_edit = False

    if is_staff:
        can_edit = True
    elif school.province:
        # Аймгийн manager эсэхийг шалгах
        if school.province.contact_person == request.user:
            can_edit = True
        elif request.user.groups.filter(name=f"Province_{school.province.id}_Managers").exists():
            can_edit = True

    if not can_edit:
        messages.error(request, 'Та энэ сургуулийн мэдээлэл засах эрхгүй.')
        return redirect('my_managed_schools')

    if request.method == 'POST':
        # Хэрэв сургуулийн мэдээлэл өөрчлөх form илгээгдсэн бол
        if 'save_school_info' in request.POST:
            info_form = EditSchoolInfoForm(request.POST, instance=school)
            if info_form.is_valid():
                info_form.save()
                messages.success(request, "Сургуулийн мэдээлэл амжилттай шинэчлэгдлээ.")
                # Province manager бол өөрийн аймгийн хуудас руу буцаах
                if not is_staff and school.province:
                    return redirect(f'/schools/manage-all/?p={school.province.id}')
                return redirect('manage_all_schools')
    else:
        info_form = EditSchoolInfoForm(instance=school)
        moderator_form = SchoolModeratorChangeForm()

    return render(request, 'schools/edit_school_info.html', {
        'school': school,
        'info_form': info_form,
    })


@login_required
def manager_change_moderator_view(request, school_id):
    """
    Менежер өөрийн сургуулийн модераторыг солих хуудас.
    Зөвхөн manager эрхтэй хэрэглэгч өөрийн сургуульд хандаж болно.
    """
    school = get_object_or_404(School, id=school_id)

    # Check if user is manager of this school
    if school.manager != request.user and not request.user.is_staff:
        messages.error(request, 'Та энэ үйлдлийг хийх эрхгүй байна.')
        return redirect('school_dashboard', school_id=school_id)

    search_results = None

    if request.method == 'POST':
        # Хэрэв "assign_moderator" үйлдэл хийгдэж байвал
        if 'assign_moderator' in request.POST:
            user_id = request.POST.get('user_id')
            new_moderator = get_object_or_404(User, id=user_id)
            school.user = new_moderator
            school.save()
            messages.success(request, f"'{school.name}' сургуулийн модераторыг '{new_moderator.get_full_name()}' хэрэглэгчээр амжилттай солилоо.")
            return redirect('school_dashboard', school_id=school_id)

        # Хэрэв "search_users" үйлдэл хийгдэж байвал
        search_form = UserSearchForm(request.POST)
        if search_form.is_valid():
            search_results = search_form.search_users()
    else:
        search_form = UserSearchForm()

    context = {
        'school': school,
        'search_form': search_form,
        'search_results': search_results,
    }
    return render(request, 'schools/manager_change_moderator.html', context)


@staff_member_required
def change_school_manager_view(request, school_id):
    """
    Staff хэрэглэгч сургуулийн менежерийг хайж олоод солих хуудас.
    """
    school = get_object_or_404(School, id=school_id)
    search_results = None

    if request.method == 'POST':
        # Хэрэв "assign_manager" үйлдэл хийгдэж байвал
        if 'assign_manager' in request.POST:
            user_id = request.POST.get('user_id')
            new_manager = get_object_or_404(User, id=user_id)
            school.manager = new_manager
            school.save()
            messages.success(request, f"'{school.name}' сургуулийн менежерийг '{new_manager.get_full_name()}' хэрэглэгчээр амжилттай солилоо.")
            return redirect('manage_all_schools')

        # Хэрэв "search_users" үйлдэл хийгдэж байвал
        search_form = UserSearchForm(request.POST)
        if search_form.is_valid():
            search_results = search_form.search_users()

    else:
        search_form = UserSearchForm()

    context = {
        'school': school,
        'search_form': search_form,
        'search_results': search_results,
    }
    return render(request, 'schools/change_school_manager.html', context)


@staff_member_required
def change_school_manager_password_view(request, user_id):
    """
    Сургуулийн менежерийн нууц үгийг солих хуудас.
    """
    target_user = get_object_or_404(User, id=user_id)

    # Аюулгүй байдлын шалгалт: staff хэрэглэгч өөр staff-ийн нууц үгийг солихыг хориглох
    if target_user.is_staff or target_user.is_superuser:
        # Өөрийнхөөс бусад staff-ийн нууц үгийг солихгүй
        if target_user != request.user:
            messages.error(request, "Та өөр staff эрхтэй хэрэглэгчийн нууц үгийг солих боломжгүй.")
            return redirect('manage_all_schools')

    if request.method == 'POST':
        form = SchoolAdminPasswordChangeForm(user=target_user, data=request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f"'{target_user.get_full_name()}' хэрэглэгчийн нууц үгийг амжилттай солилоо.")
            return redirect('manage_all_schools')
    else:
        form = SchoolAdminPasswordChangeForm(user=target_user)

    context = {
        'form': form,
        'target_user': target_user,
        'is_manager': True,
    }
    return render(request, 'schools/change_password.html', context)


@staff_member_required
def school_official_levels_view(request):
    """
    Сургуулиудын албан ёсны оролцооны түвшингүүдийг харуулах, засварлах хуудас
    """
    from accounts.models import Level, Province

    # Get all levels and provinces
    levels = Level.objects.all().order_by('id')
    provinces = Province.objects.all().order_by('name')

    # Handle POST request for updating levels
    if request.method == 'POST':
        school_id = request.POST.get('school_id')
        if school_id:
            try:
                school = School.objects.get(id=school_id)
                # Get selected level IDs
                selected_levels = request.POST.getlist('levels')

                # Clear and set levels explicitly
                if selected_levels:
                    # Convert to integers
                    level_ids = [int(lid) for lid in selected_levels]
                    school.official_levels.set(level_ids)
                else:
                    # Clear all levels if none selected
                    school.official_levels.clear()

                messages.success(request, f"'{school.name}' сургуулийн түвшингүүд амжилттай шинэчлэгдлээ.")
            except School.DoesNotExist:
                messages.error(request, "Сургууль олдсонгүй.")

        # Redirect back with filters
        redirect_url = request.path
        params = []
        if request.POST.get('p'):
            params.append(f"p={request.POST.get('p')}")
        if request.POST.get('l'):
            params.append(f"l={request.POST.get('l')}")
        if params:
            redirect_url += '?' + '&'.join(params)
        return redirect(redirect_url)

    # Get filter parameters
    selected_province_id = request.GET.get('p', '')
    selected_level_id = request.GET.get('l', '')

    # Base queryset
    schools = School.objects.select_related('province').prefetch_related('official_levels').order_by('province__name', 'name')

    # Apply filters
    if selected_province_id:
        schools = schools.filter(province_id=selected_province_id)

    if selected_level_id:
        schools = schools.filter(official_levels__id=selected_level_id)

    context = {
        'schools': schools,
        'levels': levels,
        'provinces': provinces,
        'selected_province_id': int(selected_province_id) if selected_province_id else '',
        'selected_level_id': int(selected_level_id) if selected_level_id else '',
        'total_schools': schools.count(),
    }

    return render(request, 'schools/official_levels.html', context)


@login_required
def merge_school_users(request, school_id):
    """Сургуулийн сурагчдын давхар бүртгэлийг нэгтгэх (шат дараатай)"""
    from accounts.models import UserMergeRequest
    from olympiad.models import Award, Comment, ScoreSheet

    school = get_object_or_404(School, id=school_id)
    if not school.user_has_access(request.user):
        messages.error(request, 'Та энэ сургуулийг удирдах эрхгүй.')
        return redirect('my_managed_schools')

    back_url = reverse('school_all_users', args=[school_id])

    # Parse user_ids
    user_ids_str = request.GET.get('user_ids', '') or request.POST.get('user_ids', '')
    try:
        user_ids = sorted([int(uid.strip()) for uid in user_ids_str.split(',') if uid.strip()])
    except ValueError:
        messages.error(request, 'Буруу ID байна.')
        return redirect('school_all_users', school_id=school_id)

    if len(user_ids) < 2:
        messages.error(request, 'Хамгийн багадаа 2 хэрэглэгч сонгоно уу.')
        return redirect('school_all_users', school_id=school_id)

    users = list(User.objects.filter(id__in=user_ids).select_related(
        'data', 'data__school', 'data__province', 'data__grade', 'data__level'))
    if len(users) != len(user_ids):
        messages.error(request, 'Зарим хэрэглэгч олдсонгүй.')
        return redirect('school_all_users', school_id=school_id)

    from django.utils import timezone as tz
    import datetime
    users.sort(key=lambda u: u.last_login or tz.make_aware(datetime.datetime.min), reverse=True)

    if request.method == 'POST':
        # === Execute merge ===
        primary_id = int(request.POST.get('primary_id'))
        primary_user = get_object_or_404(User, id=primary_id)
        duplicate_users = [u for u in users if u.id != primary_id]

        all_field_names = ['last_name', 'first_name', 'email', 'reg_num', 'school',
                           'province', 'grade', 'level', 'mobile', 'gender']
        field_selections = {}
        for fn in all_field_names:
            val = request.POST.get(f'field_{fn}', '').strip()
            if val:
                field_selections[fn] = val

        # Check conflicts
        has_name_conflict = False
        for nf in ['last_name', 'first_name']:
            vals = set(filter(None, [getattr(u, nf, '') or '' for u in users]))
            if len(vals) > 1:
                has_name_conflict = True
                break

        has_result_conflict = False
        for dup_user in duplicate_users:
            dup_results = Result.objects.filter(contestant=dup_user).select_related('olympiad', 'problem')
            for dup_result in dup_results:
                primary_result = Result.objects.filter(
                    contestant_id=primary_id,
                    olympiad=dup_result.olympiad,
                    problem=dup_result.problem
                ).first()
                if primary_result and (primary_result.answer != dup_result.answer or primary_result.score != dup_result.score):
                    has_result_conflict = True
                    break
            if has_result_conflict:
                break

        needs_staff_approval = (has_name_conflict or has_result_conflict) and not request.user.is_staff

        if needs_staff_approval:
            conflict_reasons = []
            if has_name_conflict:
                conflict_reasons.append('нэр зөрүүтэй')
            if has_result_conflict:
                conflict_reasons.append('дүнгийн давхцал')

            merge_request = UserMergeRequest.objects.create(
                requesting_user=request.user,
                user_ids=user_ids,
                primary_user=primary_user,
                reason=f'{school.name} сургуулийн бүртгэлээс нэгтгэх хүсэлт ({", ".join(conflict_reasons)}).',
                status=UserMergeRequest.Status.PENDING,
                requires_user_confirmation=False,
                conflicts_data={
                    'field_selections': field_selections,
                    'conflict_reasons': conflict_reasons,
                },
            )
            merge_request.detect_conflicts()
            merge_request.save()

            messages.info(
                request,
                f'Нэгтгэх хүсэлт #{merge_request.id} үүслээ ({", ".join(conflict_reasons)}). '
                f'Админ хянаад баталгаажуулна.'
            )
            return redirect('school_all_users', school_id=school_id)

        # Staff or no conflicts → execute merge
        primary_meta, _ = UserMeta.objects.get_or_create(user=primary_user)

        with transaction.atomic():
            user_fields = ['last_name', 'first_name', 'email']
            for field in user_fields:
                value = field_selections.get(field, '')
                if value:
                    setattr(primary_user, field, value)
            primary_user.save()

            meta_fields = ['reg_num', 'mobile', 'gender']
            meta_fk_fields = ['school', 'province', 'grade', 'level']

            for field in meta_fields:
                value = field_selections.get(field, '')
                if value:
                    if field == 'mobile':
                        try:
                            setattr(primary_meta, field, int(value))
                        except ValueError:
                            pass
                    else:
                        setattr(primary_meta, field, value)

            for field in meta_fk_fields:
                value = field_selections.get(field, '')
                if value:
                    try:
                        setattr(primary_meta, f'{field}_id', int(value))
                    except ValueError:
                        pass

            primary_meta.save()

            for dup_user in duplicate_users:
                primary_user.groups.add(*dup_user.groups.all())
                Result.objects.filter(contestant=dup_user).update(contestant=primary_user)
                Award.objects.filter(contestant=dup_user).update(contestant=primary_user)
                Comment.objects.filter(author=dup_user).update(author=primary_user)
                ScoreSheet.objects.filter(user=dup_user).update(user=primary_user)
                School.objects.filter(user=dup_user).update(user=primary_user)
                School.objects.filter(manager=dup_user).update(manager=primary_user)
                dup_user.delete()

            UserMergeRequest.objects.create(
                requesting_user=request.user,
                user_ids=user_ids,
                primary_user=primary_user,
                reason=f'{school.name} сургуулийн бүртгэлээс шууд нэгтгэсэн.',
                status=UserMergeRequest.Status.COMPLETED,
                requires_user_confirmation=False,
            )

        messages.success(request, f'Амжилттай нэгтгэлээ. Үндсэн хэрэглэгч: {primary_user.last_name} {primary_user.first_name} (ID: {primary_user.id})')
        return redirect('school_all_users', school_id=school_id)

    # === GET: Step 1 or Step 2 ===
    step = request.GET.get('step', '1')
    primary_id = request.GET.get('primary')

    if step == '2' and primary_id:
        primary_id = int(primary_id)
        primary_user = next((u for u in users if u.id == primary_id), None)
        if not primary_user:
            messages.error(request, 'Үндсэн хэрэглэгч олдсонгүй.')
            return redirect('school_all_users', school_id=school_id)

        compare_fields = []
        field_defs = [
            ('last_name', 'Овог', 'user'),
            ('first_name', 'Нэр', 'user'),
            ('email', 'Имэйл', 'user'),
            ('reg_num', 'РД', 'meta'),
            ('school', 'Сургууль', 'meta_fk'),
            ('province', 'Аймаг', 'meta_fk'),
            ('grade', 'Анги', 'meta_fk'),
            ('level', 'Ангилал', 'meta_fk'),
            ('mobile', 'Утас', 'meta'),
            ('gender', 'Хүйс', 'meta'),
        ]

        for field_name, label, field_type in field_defs:
            values = []
            for u in users:
                if field_type == 'user':
                    val = getattr(u, field_name, '') or ''
                    values.append({'user_id': u.id, 'value': str(val), 'raw_value': str(val)})
                elif field_type == 'meta':
                    meta = getattr(u, 'data', None)
                    val = getattr(meta, field_name, '') if meta else ''
                    val = val if val is not None else ''
                    values.append({'user_id': u.id, 'value': str(val), 'raw_value': str(val)})
                elif field_type == 'meta_fk':
                    meta = getattr(u, 'data', None)
                    obj = getattr(meta, field_name, None) if meta else None
                    display = str(obj) if obj else ''
                    raw = str(obj.id) if obj else ''
                    values.append({'user_id': u.id, 'value': display, 'raw_value': raw})

            non_empty = [v['value'] for v in values if v['value']]
            unique_values = set(non_empty)
            is_different = len(unique_values) > 1

            primary_val = next((v for v in values if v['user_id'] == primary_id), None)
            if primary_val and primary_val['raw_value']:
                auto_selected = primary_val['raw_value']
                auto_selected_display = primary_val['value']
            else:
                first_non_empty = next((v for v in values if v['raw_value']), None)
                auto_selected = first_non_empty['raw_value'] if first_non_empty else ''
                auto_selected_display = first_non_empty['value'] if first_non_empty else ''

            compare_fields.append({
                'name': field_name,
                'label': label,
                'field_type': field_type,
                'values': values,
                'is_different': is_different,
                'all_empty': len(non_empty) == 0,
                'auto_selected': auto_selected,
                'auto_selected_display': auto_selected_display,
            })

        result_conflicts = []
        duplicate_users_list = [u for u in users if u.id != primary_id]
        for dup_user in duplicate_users_list:
            dup_results = Result.objects.filter(contestant=dup_user).select_related('olympiad', 'problem')
            for dup_result in dup_results:
                primary_result = Result.objects.filter(
                    contestant_id=primary_id,
                    olympiad=dup_result.olympiad,
                    problem=dup_result.problem
                ).first()
                if primary_result and (primary_result.answer != dup_result.answer or primary_result.score != dup_result.score):
                    result_conflicts.append({
                        'dup_user': dup_user,
                        'olympiad_name': dup_result.olympiad.name,
                        'problem_order': dup_result.problem.order,
                        'primary_score': primary_result.score,
                        'dup_score': dup_result.score,
                        'primary_answer': primary_result.answer,
                        'dup_answer': dup_result.answer,
                    })

        has_name_conflict = any(
            f['is_different'] for f in compare_fields if f['name'] in ('last_name', 'first_name')
        )
        needs_staff_approval = (has_name_conflict or bool(result_conflicts)) and not request.user.is_staff

        province_schools = School.objects.filter(province=school.province).order_by('name')

        context = {
            'school': school,
            'users': users,
            'primary_user': primary_user,
            'primary_id': primary_id,
            'user_ids_str': ','.join(str(uid) for uid in user_ids),
            'compare_fields': compare_fields,
            'result_conflicts': result_conflicts,
            'needs_staff_approval': needs_staff_approval,
            'province_schools': province_schools,
            'step': 2,
        }
        return render(request, 'schools/merge_school_users.html', context)

    else:
        user_data = []
        for u in users:
            meta = getattr(u, 'data', None)
            user_data.append({
                'user': u,
                'school': meta.school if meta else None,
                'reg_num': meta.reg_num if meta else '',
                'mobile': meta.mobile if meta else '',
                'last_login': u.last_login,
            })

        context = {
            'school': school,
            'users': users,
            'user_data': user_data,
            'user_ids_str': ','.join(str(uid) for uid in user_ids),
            'default_primary_id': users[0].id if users else None,
            'step': 1,
        }
        return render(request, 'schools/merge_school_users.html', context)
