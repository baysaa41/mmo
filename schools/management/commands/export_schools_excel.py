from django.core.management.base import BaseCommand
from schools.models import School
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os


class Command(BaseCommand):
    help = 'Сургуулиудын нэр, ID бүхий Excel файл үүсгэх. Аймаг/Дүүрэг бүрээр тусдаа файл.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output-dir',
            type=str,
            default='schools_export',
            help='Файлууд хадгалах хавтас (default: schools_export)'
        )

    def handle(self, *args, **kwargs):
        output_dir = kwargs['output_dir']

        # Хавтас үүсгэх
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            self.stdout.write(self.style.SUCCESS(f'📁 Хавтас үүслээ: {output_dir}'))

        # Сургуулиудыг дүүргээр нь бүлэглэх
        schools_by_province = {}
        schools = School.objects.select_related('province').all()

        for school in schools:
            province_name = school.province.name if school.province else 'Аймаг/Дүүрэггүй'

            if province_name not in schools_by_province:
                schools_by_province[province_name] = []

            schools_by_province[province_name].append(school)

        # Province бүрээр тусдаа файл үүсгэх
        files_created = []
        total_schools = 0

        for province_name in sorted(schools_by_province.keys()):
            # Excel workbook үүсгэх
            wb = Workbook()
            ws = wb.active
            ws.title = province_name[:31]  # Sheet нэр - 31 тэмдэгтийн хязгаар

            # Толгой мөр
            ws['A1'] = '№'
            ws['B1'] = 'ID'
            ws['C1'] = 'Сургуулийн нэр'
            ws['D1'] = 'Удирдлагын овог'
            ws['E1'] = 'Удирдлагын нэр'
            ws['F1'] = 'Удирдлагын имэйл'

            # Толгой мөрийг тодотгох
            for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
                ws[cell].font = Font(bold=True)
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')

            # Сургуулиудыг нэрээр нь эрэмбэлэх
            schools_list = sorted(schools_by_province[province_name], key=lambda s: s.name)

            # Өгөгдөл оруулах
            for idx, school in enumerate(schools_list, start=1):
                row = idx + 1
                ws[f'A{row}'] = idx
                ws[f'B{row}'] = school.id
                ws[f'C{row}'] = school.name

                # Удирдлагын мэдээлэл - хоосон үлдээх (цуглуулах зорилгоор)
                ws[f'D{row}'] = ''
                ws[f'E{row}'] = ''
                ws[f'F{row}'] = ''

                # Дугаарыг төвлөрүүлэх
                ws[f'A{row}'].alignment = Alignment(horizontal='center')
                ws[f'B{row}'].alignment = Alignment(horizontal='center')

            # Баганын өргөнийг тохируулах
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 60
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 30

            # Файлын нэр үүсгэх (аймгийн нэртэй, зайгүй)
            safe_province_name = province_name.replace('/', '_').replace('\\', '_').replace(',', '').replace(' ', '_')
            output_file = os.path.join(output_dir, f'{safe_province_name}.xlsx')

            # Файл хадгалах
            wb.save(output_file)
            files_created.append(output_file)
            total_schools += len(schools_list)

            self.stdout.write(
                self.style.SUCCESS(f'✅ {province_name}: {len(schools_list)} сургууль → {output_file}')
            )

        # Эцсийн тайлан
        self.stdout.write('')
        self.stdout.write('=' * 80)
        self.stdout.write(self.style.SUCCESS('📊 ЭЦСИЙН ТАЙЛАН'))
        self.stdout.write('=' * 80)
        self.stdout.write(f'📁 Хавтас: {os.path.abspath(output_dir)}')
        self.stdout.write(f'📄 Нийт файл: {len(files_created)}')
        self.stdout.write(f'🏫 Нийт сургууль: {total_schools}')
        self.stdout.write('=' * 80)
