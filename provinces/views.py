from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Q

from accounts.models import Province, Zone
from olympiad.models import Olympiad, Award, ScoreSheet, Result, Problem, SchoolYear
from olympiad.utils.group_management import ensure_olympiad_has_group, get_or_create_round2_group
from schools.models import School
from django.contrib.auth.models import User, Group

import pandas as pd
import io
from openpyxl.styles import Border, Side, Font, Alignment


def user_can_manage_province(user, province):
    """Хэрэглэгч аймгаа удирдах эрхтэй эсэхийг шалгах"""
    if user.is_staff:
        return True

    if province.contact_person == user:
        return True

    # Province_{id}_Managers группт орсон эсэх
    group_name = f"Province_{province.id}_Managers"
    if user.groups.filter(name=group_name).exists():
        return True

    # edit_province эрхтэй эсэх
    if user.has_perm('accounts.edit_province'):
        return True

    return False


@login_required
def my_managed_provinces(request):
    """Удирдаж байгаа аймгуудын жагсаалт - зөвхөн өөрийн удирдаж байгаа аймгууд"""
    from django.db.models import Q

    is_staff_access = False

    if request.user.is_staff:
        # Staff бол бүх аймгийг харуулна
        managed_provinces = Province.objects.all().select_related('zone', 'contact_person').order_by('name')
        is_staff_access = True
    else:
        # Province_{id}_Managers group-д байгаа аймгуудын ID-г олох
        province_ids = []
        user_groups = request.user.groups.filter(name__startswith='Province_', name__endswith='_Managers')
        for group in user_groups:
            try:
                province_id = int(group.name.replace('Province_', '').replace('_Managers', ''))
                province_ids.append(province_id)
            except (ValueError, AttributeError):
                continue

        # Contact person эсвэл Province_{id}_Managers group-д байгаа аймгууд
        managed_provinces = Province.objects.filter(
            Q(contact_person=request.user) | Q(id__in=province_ids)
        ).select_related('zone', 'contact_person').distinct().order_by('name')

    context = {
        'provinces': managed_provinces,
        'is_staff_access': is_staff_access,
    }
    return render(request, 'provinces/my_managed_provinces.html', context)


@staff_member_required
def province_change_contact(request, province_id):
    """Аймгийн удирдах хүнийг солих (staff only)"""
    province = get_object_or_404(Province, id=province_id)

    if request.method == 'POST':
        user_id = request.POST.get('contact_person_id', '').strip()

        if not user_id:
            # Хоосон болгох
            province.contact_person = None
            province.save(update_fields=['contact_person'])
            messages.success(request, f'"{province.name}" аймгийн удирдах хүнийг хаслаа.')
        else:
            try:
                user = User.objects.get(id=int(user_id))
                province.contact_person = user
                province.save(update_fields=['contact_person'])
                messages.success(request, f'"{province.name}" аймгийн удирдах хүнийг {user.last_name} {user.first_name} (ID: {user.id}) болгож солилоо.')
            except (ValueError, User.DoesNotExist):
                messages.error(request, f'ID={user_id} хэрэглэгч олдсонгүй.')

    return redirect('province_dashboard', province_id=province_id)


@login_required
def province_dashboard(request, province_id):
    """Аймгийн удирдлагын самбар - Round 2 олимпиадуудыг харуулна"""
    province = get_object_or_404(Province, id=province_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Та энэ аймгийг удирдах эрхгүй.')
        return redirect('my_managed_provinces')

    # Одоогийн хичээлийн жилийг олох
    current_school_year = SchoolYear.get_current()

    # Зөвхөн одоогийн хичээлийн жилийн Round 2 олимпиадууд
    round2_olympiads = Olympiad.objects.filter(
        round=2,
        school_year=current_school_year
    ).select_related('level', 'school_year', 'group').order_by('level__name')

    # Бүлэг байхгүй олимпиадад автоматаар бүлэг үүсгэх
    for olympiad in round2_olympiads:
        group, created = ensure_olympiad_has_group(olympiad)
        if created:
            messages.info(request, f'"{olympiad.name}" олимпиадад "{group.name}" бүлэг автоматаар үүсгэгдлээ.')

    # Олимпиад бүрт province-specific тоо нэмэх
    from accounts.models import UserMeta

    for olympiad in round2_olympiads:
        # S, T ангилал эсэхийг шалгах (S (ББ), T (ДБ) гэх мэт форматтай байж болно)
        is_teacher = olympiad.level and (olympiad.level.name.startswith('S') or olympiad.level.name.startswith('T'))
        olympiad.is_teacher_olympiad = is_teacher

        if is_teacher:
            # Багш нарын олимпиад
            # Ангилалын эхний үсгийг авах (S эсвэл T)
            category_prefix = olympiad.level.name[0]

            # Тухайн аймгийн бүртгүүлсэн багш нар (олимпиадын group-д байгаа)
            if olympiad.group:
                olympiad.province_registered = olympiad.group.user_set.filter(
                    data__province=province,
                    data__level__name__startswith=category_prefix
                ).count()
            else:
                olympiad.province_registered = 0

            # Тухайн аймгийн нийт багш нар (тухайн ангилалаар)
            olympiad.province_total = User.objects.filter(
                data__province=province,
                data__level__name__startswith=category_prefix
            ).count()
        else:
            # Сурагчдын олимпиад
            # Тухайн аймгийн Round 2-т бүртгэлтэй сурагчид
            if olympiad.group:
                olympiad.province_registered = olympiad.group.user_set.filter(
                    data__province=province
                ).count()
            else:
                olympiad.province_registered = 0

            # Round 1 олимпиадууд
            round1_olympiads = Olympiad.objects.filter(next_round=olympiad)

            # Тухайн аймгийн Round 1-д оролцсон сурагчид
            olympiad.province_total = ScoreSheet.objects.filter(
                olympiad__in=round1_olympiads,
                user__data__province=province,
                is_official=True
            ).values('user').distinct().count()

    # Багш болон сурагчдын олимпиадыг ялгах
    teacher_olympiads = [o for o in round2_olympiads if o.is_teacher_olympiad]
    student_olympiads = [o for o in round2_olympiads if not o.is_teacher_olympiad]

    # Ангилалуудыг цуглуулах (давхцалгүй)
    teacher_levels = sorted(set(o.level.name for o in teacher_olympiads if o.level))
    student_levels = sorted(set(o.level.name for o in student_olympiads if o.level))

    context = {
        'province': province,
        'teacher_olympiads': teacher_olympiads,
        'student_olympiads': student_olympiads,
        'teacher_levels': ', '.join(teacher_levels),
        'student_levels': ', '.join(student_levels),
        'current_school_year': current_school_year,
    }
    return render(request, 'provinces/province_dashboard.html', context)


@login_required
def province_olympiad_view(request, province_id, olympiad_id):
    """Олимпиад удирдах үндсэн хуудас - 3 үндсэн үйлдэл"""
    province = get_object_or_404(Province, id=province_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Та энэ аймгийг удирдах эрхгүй.')
        return redirect('my_managed_provinces')

    # Олимпиад Round 2 эсэхийг шалгах
    if olympiad.round != 2:
        messages.error(request, 'Энэ олимпиад 2-р давааны олимпиад биш байна.')
        return redirect('province_dashboard', province_id=province_id)

    # Олимпиадад бүлэг байгаа эсэхийг шалгаад, байхгүй бол үүсгэх
    group, created = ensure_olympiad_has_group(olympiad)
    if created:
        messages.info(request, f'Олимпиадад "{group.name}" бүлэг автоматаар үүсгэгдлээ.')

    # S, T ангилал эсэхийг шалгах (багш нарын олимпиад)
    is_teacher_olympiad = olympiad.level and (olympiad.level.name.startswith('S') or olympiad.level.name.startswith('T'))

    # Олимпиадын нийт бүртгэлтэй тоо
    total_registered = olympiad.group.user_set.count() if olympiad.group else 0

    # Энэ аймгийн бүртгэлтэй тоо
    province_registered = 0
    if olympiad.group:
        province_registered = olympiad.group.user_set.filter(
            data__province=province
        ).count()

    if is_teacher_olympiad:
        # S, T ангилал - багш нарын мэдээлэл
        # Round 1 байхгүй, зөвхөн багш нарын тоо
        # Ангилалын эхний үсгийг авах (S эсвэл T)
        category_prefix = olympiad.level.name[0]

        total_teachers = User.objects.filter(
            data__province=province,
            data__level__name__startswith=category_prefix
        ).count()

        context = {
            'province': province,
            'olympiad': olympiad,
            'total_registered': total_registered,
            'province_registered': province_registered,
            'total_teachers': total_teachers,
            'is_teacher_olympiad': True,
        }
    else:
        # Сурагчдын олимпиад - Round 1 мэдээлэлтэй
        # Round 1 олимпиадууд
        round1_olympiads = Olympiad.objects.filter(next_round=olympiad)

        # Энэ аймгийн Round 1-д оролцсон сурагчдын тоо
        round1_participants = ScoreSheet.objects.filter(
            olympiad__in=round1_olympiads,
            user__data__province=province,
            is_official=True
        ).values('user').distinct().count()

        # Автомат бүртгэлтэй (Award-тай) сурагчдын тоо (энэ аймгийн)
        auto_registered = Award.objects.filter(
            olympiad__in=round1_olympiads,
            place__startswith='2.1',
            contestant__data__province=province
        ).count()

        context = {
            'province': province,
            'olympiad': olympiad,
            'total_registered': total_registered,
            'province_registered': province_registered,
            'round1_participants': round1_participants,
            'auto_registered_count': auto_registered,
            'is_teacher_olympiad': False,
        }
    return render(request, 'provinces/province_olympiad_detail.html', context)


@login_required
def add_students_by_threshold(request, province_id, olympiad_id):
    """Онооны босгоор нэмэлт сурагч бүртгэх"""
    province = get_object_or_404(Province, id=province_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_provinces')

    if request.method == 'POST':
        from .forms import ThresholdScoreForm
        form = ThresholdScoreForm(request.POST)

        if form.is_valid():
            threshold = form.cleaned_data['threshold_score']

            # Олимпиадад бүлэг байгаа эсэхийг шалгаад, байхгүй бол үүсгэх
            group, created = ensure_olympiad_has_group(olympiad)

            # Round 1 олимпиадуудыг олох (энэ Round 2-г зааж байгаа)
            round1_olympiads = Olympiad.objects.filter(next_round=olympiad)

            # Босгоос дээш оноотой, аль хэдийн группт ороогүй сурагчдыг олох
            eligible_scoresheets = ScoreSheet.objects.filter(
                olympiad__in=round1_olympiads,
                total__gte=threshold,
                is_official=True,
                user__data__province=province
            ).select_related('user').exclude(
                user__groups=group
            )

            # Award үүсгэж, группт нэмэх
            added_count = 0
            with transaction.atomic():
                for scoresheet in eligible_scoresheets:
                    # Award үүсгэх
                    Award.objects.get_or_create(
                        olympiad=scoresheet.olympiad,
                        contestant=scoresheet.user,
                        defaults={'place': '2.1 эрх нэмэлтээр (босго)'}
                    )

                    # Группт нэмэх
                    group.user_set.add(scoresheet.user)
                    added_count += 1

            messages.success(request, f'{added_count} сурагч нэмэгдлээ.')
            return redirect('province_olympiad_view', province_id=province_id, olympiad_id=olympiad_id)

    else:
        from .forms import ThresholdScoreForm
        form = ThresholdScoreForm()

    context = {
        'province': province,
        'olympiad': olympiad,
        'form': form,
    }
    return render(request, 'provinces/add_by_threshold.html', context)


@login_required
def select_top_students_by_school(request, province_id, olympiad_id):
    """Сургууль бүрээс эрх аваагүй 3 хүртэл өндөр оноотой сурагч сонгох"""
    province = get_object_or_404(Province, id=province_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_provinces')

    # Round 1 олимпиадууд
    round1_olympiads = Olympiad.objects.filter(next_round=olympiad)

    # Эрх авсан хүмүүсийн ID → place
    qualified_awards = {}
    for award in Award.objects.filter(
        olympiad__in=round1_olympiads,
        place__startswith='2.1 эрх',
    ):
        qualified_awards[award.contestant_id] = award.place

    # Сургуулиудаар бүлэглэж, эрх авсан + ижил оноотой сурагчдыг харуулах
    schools_data = []
    schools = School.objects.filter(province=province)

    for school in schools:
        # Энэ сургуулийн бүх ScoreSheet (оноогоор буурахаар)
        all_school_scoresheets = ScoreSheet.objects.filter(
            olympiad__in=round1_olympiads,
            school=school,
            is_official=True
        ).select_related('user').order_by('-total')

        if not all_school_scoresheets.exists():
            continue

        # Эрх авсан сурагчдыг ялгах
        qualified_scoresheets = []
        for ss in all_school_scoresheets:
            if ss.user_id in qualified_awards:
                ss.award_place = qualified_awards[ss.user_id]
                ss.is_qualified = True
                qualified_scoresheets.append(ss)

        if not qualified_scoresheets:
            continue

        # Эрх аваагүй сурагчдаас хамгийн өндөр оноотой + ижил оноотой сурагчид
        top_unqualified_score = None
        for ss in all_school_scoresheets:
            if ss.user_id not in qualified_awards:
                top_unqualified_score = ss.total
                break

        tie_scoresheets = []
        if top_unqualified_score is not None:
            for ss in all_school_scoresheets:
                if ss.user_id not in qualified_awards and ss.total == top_unqualified_score:
                    ss.is_qualified = False
                    ss.award_place = None
                    tie_scoresheets.append(ss)

        students_data = qualified_scoresheets + tie_scoresheets
        # Оноогоор буурахаар эрэмбэлэх
        students_data.sort(key=lambda s: (-s.total, not s.is_qualified))

        schools_data.append({
            'school': school,
            'students': students_data,
            'top_unqualified_score': top_unqualified_score,
            'qualified_count': len(qualified_scoresheets),
            'tie_count': len(tie_scoresheets),
        })

    # POST: сонгосон сурагчдыг бүртгэх
    if request.method == 'POST':
        selected_user_ids = request.POST.getlist('selected_users')

        selected_scoresheets = ScoreSheet.objects.filter(
            olympiad__in=round1_olympiads,
            user_id__in=selected_user_ids
        ).select_related('school', 'user')

        # Олимпиадад бүлэг байгаа эсэхийг шалгаад, байхгүй бол үүсгэх
        group, created = ensure_olympiad_has_group(olympiad)

        # Бүртгэх
        added_count = 0
        with transaction.atomic():
            for ss in selected_scoresheets:
                # Award үүсгэх
                Award.objects.get_or_create(
                    olympiad=ss.olympiad,
                    contestant=ss.user,
                    defaults={'place': '2.1 эрх нэмэлтээр (гараар)'}
                )

                # Группт нэмэх
                group.user_set.add(ss.user)
                added_count += 1

        messages.success(request, f'{added_count} сурагч нэмэгдлээ.')
        return redirect('province_olympiad_view', province_id=province_id, olympiad_id=olympiad_id)

    context = {
        'province': province,
        'olympiad': olympiad,
        'schools_data': schools_data,
    }
    return render(request, 'provinces/select_top_students.html', context)


@login_required
def add_students_by_id(request, province_id, olympiad_id):
    """ID-гаар сурагч/багш нэмэх"""
    province = get_object_or_404(Province, id=province_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_provinces')

    # S, T ангилал эсэхийг шалгах (багш нарын олимпиад)
    is_teacher_olympiad = olympiad.level and (olympiad.level.name.startswith('S') or olympiad.level.name.startswith('T'))

    if request.method == 'POST':
        user_ids_input = request.POST.get('user_ids', '').strip()

        if not user_ids_input:
            messages.error(request, 'Хэрэглэгчийн ID оруулна уу.')
            return redirect('add_students_by_id', province_id=province_id, olympiad_id=olympiad_id)

        # Parse user IDs (comma, space, newline separated)
        import re
        user_ids = re.split(r'[,\s\n]+', user_ids_input)
        user_ids = [uid.strip() for uid in user_ids if uid.strip()]

        # Convert to integers
        try:
            user_ids = [int(uid) for uid in user_ids]
        except ValueError:
            messages.error(request, 'ID буруу байна. Зөвхөн тоон утга оруулна уу.')
            return redirect('add_students_by_id', province_id=province_id, olympiad_id=olympiad_id)

        # Олимпиадад бүлэг байгаа эсэхийг шалгаад, байхгүй бол үүсгэх
        group, created = ensure_olympiad_has_group(olympiad)

        # User-үүдийг олох
        users = User.objects.filter(id__in=user_ids).select_related('data', 'data__province')

        found_ids = set(users.values_list('id', flat=True))
        not_found_ids = set(user_ids) - found_ids

        if not_found_ids:
            messages.warning(request, f'Олдоогүй ID: {", ".join(map(str, not_found_ids))}')

        # Бүртгэх
        added_count = 0
        already_in_group = 0
        wrong_province = 0

        if is_teacher_olympiad:
            # Багш нарын олимпиад - Province шалгалтгүй, Round 1 байхгүй
            with transaction.atomic():
                for user in users:
                    # Группт аль хэдийн байгаа эсэхийг шалгах
                    if group.user_set.filter(id=user.id).exists():
                        already_in_group += 1
                        continue

                    # Award үүсгэх - Round 2 олимпиадад шууд
                    Award.objects.get_or_create(
                        olympiad=olympiad,
                        contestant=user,
                        defaults={'place': '2.1 эрх (ID)'}
                    )

                    # Группт нэмэх
                    group.user_set.add(user)
                    added_count += 1

            # Мессеж
            if added_count > 0:
                messages.success(request, f'{added_count} багш нэмэгдлээ.')
            if already_in_group > 0:
                messages.info(request, f'{already_in_group} багш аль хэдийн группт байна.')
        else:
            # Сурагчдын олимпиад - Province шалгах, Round 1 хайх
            # Round 1 олимпиадууд
            round1_olympiads = Olympiad.objects.filter(next_round=olympiad)

            with transaction.atomic():
                for user in users:
                    # Province шалгах
                    if hasattr(user, 'data') and user.data and user.data.province != province:
                        wrong_province += 1
                        continue

                    # Группт аль хэдийн байгаа эсэхийг шалгах
                    if group.user_set.filter(id=user.id).exists():
                        already_in_group += 1
                        continue

                    # Award үүсгэх - Round 1-д оролцсон бол тухайн олимпиадад, үгүй бол эхний Round 1-д
                    scoresheet = ScoreSheet.objects.filter(
                        olympiad__in=round1_olympiads,
                        user=user,
                        is_official=True
                    ).first()

                    if scoresheet:
                        award_olympiad = scoresheet.olympiad
                    else:
                        # Round 1 олимпиад олдохгүй бол анхны олимпиад руу
                        award_olympiad = round1_olympiads.first() if round1_olympiads.exists() else olympiad

                    Award.objects.get_or_create(
                        olympiad=award_olympiad,
                        contestant=user,
                        defaults={'place': '2.1 эрх нэмэлтээр (ID)'}
                    )

                    # Группт нэмэх
                    group.user_set.add(user)
                    added_count += 1

            # Мессеж
            if added_count > 0:
                messages.success(request, f'{added_count} сурагч нэмэгдлээ.')
            if already_in_group > 0:
                messages.info(request, f'{already_in_group} сурагч аль хэдийн группт байна.')
            if wrong_province > 0:
                messages.warning(request, f'{wrong_province} сурагч өөр аймгийнх байна.')

        return redirect('province_olympiad_view', province_id=province_id, olympiad_id=olympiad_id)

    context = {
        'province': province,
        'olympiad': olympiad,
        'is_teacher_olympiad': is_teacher_olympiad,
    }
    return render(request, 'provinces/add_students_by_id.html', context)


@login_required
def select_teachers(request, province_id, olympiad_id):
    """Тухайн аймгийн багш нарын жагсаалтаас сонгож бүртгэх"""
    province = get_object_or_404(Province, id=province_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_provinces')

    # Багш нарын олимпиад эсэхийг шалгах
    is_teacher_olympiad = olympiad.level and (olympiad.level.name.startswith('S') or olympiad.level.name.startswith('T'))
    if not is_teacher_olympiad:
        messages.error(request, 'Энэ олимпиад багш нарын олимпиад биш байна.')
        return redirect('province_olympiad_view', province_id=province_id, olympiad_id=olympiad_id)

    # Олимпиадад бүлэг байгаа эсэхийг шалгаад, байхгүй бол үүсгэх
    group, created = ensure_olympiad_has_group(olympiad)

    # Ангилалын эхний үсгийг авах (S эсвэл T)
    category_prefix = olympiad.level.name[0]

    # Тухайн аймгийн тохирох ангилалын бүх багш нар
    teachers = User.objects.filter(
        data__province=province,
        data__level__name__startswith=category_prefix
    ).select_related('data', 'data__school').order_by('last_name', 'first_name')

    # Аль хэдийн бүртгэлтэй багш нарын ID
    registered_user_ids = set(group.user_set.values_list('id', flat=True))

    # Өмнөх жил оролцсон багш нарыг олох
    previous_year_ids = set()
    current_school_year = SchoolYear.get_current()
    if current_school_year:
        previous_school_year = SchoolYear.objects.filter(
            name__lt=current_school_year.name
        ).first()
        if previous_school_year:
            # Өмнөх жилийн ижил level, round=2 олимпиад
            prev_olympiads = Olympiad.objects.filter(
                school_year=previous_school_year,
                level=olympiad.level,
                round=2
            )
            for prev_oly in prev_olympiads:
                # Group гишүүд
                if prev_oly.group:
                    prev_group_ids = set(prev_oly.group.user_set.values_list('id', flat=True))
                    previous_year_ids.update(prev_group_ids)
                # ScoreSheet-тэй хэрэглэгчид
                prev_ss_ids = set(ScoreSheet.objects.filter(
                    olympiad=prev_oly
                ).values_list('user_id', flat=True))
                previous_year_ids.update(prev_ss_ids)

    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_users')
        selected_ids_set = {int(uid) for uid in selected_ids if uid.isdigit()}

        # Тухайн аймгийн багш нарын ID (form-д байгаа бүх багш)
        all_teacher_ids = set(teachers.values_list('id', flat=True))

        added_count = 0
        removed_count = 0
        with transaction.atomic():
            # Шинээр нэмэх (сонгосон боловч бүртгэлгүй)
            to_add = selected_ids_set - registered_user_ids
            for user_id in to_add:
                if user_id not in all_teacher_ids:
                    continue
                try:
                    user = User.objects.get(id=user_id)
                except User.DoesNotExist:
                    continue

                Award.objects.get_or_create(
                    olympiad=olympiad,
                    contestant=user,
                    defaults={'place': '2.1 эрх (сонгосон)'}
                )
                group.user_set.add(user)
                added_count += 1

            # Цуцлах (бүртгэлтэй байсан боловч сонгоогүй)
            to_remove = (registered_user_ids & all_teacher_ids) - selected_ids_set
            for user_id in to_remove:
                group.user_set.remove(user_id)
                Award.objects.filter(
                    olympiad=olympiad,
                    contestant_id=user_id
                ).delete()
                removed_count += 1

        if added_count > 0:
            messages.success(request, f'{added_count} багш нэмэгдлээ.')
        if removed_count > 0:
            messages.warning(request, f'{removed_count} багшийн бүртгэл цуцлагдлаа.')
        if added_count == 0 and removed_count == 0:
            messages.info(request, 'Өөрчлөлт байхгүй.')
        return redirect('province_olympiad_view', province_id=province_id, olympiad_id=olympiad_id)

    # Багш нарын жагсаалт бэлдэх
    teachers_list = []
    for teacher in teachers:
        teachers_list.append({
            'user': teacher,
            'school': teacher.data.school if hasattr(teacher, 'data') and teacher.data else None,
            'is_registered': teacher.id in registered_user_ids,
            'is_previous_year': teacher.id in previous_year_ids,
        })

    context = {
        'province': province,
        'olympiad': olympiad,
        'teachers_list': teachers_list,
        'registered_count': len(registered_user_ids),
        'previous_year_count': len(previous_year_ids),
        'total_teachers': len(teachers_list),
    }
    return render(request, 'provinces/select_teachers.html', context)


@login_required
def merge_teacher_list(request, province_id, olympiad_id):
    """Багш нарын жагсаалтаас нэгтгэх хэрэглэгчдийг сонгох"""
    province = get_object_or_404(Province, id=province_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_provinces')

    is_teacher_olympiad = olympiad.level and (olympiad.level.name.startswith('S') or olympiad.level.name.startswith('T'))
    if not is_teacher_olympiad:
        messages.error(request, 'Энэ олимпиад багш нарын олимпиад биш байна.')
        return redirect('province_olympiad_view', province_id=province_id, olympiad_id=olympiad_id)

    category_prefix = olympiad.level.name[0]

    teachers = User.objects.filter(
        data__province=province,
        data__level__name__startswith=category_prefix
    ).select_related('data', 'data__school').order_by('last_name', 'first_name')

    teachers_list = []
    for teacher in teachers:
        teachers_list.append({
            'user': teacher,
            'school': teacher.data.school if hasattr(teacher, 'data') and teacher.data else None,
            'reg_num': teacher.data.reg_num if hasattr(teacher, 'data') and teacher.data else '',
        })

    context = {
        'province': province,
        'olympiad': olympiad,
        'teachers_list': teachers_list,
        'total_teachers': len(teachers_list),
    }
    return render(request, 'provinces/merge_teacher_list.html', context)


@login_required
def request_teacher_merge(request, province_id, olympiad_id):
    """Давхар бүртгэлтэй багш нарыг нэгтгэх хүсэлт явуулах"""
    from accounts.models import UserMergeRequest

    province = get_object_or_404(Province, id=province_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_provinces')

    if request.method != 'POST':
        return redirect('select_teachers', province_id=province_id, olympiad_id=olympiad_id)

    user_ids_str = request.POST.get('user_ids', '')
    try:
        user_ids = sorted([int(uid.strip()) for uid in user_ids_str.split(',') if uid.strip()])
    except ValueError:
        messages.error(request, 'Буруу ID байна.')
        return redirect('select_teachers', province_id=province_id, olympiad_id=olympiad_id)

    if len(user_ids) < 2:
        messages.error(request, 'Хамгийн багадаа 2 хэрэглэгч сонгоно уу.')
        return redirect('select_teachers', province_id=province_id, olympiad_id=olympiad_id)

    # Аль хэдийн хүсэлт байгаа эсэхийг шалгах
    existing = UserMergeRequest.objects.filter(
        user_ids=user_ids,
        status__in=['pending', 'approved']
    ).first()
    if existing:
        messages.info(request, f'Энэ бүлгийн нэгтгэх хүсэлт #{existing.id} аль хэдийн үүссэн байна.')
        return redirect('select_teachers', province_id=province_id, olympiad_id=olympiad_id)

    # Хэрэглэгчид байгаа эсэхийг шалгах
    users = User.objects.filter(id__in=user_ids)
    if users.count() != len(user_ids):
        messages.error(request, 'Зарим хэрэглэгч олдсонгүй.')
        return redirect('select_teachers', province_id=province_id, olympiad_id=olympiad_id)

    # Нэгтгэх хүсэлт үүсгэх
    merge_request = UserMergeRequest.objects.create(
        requesting_user=request.user,
        user_ids=user_ids,
        reason=f'{province.name} - {olympiad.name} багш нарын бүртгэлээс давхар бүртгэл илэрсэн.',
        status=UserMergeRequest.Status.PENDING,
        requires_user_confirmation=False
    )

    merge_request.detect_conflicts()
    merge_request.save()

    messages.success(
        request,
        f'Нэгтгэх хүсэлт #{merge_request.id} амжилттай үүслээ. '
        f'Админ хянаад нэгтгэнэ.'
    )
    return redirect('select_teachers', province_id=province_id, olympiad_id=olympiad_id)


@login_required
def merge_teachers(request, province_id, olympiad_id):
    """Давхар бүртгэлтэй багш нарыг шууд нэгтгэх (шат дараатай)"""
    from accounts.models import UserMergeRequest, UserMeta
    from olympiad.models import Result, Award, Comment
    import re

    province = get_object_or_404(Province, id=province_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_provinces')

    # Parse user_ids
    user_ids_str = request.GET.get('user_ids', '') or request.POST.get('user_ids', '')
    try:
        user_ids = sorted([int(uid.strip()) for uid in user_ids_str.split(',') if uid.strip()])
    except ValueError:
        messages.error(request, 'Буруу ID байна.')
        return redirect('select_teachers', province_id=province_id, olympiad_id=olympiad_id)

    if len(user_ids) < 2:
        messages.error(request, 'Хамгийн багадаа 2 хэрэглэгч сонгоно уу.')
        return redirect('select_teachers', province_id=province_id, olympiad_id=olympiad_id)

    users = list(User.objects.filter(id__in=user_ids).select_related('data', 'data__school', 'data__province', 'data__grade', 'data__level'))
    if len(users) != len(user_ids):
        messages.error(request, 'Зарим хэрэглэгч олдсонгүй.')
        return redirect('select_teachers', province_id=province_id, olympiad_id=olympiad_id)

    # Sort by last_login descending (most recent first)
    from django.utils import timezone as tz
    import datetime
    users.sort(key=lambda u: u.last_login or tz.make_aware(datetime.datetime.min), reverse=True)

    if request.method == 'POST':
        # === STEP 3: Execute merge ===
        primary_id = int(request.POST.get('primary_id'))
        primary_user = get_object_or_404(User, id=primary_id)
        duplicate_users = [u for u in users if u.id != primary_id]

        # Collect selected field values from POST
        all_field_names = ['last_name', 'first_name', 'email', 'reg_num', 'school',
                           'province', 'grade', 'level', 'mobile', 'gender']
        field_selections = {}
        for fn in all_field_names:
            val = request.POST.get(f'field_{fn}', '').strip()
            if val:
                field_selections[fn] = val

        # Check for name differences and result conflicts
        has_name_conflict = False
        name_fields = ['last_name', 'first_name']
        for nf in name_fields:
            vals = set()
            for u in users:
                v = getattr(u, nf, '') or ''
                if v:
                    vals.add(v)
            if len(vals) > 1:
                has_name_conflict = True
                break

        has_result_conflict = False
        for dup_user in duplicate_users:
            dup_results = Result.objects.filter(contestant=dup_user).select_related('olympiad', 'problem')
            for dup_result in dup_results:
                primary_result = Result.objects.filter(
                    contestant_id=primary_id,
                    olympiad=dup_result.olympiad,
                    problem=dup_result.problem
                ).first()
                if primary_result and (primary_result.answer != dup_result.answer or primary_result.score != dup_result.score):
                    has_result_conflict = True
                    break
            if has_result_conflict:
                break

        needs_staff_approval = has_name_conflict or has_result_conflict

        # Non-staff with conflicts → create pending merge request
        if needs_staff_approval and not request.user.is_staff:
            conflict_reasons = []
            if has_name_conflict:
                conflict_reasons.append('нэр зөрүүтэй')
            if has_result_conflict:
                conflict_reasons.append('дүнгийн давхцал')

            merge_request = UserMergeRequest.objects.create(
                requesting_user=request.user,
                user_ids=user_ids,
                primary_user=primary_user,
                reason=f'{province.name} - {olympiad.name} багш нарын бүртгэлээс нэгтгэх хүсэлт ({", ".join(conflict_reasons)}).',
                status=UserMergeRequest.Status.PENDING,
                requires_user_confirmation=False,
                conflicts_data={
                    'field_selections': field_selections,
                    'conflict_reasons': conflict_reasons,
                },
            )
            merge_request.detect_conflicts()
            merge_request.save()

            messages.info(
                request,
                f'Нэгтгэх хүсэлт #{merge_request.id} үүслээ ({", ".join(conflict_reasons)}). '
                f'Админ хянаад баталгаажуулна.'
            )
            return redirect('select_teachers', province_id=province_id, olympiad_id=olympiad_id)

        # Staff or no conflicts → execute merge directly
        primary_meta, _ = UserMeta.objects.get_or_create(user=primary_user)

        with transaction.atomic():
            # Apply selected field values
            user_fields = ['last_name', 'first_name', 'email']
            for field in user_fields:
                value = field_selections.get(field, '')
                if value:
                    setattr(primary_user, field, value)
            primary_user.save()

            meta_fields = ['reg_num', 'mobile', 'gender']
            meta_fk_fields = ['school', 'province', 'grade', 'level']

            for field in meta_fields:
                value = field_selections.get(field, '')
                if value:
                    if field == 'mobile':
                        try:
                            setattr(primary_meta, field, int(value))
                        except ValueError:
                            pass
                    else:
                        setattr(primary_meta, field, value)

            for field in meta_fk_fields:
                value = field_selections.get(field, '')
                if value:
                    try:
                        setattr(primary_meta, f'{field}_id', int(value))
                    except ValueError:
                        pass

            primary_meta.save()

            # Migrate relationships from duplicates
            for dup_user in duplicate_users:
                # Add groups
                primary_user.groups.add(*dup_user.groups.all())

                # Update foreign keys
                Result.objects.filter(contestant=dup_user).update(contestant=primary_user)
                Award.objects.filter(contestant=dup_user).update(contestant=primary_user)
                Comment.objects.filter(author=dup_user).update(author=primary_user)
                ScoreSheet.objects.filter(user=dup_user).update(user=primary_user)
                School.objects.filter(user=dup_user).update(user=primary_user)
                School.objects.filter(manager=dup_user).update(manager=primary_user)

                # Delete duplicate user
                dup_user.delete()

            # Create completed merge request record
            UserMergeRequest.objects.create(
                requesting_user=request.user,
                user_ids=user_ids,
                primary_user=primary_user,
                reason=f'{province.name} - {olympiad.name} багш нарын бүртгэлээс шууд нэгтгэсэн.',
                status=UserMergeRequest.Status.COMPLETED,
                requires_user_confirmation=False,
            )

        messages.success(request, f'Амжилттай нэгтгэлээ. Үндсэн хэрэглэгч: {primary_user.last_name} {primary_user.first_name} (ID: {primary_user.id})')
        return redirect('select_teachers', province_id=province_id, olympiad_id=olympiad_id)

    # === GET: Step 1 or Step 2 ===
    step = request.GET.get('step', '1')
    primary_id = request.GET.get('primary')

    if step == '2' and primary_id:
        # STEP 2: Compare fields
        primary_id = int(primary_id)
        primary_user = next((u for u in users if u.id == primary_id), None)
        if not primary_user:
            messages.error(request, 'Үндсэн хэрэглэгч олдсонгүй.')
            return redirect('select_teachers', province_id=province_id, olympiad_id=olympiad_id)

        # Build comparison fields
        compare_fields = []
        field_defs = [
            ('last_name', 'Овог', 'user'),
            ('first_name', 'Нэр', 'user'),
            ('email', 'Имэйл', 'user'),
            ('reg_num', 'РД', 'meta'),
            ('school', 'Сургууль', 'meta_fk'),
            ('province', 'Аймаг', 'meta_fk'),
            ('grade', 'Анги', 'meta_fk'),
            ('level', 'Ангилал', 'meta_fk'),
            ('mobile', 'Утас', 'meta'),
            ('gender', 'Хүйс', 'meta'),
        ]

        for field_name, label, field_type in field_defs:
            values = []
            for u in users:
                if field_type == 'user':
                    val = getattr(u, field_name, '') or ''
                    values.append({'user_id': u.id, 'value': str(val), 'raw_value': str(val)})
                elif field_type == 'meta':
                    meta = getattr(u, 'data', None)
                    val = getattr(meta, field_name, '') if meta else ''
                    val = val if val is not None else ''
                    values.append({'user_id': u.id, 'value': str(val), 'raw_value': str(val)})
                elif field_type == 'meta_fk':
                    meta = getattr(u, 'data', None)
                    obj = getattr(meta, field_name, None) if meta else None
                    display = str(obj) if obj else ''
                    raw = str(obj.id) if obj else ''
                    values.append({'user_id': u.id, 'value': display, 'raw_value': raw})

            # Determine if values differ
            non_empty = [v['value'] for v in values if v['value']]
            unique_values = set(non_empty)
            is_different = len(unique_values) > 1

            # Auto-select: primary's value if exists, otherwise first non-empty
            primary_val = next((v for v in values if v['user_id'] == primary_id), None)
            if primary_val and primary_val['raw_value']:
                auto_selected = primary_val['raw_value']
                auto_selected_display = primary_val['value']
            else:
                first_non_empty = next((v for v in values if v['raw_value']), None)
                auto_selected = first_non_empty['raw_value'] if first_non_empty else ''
                auto_selected_display = first_non_empty['value'] if first_non_empty else ''

            compare_fields.append({
                'name': field_name,
                'label': label,
                'field_type': field_type,
                'values': values,
                'is_different': is_different,
                'all_empty': len(non_empty) == 0,
                'auto_selected': auto_selected,
                'auto_selected_display': auto_selected_display,
            })

        # Detect Result conflicts
        result_conflicts = []
        duplicate_users_list = [u for u in users if u.id != primary_id]
        for dup_user in duplicate_users_list:
            dup_results = Result.objects.filter(contestant=dup_user).select_related('olympiad', 'problem')
            for dup_result in dup_results:
                primary_result = Result.objects.filter(
                    contestant_id=primary_id,
                    olympiad=dup_result.olympiad,
                    problem=dup_result.problem
                ).first()
                if primary_result and (primary_result.answer != dup_result.answer or primary_result.score != dup_result.score):
                    result_conflicts.append({
                        'dup_user': dup_user,
                        'olympiad_name': dup_result.olympiad.name,
                        'problem_order': dup_result.problem.order,
                        'primary_score': primary_result.score,
                        'dup_score': dup_result.score,
                        'primary_answer': primary_result.answer,
                        'dup_answer': dup_result.answer,
                    })

        # Check if name fields differ
        has_name_conflict = any(
            f['is_different'] for f in compare_fields if f['name'] in ('last_name', 'first_name')
        )
        needs_staff_approval = (has_name_conflict or bool(result_conflicts)) and not request.user.is_staff

        # Province schools for school selector
        province_schools = School.objects.filter(province=province).order_by('name')

        context = {
            'province': province,
            'olympiad': olympiad,
            'users': users,
            'primary_user': primary_user,
            'primary_id': primary_id,
            'user_ids_str': ','.join(str(uid) for uid in user_ids),
            'compare_fields': compare_fields,
            'result_conflicts': result_conflicts,
            'needs_staff_approval': needs_staff_approval,
            'province_schools': province_schools,
            'step': 2,
        }
        return render(request, 'provinces/merge_teachers.html', context)

    else:
        # STEP 1: Select primary user
        user_data = []
        for u in users:
            meta = getattr(u, 'data', None)
            user_data.append({
                'user': u,
                'school': meta.school if meta else None,
                'reg_num': meta.reg_num if meta else '',
                'mobile': meta.mobile if meta else '',
                'last_login': u.last_login,
            })

        context = {
            'province': province,
            'olympiad': olympiad,
            'users': users,
            'user_data': user_data,
            'user_ids_str': ','.join(str(uid) for uid in user_ids),
            'default_primary_id': users[0].id if users else None,
            'step': 1,
        }
        return render(request, 'provinces/merge_teachers.html', context)


@login_required
def generate_province_answer_sheet(request, province_id, olympiad_id):
    """Excel загвар үүсгэж татах"""
    province = get_object_or_404(Province, id=province_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_provinces')

    # Олимпиадад бүлэг байгаа эсэхийг шалгаад, байхгүй бол үүсгэх
    group, created = ensure_olympiad_has_group(olympiad)

    # Бүртгэлтэй сурагчид
    contestants = group.user_set.select_related('data__province', 'data__school').order_by(
        'data__school__name', 'last_name', 'first_name'
    )
    problems = olympiad.problem_set.all().order_by('order')

    # Excel Sheet 1: Хариулт
    answers_data = []
    for c in contestants:
        school_name = c.data.school.name if hasattr(c, 'data') and c.data.school else 'Тодорхойгүй'
        answers_data.append({
            'ID': c.id,
            'Овог': c.last_name,
            'Нэр': c.first_name,
            'Сургууль': school_name,
            **{f'№{p.order}': '' for p in problems}
        })

    header = ['ID', 'Овог', 'Нэр', 'Сургууль'] + [f'№{p.order}' for p in problems]
    answers_df = pd.DataFrame(answers_data, columns=header)

    # Sheet 2: Мэдээлэл
    info_data = {
        'Түлхүүр': ['olympiad_id', 'olympiad_name', 'province_id', 'province_name', 'level_id', 'level_name'],
        'Утга': [
            olympiad.id,
            olympiad.name,
            province.id,
            province.name,
            olympiad.level.id if olympiad.level else '',
            olympiad.level.name if olympiad.level else ''
        ]
    }
    info_df = pd.DataFrame(info_data)

    # Excel үүсгэх
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        answers_df.to_excel(writer, sheet_name='Хариулт', index=False)
        info_df.to_excel(writer, sheet_name='Мэдээлэл', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Хариулт']

        # Загвар
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')

        # Header
        for cell in worksheet["1:1"]:
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_align

        # Cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border

        # Column widths
        for i, column_cells in enumerate(worksheet.columns):
            column_letter = column_cells[0].column_letter
            if i < 4:
                max_length = 0
                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 30)
            else:
                worksheet.column_dimensions[column_letter].width = 5

    output.seek(0)

    filename = f"province_{province.id}_olympiad_{olympiad.id}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def import_province_answer_sheet(request, province_id, olympiad_id):
    """Excel файл уншиж дүн оруулах"""
    province = get_object_or_404(Province, id=province_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_provinces')

    if request.method == 'POST':
        from .forms import UploadExcelForm
        form = UploadExcelForm(request.POST, request.FILES)

        if form.is_valid():
            excel_file = request.FILES['excel_file']

            try:
                # Excel уншиж, баталгаажуулах
                xls = pd.ExcelFile(excel_file)

                if 'Хариулт' not in xls.sheet_names or 'Мэдээлэл' not in xls.sheet_names:
                    messages.error(request, 'Excel файлд "Хариулт" болон "Мэдээлэл" sheet байх ёстой.')
                    return redirect('province_olympiad_view', province_id=province_id, olympiad_id=olympiad_id)

                # Мэдээлэл шалгах
                info_df = pd.read_excel(excel_file, sheet_name='Мэдээлэл')
                info_dict = dict(zip(info_df['Түлхүүр'], info_df['Утга']))

                if int(info_dict.get('olympiad_id', 0)) != olympiad_id:
                    messages.error(request, 'Олимпиадын ID таарахгүй байна.')
                    return redirect('province_olympiad_view', province_id=province_id, olympiad_id=olympiad_id)

                if int(info_dict.get('province_id', 0)) != province_id:
                    messages.error(request, 'Аймгийн ID таарахгүй байна.')
                    return redirect('province_olympiad_view', province_id=province_id, olympiad_id=olympiad_id)

                # Хариулт уншиж Result үүсгэх
                answers_df = pd.read_excel(excel_file, sheet_name='Хариулт')
                problems = olympiad.problem_set.all().order_by('order')
                problem_dict = {f'№{p.order}': p for p in problems}

                created_count = 0
                updated_count = 0

                with transaction.atomic():
                    for idx, row in answers_df.iterrows():
                        user_id = row['ID']

                        try:
                            user = User.objects.get(id=user_id)

                            # Группт байгаа эсэхийг шалгах
                            if olympiad.group and not olympiad.group.user_set.filter(id=user_id).exists():
                                continue

                            # Бодлого бүрээр Result үүсгэх/шинэчлэх (оноог шууд оруулна)
                            for col_name, problem in problem_dict.items():
                                if col_name in row and pd.notna(row[col_name]):
                                    try:
                                        score = float(row[col_name])
                                    except (ValueError, TypeError):
                                        score = 0

                                    result, created = Result.objects.update_or_create(
                                        contestant=user,
                                        olympiad=olympiad,
                                        problem=problem,
                                        defaults={
                                            'score': score,
                                            'state': 2
                                        }
                                    )

                                    if created:
                                        created_count += 1
                                    else:
                                        updated_count += 1

                        except User.DoesNotExist:
                            continue

                messages.success(request, f'Амжилттай. Шинээр үүссэн: {created_count}, Шинэчлэгдсэн: {updated_count}')
                return redirect('province_olympiad_view', province_id=province_id, olympiad_id=olympiad_id)

            except Exception as e:
                messages.error(request, f'Алдаа гарлаа: {str(e)}')
                return redirect('province_olympiad_view', province_id=province_id, olympiad_id=olympiad_id)

    return redirect('province_olympiad_view', province_id=province_id, olympiad_id=olympiad_id)


@login_required
def view_province_participants(request, province_id, olympiad_id):
    """Олимпиадад эрх авсан багш, сурагчдын жагсаалт"""
    province = get_object_or_404(Province, id=province_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_provinces')

    if olympiad.round != 2:
        messages.error(request, 'Энэ олимпиад 2-р давааны олимпиад биш байна.')
        return redirect('province_dashboard', province_id=province_id)

    # Олимпиадад бүлэг байгаа эсэхийг шалгаад, байхгүй бол үүсгэх
    group, created = ensure_olympiad_has_group(olympiad)

    # Round 1 олимпиадуудыг олох (энэ Round 2 олимпиадыг зааж байгаа)
    round1_olympiads = Olympiad.objects.filter(next_round=olympiad)

    # Энэ аймгийн бүртгэлтэй сурагчид
    participants = group.user_set.filter(
        data__province=province
    ).select_related('data__school', 'data__grade', 'data__province').order_by(
        'data__school__name', 'last_name', 'first_name'
    )

    # POST үйлдлүүд
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'remove':
            remove_ids = request.POST.getlist('remove_users')
            if remove_ids:
                try:
                    remove_ids = [int(uid) for uid in remove_ids]
                except ValueError:
                    remove_ids = []
                if remove_ids:
                    users_to_remove = User.objects.filter(id__in=remove_ids)
                    count = users_to_remove.count()
                    group.user_set.remove(*users_to_remove)
                    messages.success(request, f'{count} оролцогч хасагдлаа.')

        elif action == 'add_qualified':
            qualified_awards = Award.objects.filter(
                olympiad__in=round1_olympiads,
                place__startswith='2.1 эрх',
                contestant__data__province=province,
            ).select_related('contestant')

            added_count = 0
            already_count = 0
            for award in qualified_awards:
                user = award.contestant
                if group.user_set.filter(id=user.id).exists():
                    already_count += 1
                else:
                    group.user_set.add(user)
                    added_count += 1

            if added_count:
                messages.success(request, f'{added_count} сурагч нэмэгдлээ.')
            if already_count:
                messages.info(request, f'{already_count} сурагч аль хэдийн бүртгэлтэй байсан.')
            if not added_count and not already_count:
                messages.warning(request, '1-р шатнаас 2.1 эрх авсан сурагч олдсонгүй.')

        return redirect('view_province_participants', province_id=province_id, olympiad_id=olympiad_id)

    # Round 1 мэдээлэл нэмэх
    participants_list = []
    for user in participants:
        award = Award.objects.filter(
            olympiad__in=round1_olympiads,
            contestant=user
        ).first()

        scoresheet = ScoreSheet.objects.filter(
            olympiad__in=round1_olympiads,
            user=user,
            is_official=True
        ).order_by('-total').first()

        participants_list.append({
            'user': user,
            'school': user.data.school if hasattr(user, 'data') and user.data else None,
            'grade': user.data.grade if hasattr(user, 'data') and user.data else None,
            'award_place': award.place if award else 'Тодорхойгүй',
            'round1_score': scoresheet.total if scoresheet else 0,
        })

    context = {
        'province': province,
        'olympiad': olympiad,
        'participants_list': participants_list,
        'total_count': len(participants_list),
    }
    return render(request, 'provinces/province_participants.html', context)


@login_required
def view_province_olympiad_results(request, province_id, olympiad_id):
    """Үр дүн харах"""
    province = get_object_or_404(Province, id=province_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_provinces')

    # Олимпиадад бүлэг байгаа эсэхийг шалгаад, байхгүй бол үүсгэх
    group, created = ensure_olympiad_has_group(olympiad)

    # Бүртгэлтэй сурагчдын Result-үүд
    group_user_ids = group.user_set.values_list('id', flat=True)
    results = Result.objects.filter(
        olympiad=olympiad,
        contestant_id__in=group_user_ids
    ).select_related('contestant', 'problem', 'contestant__data__school')

    # Pivot table үүсгэх
    students_scores = {}
    problems = olympiad.problem_set.all().order_by('order')

    for result in results:
        user_id = result.contestant.id
        if user_id not in students_scores:
            students_scores[user_id] = {
                'user': result.contestant,
                'school': result.contestant.data.school if hasattr(result.contestant, 'data') else None,
                'scores': {},
                'total': 0
            }

        students_scores[user_id]['scores'][result.problem.order] = result.score
        students_scores[user_id]['total'] += result.score or 0

    # Эрэмбэлэх (total буурах, нэр өсөх)
    students_list = sorted(
        students_scores.values(),
        key=lambda x: (-x['total'], x['user'].last_name)
    )

    context = {
        'province': province,
        'olympiad': olympiad,
        'students': students_list,
        'problems': problems,
    }
    return render(request, 'provinces/province_olympiad_results.html', context)


@login_required
def view_round1_province_results(request, province_id, olympiad_id):
    """Round 1-ийн аймгийн дүнг харах (Round 2 олимпиадаас)"""
    province = get_object_or_404(Province, id=province_id)
    round2_olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_province(request.user, province):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_provinces')

    # Round 1 олимпиадуудыг олох
    round1_olympiads = Olympiad.objects.filter(next_round=round2_olympiad)

    if not round1_olympiads.exists():
        messages.warning(request, 'Round 1 олимпиад олдсонгүй.')
        return redirect('province_olympiad_view', province_id=province_id, olympiad_id=olympiad_id)

    # Аймгийн Round 1 оролцогчдын ScoreSheet-үүд
    scoresheets = ScoreSheet.objects.filter(
        olympiad__in=round1_olympiads,
        user__data__province=province,
        is_official=True
    ).select_related('user', 'user__data__school', 'olympiad').order_by('-total', 'user__last_name')

    # ScoreSheet-үүдээс сурагчдын мэдээллийг бэлтгэх
    # Award-аар 2-р давааны эрхийн мэдээллийг олох
    students_list = []
    for scoresheet in scoresheets:
        # 2-р давааны эрх шалгах
        award = Award.objects.filter(
            olympiad=round2_olympiad,
            contestant=scoresheet.user
        ).first()

        award_status = award.place if award else None

        students_list.append({
            'user': scoresheet.user,
            'school': scoresheet.user.data.school if hasattr(scoresheet.user, 'data') else None,
            'total': scoresheet.total,
            'olympiad': scoresheet.olympiad,
            'award_status': award_status,
        })

    context = {
        'province': province,
        'round2_olympiad': round2_olympiad,
        'round1_olympiads': round1_olympiads,
        'students': students_list,
    }
    return render(request, 'provinces/round1_province_results.html', context)


# =============================================================================
# Zone (Бүс) level views - Round 3
# =============================================================================

def user_can_manage_zone(user, zone):
    """Хэрэглэгч бүсийг удирдах эрхтэй эсэхийг шалгах"""
    if user.is_staff:
        return True

    if zone.contact_person == user:
        return True

    group_name = f"Zone_{zone.id}_Managers"
    if user.groups.filter(name=group_name).exists():
        return True

    return False


@login_required
def my_managed_zones(request):
    """Удирдаж байгаа бүсүүдийн жагсаалт"""
    is_staff_access = False

    if request.user.is_staff:
        # Staff бол бүх бүсийг харуулна
        managed_zones = Zone.objects.all().select_related('contact_person').order_by('name')
        is_staff_access = True
    else:
        zone_ids = []
        user_groups = request.user.groups.filter(name__startswith='Zone_', name__endswith='_Managers')
        for group in user_groups:
            try:
                zone_id = int(group.name.replace('Zone_', '').replace('_Managers', ''))
                zone_ids.append(zone_id)
            except (ValueError, AttributeError):
                continue

        managed_zones = Zone.objects.filter(
            Q(contact_person=request.user) | Q(id__in=zone_ids)
        ).select_related('contact_person').distinct().order_by('name')

    context = {
        'zones': managed_zones,
        'is_staff_access': is_staff_access,
    }
    return render(request, 'provinces/my_managed_zones.html', context)


@staff_member_required
def zone_change_contact(request, zone_id):
    """Бүсийн удирдах хүнийг солих (staff only)"""
    zone = get_object_or_404(Zone, id=zone_id)

    if request.method == 'POST':
        user_id = request.POST.get('contact_person_id', '').strip()

        if not user_id:
            zone.contact_person = None
            zone.save(update_fields=['contact_person'])
            messages.success(request, f'"{zone.name}" бүсийн удирдах хүнийг хаслаа.')
        else:
            try:
                user = User.objects.get(id=int(user_id))
                zone.contact_person = user
                zone.save(update_fields=['contact_person'])
                messages.success(request, f'"{zone.name}" бүсийн удирдах хүнийг {user.last_name} {user.first_name} (ID: {user.id}) болгож солилоо.')
            except (ValueError, User.DoesNotExist):
                messages.error(request, f'ID={user_id} хэрэглэгч олдсонгүй.')

    return redirect('zone_dashboard', zone_id=zone_id)


@login_required
def zone_select_olympiads(request, zone_id):
    """Бүсийн оролцох олимпиадуудыг сонгох"""
    zone = get_object_or_404(Zone, id=zone_id)

    if not user_can_manage_zone(request.user, zone):
        messages.error(request, 'Та энэ бүсийг удирдах эрхгүй.')
        return redirect('my_managed_zones')

    current_school_year = SchoolYear.get_current()

    # Энэ жилийн бүх Round 3 олимпиадууд
    all_round3 = Olympiad.objects.filter(
        round=3,
        school_year=current_school_year
    ).select_related('level', 'school_year').order_by('level__name')

    # Аль хэдийн сонгогдсон олимпиадуудын ID
    selected_ids = set(zone.olympiads.values_list('id', flat=True))

    if request.method == 'POST':
        chosen_ids = request.POST.getlist('olympiad_ids')
        chosen_ids = [int(x) for x in chosen_ids if x.isdigit()]

        # Зөвхөн энэ жилийн round=3 олимпиадуудаас сонгохыг баталгаажуулах
        valid_ids = set(all_round3.values_list('id', flat=True))
        chosen_ids = [oid for oid in chosen_ids if oid in valid_ids]

        zone.olympiads.set(chosen_ids)
        messages.success(request, f'{len(chosen_ids)} олимпиад сонгогдлоо.')
        return redirect('zone_dashboard', zone_id=zone_id)

    # Олимпиад бүрт сонгогдсон эсэхийг тэмдэглэх
    for olympiad in all_round3:
        olympiad.is_selected = olympiad.id in selected_ids

    context = {
        'zone': zone,
        'all_round3': all_round3,
        'current_school_year': current_school_year,
    }
    return render(request, 'provinces/zone_select_olympiads.html', context)


@login_required
def zone_dashboard(request, zone_id):
    """Бүсийн удирдлагын самбар - Round 3 олимпиадуудыг харуулна"""
    zone = get_object_or_404(Zone, id=zone_id)

    if not user_can_manage_zone(request.user, zone):
        messages.error(request, 'Та энэ бүсийг удирдах эрхгүй.')
        return redirect('my_managed_zones')

    current_school_year = SchoolYear.get_current()

    # Бүсийн сонгосон Round 3 олимпиадууд (энэ жилийн)
    round3_olympiads = zone.olympiads.filter(
        round=3,
        school_year=current_school_year
    ).select_related('level', 'school_year', 'group').order_by('level__name')

    # Бүлэг байхгүй олимпиадад автоматаар бүлэг үүсгэх
    for olympiad in round3_olympiads:
        group, created = ensure_olympiad_has_group(olympiad, group_name_template="Round3_Olympiad_{olympiad.id}")
        if created:
            messages.info(request, f'"{olympiad.name}" олимпиадад "{group.name}" бүлэг автоматаар үүсгэгдлээ.')

    from accounts.models import UserMeta

    # Zone доторх Province-уудын ID (zone_id=5 бол бүгд)
    if zone.id == 5:
        zone_province_ids = None
    else:
        zone_province_ids = Province.objects.filter(zone=zone).values_list('id', flat=True)

    for olympiad in round3_olympiads:
        is_teacher = olympiad.level and (olympiad.level.name.startswith('S') or olympiad.level.name.startswith('T'))
        olympiad.is_teacher_olympiad = is_teacher

        if is_teacher:
            category_prefix = olympiad.level.name[0]

            if olympiad.group:
                qs = olympiad.group.user_set.filter(data__level__name__startswith=category_prefix)
                if zone_province_ids is not None:
                    qs = qs.filter(data__province_id__in=zone_province_ids)
                olympiad.zone_registered = qs.count()
            else:
                olympiad.zone_registered = 0

            qs = User.objects.filter(data__level__name__startswith=category_prefix)
            if zone_province_ids is not None:
                qs = qs.filter(data__province_id__in=zone_province_ids)
            olympiad.zone_total = qs.count()
        else:
            if olympiad.group:
                qs = olympiad.group.user_set.all()
                if zone_province_ids is not None:
                    qs = qs.filter(data__province_id__in=zone_province_ids)
                olympiad.zone_registered = qs.count()
            else:
                olympiad.zone_registered = 0

            # Round 2 олимпиадууд (энэ Round 3-г зааж байгаа)
            round2_olympiads = Olympiad.objects.filter(next_round=olympiad)

            qs = ScoreSheet.objects.filter(
                olympiad__in=round2_olympiads,
                is_official=True
            )
            if zone_province_ids is not None:
                qs = qs.filter(user__data__province_id__in=zone_province_ids)
            olympiad.zone_total = qs.values('user').distinct().count()

    teacher_olympiads = [o for o in round3_olympiads if o.is_teacher_olympiad]
    student_olympiads = [o for o in round3_olympiads if not o.is_teacher_olympiad]

    teacher_levels = sorted(set(o.level.name for o in teacher_olympiads if o.level))
    student_levels = sorted(set(o.level.name for o in student_olympiads if o.level))

    # Zone доторх аймгуудын жагсаалт (zone_id=5 бол бүгд)
    if zone.id == 5:
        zone_provinces = Province.objects.all().order_by('name')
    else:
        zone_provinces = Province.objects.filter(zone=zone).order_by('name')

    context = {
        'zone': zone,
        'teacher_olympiads': teacher_olympiads,
        'student_olympiads': student_olympiads,
        'teacher_levels': ', '.join(teacher_levels),
        'student_levels': ', '.join(student_levels),
        'current_school_year': current_school_year,
        'zone_provinces': zone_provinces,
    }
    return render(request, 'provinces/zone_dashboard.html', context)


@login_required
def zone_olympiad_view(request, zone_id, olympiad_id):
    """Олимпиад удирдах үндсэн хуудас - 3 үндсэн үйлдэл (Zone level)"""
    zone = get_object_or_404(Zone, id=zone_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_zone(request.user, zone):
        messages.error(request, 'Та энэ бүсийг удирдах эрхгүй.')
        return redirect('my_managed_zones')

    if olympiad.round != 3:
        messages.error(request, 'Энэ олимпиад 3-р давааны олимпиад биш байна.')
        return redirect('zone_dashboard', zone_id=zone_id)

    if not zone.olympiads.filter(id=olympiad.id).exists():
        messages.error(request, 'Энэ олимпиад таны бүсийн сонгосон олимпиадад ороогүй байна.')
        return redirect('zone_dashboard', zone_id=zone_id)

    group, created = ensure_olympiad_has_group(olympiad, group_name_template="Round3_Olympiad_{olympiad.id}")
    if created:
        messages.info(request, f'Олимпиадад "{group.name}" бүлэг автоматаар үүсгэгдлээ.')

    is_teacher_olympiad = olympiad.level and (olympiad.level.name.startswith('S') or olympiad.level.name.startswith('T'))

    total_registered = olympiad.group.user_set.count() if olympiad.group else 0

    if zone.id == 5:
        zone_province_ids = None
    else:
        zone_province_ids = Province.objects.filter(zone=zone).values_list('id', flat=True)

    zone_registered = 0
    if olympiad.group:
        qs = olympiad.group.user_set.all()
        if zone_province_ids is not None:
            qs = qs.filter(data__province_id__in=zone_province_ids)
        zone_registered = qs.count()

    round2_olympiads = Olympiad.objects.filter(next_round=olympiad)
    if not round2_olympiads.exists():
        round2_olympiads = Olympiad.objects.filter(
            school_year=olympiad.school_year,
            level=olympiad.level,
            round=2,
        )

    qs = ScoreSheet.objects.filter(
        olympiad__in=round2_olympiads,
        is_official=True
    )
    if zone_province_ids is not None:
        qs = qs.filter(user__data__province_id__in=zone_province_ids)
    round2_participants = qs.values('user').distinct().count()

    qs = Award.objects.filter(
        olympiad__in=round2_olympiads,
        place__startswith='2.2 эрх',
    )
    if zone_province_ids is not None:
        qs = qs.filter(contestant__data__province_id__in=zone_province_ids)
    auto_registered = qs.count()

    context = {
        'zone': zone,
        'olympiad': olympiad,
        'total_registered': total_registered,
        'zone_registered': zone_registered,
        'round2_participants': round2_participants,
        'auto_registered_count': auto_registered,
    }
    return render(request, 'provinces/zone_olympiad_detail.html', context)


@login_required
def zone_add_all_qualified(request, zone_id, olympiad_id):
    """2.1 эрх авсан тухайн бүсийн бүх сурагчдыг Round 3 олимпиадад нэмэх"""
    zone = get_object_or_404(Zone, id=zone_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_zone(request.user, zone):
        messages.error(request, 'Та энэ бүсийг удирдах эрхгүй.')
        return redirect('my_managed_zones')

    if request.method != 'POST':
        return redirect('zone_olympiad_view', zone_id=zone_id, olympiad_id=olympiad_id)

    group, _ = ensure_olympiad_has_group(olympiad, group_name_template="Round3_Olympiad_{olympiad.id}")

    # Round 2 олимпиадуудаас 2.2 эрх авсан сурагчдыг олох
    round2_olympiads = Olympiad.objects.filter(next_round=olympiad)
    if not round2_olympiads.exists():
        round2_olympiads = Olympiad.objects.filter(
            school_year=olympiad.school_year,
            level=olympiad.level,
            round=2,
        )

    zone_province_ids = Province.objects.filter(zone=zone).values_list('id', flat=True)
    qualified_awards = Award.objects.filter(
        olympiad__in=round2_olympiads,
        place__startswith='2.2 эрх',
        contestant__data__province_id__in=zone_province_ids,
    ).select_related('contestant')

    added_count = 0
    already_count = 0
    for award in qualified_awards:
        user = award.contestant
        if group.user_set.filter(id=user.id).exists():
            already_count += 1
        else:
            group.user_set.add(user)
            added_count += 1

    if added_count:
        messages.success(request, f'{added_count} сурагчийг амжилттай нэмлээ.')
    if already_count:
        messages.info(request, f'{already_count} сурагч аль хэдийн бүртгэлтэй байсан.')
    if not added_count and not already_count:
        messages.warning(request, '2.2 эрх авсан сурагч олдсонгүй.')

    return redirect('zone_olympiad_view', zone_id=zone_id, olympiad_id=olympiad_id)


@login_required
def zone_add_students_by_threshold(request, zone_id, olympiad_id):
    """Онооны босгоор нэмэлт сурагч бүртгэх (Zone level)"""
    zone = get_object_or_404(Zone, id=zone_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_zone(request.user, zone):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_zones')

    if request.method == 'POST':
        from .forms import ThresholdScoreForm
        form = ThresholdScoreForm(request.POST)

        if form.is_valid():
            threshold = form.cleaned_data['threshold_score']

            group, created = ensure_olympiad_has_group(olympiad, group_name_template="Round3_Olympiad_{olympiad.id}")

            # Round 2 олимпиадуудыг олох (энэ Round 3-г зааж байгаа)
            round2_olympiads = Olympiad.objects.filter(next_round=olympiad)

            eligible_scoresheets = ScoreSheet.objects.filter(
                olympiad__in=round2_olympiads,
                total__gte=threshold,
                is_official=True,
            )
            if zone.id != 5:
                zone_province_ids = Province.objects.filter(zone=zone).values_list('id', flat=True)
                eligible_scoresheets = eligible_scoresheets.filter(user__data__province_id__in=zone_province_ids)
            eligible_scoresheets = eligible_scoresheets.select_related('user').exclude(
                user__groups=group
            )

            added_count = 0
            with transaction.atomic():
                for scoresheet in eligible_scoresheets:
                    Award.objects.get_or_create(
                        olympiad=scoresheet.olympiad,
                        contestant=scoresheet.user,
                        defaults={'place': '3.1 эрх нэмэлтээр (босго)'}
                    )
                    group.user_set.add(scoresheet.user)
                    added_count += 1

            messages.success(request, f'{added_count} сурагч нэмэгдлээ.')
            return redirect('zone_olympiad_view', zone_id=zone_id, olympiad_id=olympiad_id)

    else:
        from .forms import ThresholdScoreForm
        form = ThresholdScoreForm()

    context = {
        'zone': zone,
        'olympiad': olympiad,
        'form': form,
    }
    return render(request, 'provinces/zone_add_by_threshold.html', context)


@login_required
def zone_select_top_students(request, zone_id, olympiad_id):
    """Аймаг бүрээс топ 10 (+ ижил оноотой) сурагчдыг харуулж, бүртгэх (Zone level)"""
    zone = get_object_or_404(Zone, id=zone_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_zone(request.user, zone):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_zones')

    if zone.id == 5:
        zone_provinces = Province.objects.all().order_by('name')
    else:
        zone_provinces = Province.objects.filter(zone=zone).order_by('name')

    # Round 2 олимпиадууд
    round2_olympiads = Olympiad.objects.filter(next_round=olympiad)
    if not round2_olympiads.exists():
        round2_olympiads = Olympiad.objects.filter(
            school_year=olympiad.school_year,
            level=olympiad.level,
            round=2,
        )

    # Round 3 группт бүртгэлтэй хэрэглэгчдийн ID
    group, _ = ensure_olympiad_has_group(olympiad, group_name_template="Round3_Olympiad_{olympiad.id}")
    registered_user_ids = set(group.user_set.values_list('id', flat=True))

    # 2.2 эрх авсан хэрэглэгчдийн ID
    qualified_user_ids = set(Award.objects.filter(
        olympiad__in=round2_olympiads,
        place__startswith='2.2 эрх',
    ).values_list('contestant_id', flat=True))

    # Аймаг бүрээр бүлэглэж харуулах
    provinces_data = []
    for province in zone_provinces:
        all_province_scoresheets = ScoreSheet.objects.filter(
            olympiad__in=round2_olympiads,
            user__data__province=province,
            is_official=True
        ).select_related('user', 'user__data__school').order_by('-total')

        if not all_province_scoresheets.exists():
            continue

        # Эрх авсан сурагчид + эрх аваагүй сурагчдаас топ 10 (+ тэнцсэн)
        qualified_ss = []
        unqualified_ss = []
        for ss in all_province_scoresheets:
            ss.is_registered = ss.user_id in registered_user_ids
            ss.is_qualified = ss.user_id in qualified_user_ids
            if ss.is_qualified:
                qualified_ss.append(ss)
            else:
                unqualified_ss.append(ss)

        # Эрх аваагүй, 0-ээс дээш оноотой сурагчдаас топ 10 + тэнцсэн
        next_students = []
        cutoff_score = 0
        unqualified_nonzero = [ss for ss in unqualified_ss if (ss.total or 0) > 0]
        if unqualified_nonzero:
            top_10_unqualified = unqualified_nonzero[:10]
            cutoff_score = top_10_unqualified[-1].total or 0
            for ss in unqualified_nonzero:
                if (ss.total or 0) >= cutoff_score:
                    ss.is_tied = ss not in top_10_unqualified
                    next_students.append(ss)

        if not qualified_ss and not next_students:
            continue

        provinces_data.append({
            'province': province,
            'qualified': qualified_ss,
            'next_students': next_students,
            'cutoff_score': cutoff_score,
            'qualified_count': len(qualified_ss),
            'next_count': len(next_students),
        })

    # POST: сонгосон сурагчдыг бүртгэх
    if request.method == 'POST':
        selected_user_ids = request.POST.getlist('selected_users')

        added_count = 0
        already_count = 0
        with transaction.atomic():
            for user_id in selected_user_ids:
                try:
                    user = User.objects.get(id=int(user_id))
                    if group.user_set.filter(id=user.id).exists():
                        already_count += 1
                    else:
                        group.user_set.add(user)
                        added_count += 1
                except (ValueError, User.DoesNotExist):
                    pass

        if added_count:
            messages.success(request, f'{added_count} сурагч нэмэгдлээ.')
        if already_count:
            messages.info(request, f'{already_count} сурагч аль хэдийн бүртгэлтэй байсан.')
        return redirect('zone_select_top_students', zone_id=zone_id, olympiad_id=olympiad_id)

    context = {
        'zone': zone,
        'olympiad': olympiad,
        'provinces_data': provinces_data,
    }
    return render(request, 'provinces/zone_select_top_students.html', context)


@login_required
def zone_add_students_by_id(request, zone_id, olympiad_id):
    """ID-гаар сурагч/багш нэмэх (Zone level)"""
    zone = get_object_or_404(Zone, id=zone_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_zone(request.user, zone):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_zones')

    is_teacher_olympiad = olympiad.level and (olympiad.level.name.startswith('S') or olympiad.level.name.startswith('T'))
    zone_province_ids = set(Province.objects.filter(zone=zone).values_list('id', flat=True))

    if request.method == 'POST':
        user_ids_input = request.POST.get('user_ids', '').strip()

        if not user_ids_input:
            messages.error(request, 'Хэрэглэгчийн ID оруулна уу.')
            return redirect('zone_add_students_by_id', zone_id=zone_id, olympiad_id=olympiad_id)

        import re
        user_ids = re.split(r'[,\s\n]+', user_ids_input)
        user_ids = [uid.strip() for uid in user_ids if uid.strip()]

        try:
            user_ids = [int(uid) for uid in user_ids]
        except ValueError:
            messages.error(request, 'ID буруу байна. Зөвхөн тоон утга оруулна уу.')
            return redirect('zone_add_students_by_id', zone_id=zone_id, olympiad_id=olympiad_id)

        group, created = ensure_olympiad_has_group(olympiad, group_name_template="Round3_Olympiad_{olympiad.id}")

        users = User.objects.filter(id__in=user_ids).select_related('data', 'data__province')

        found_ids = set(users.values_list('id', flat=True))
        not_found_ids = set(user_ids) - found_ids

        if not_found_ids:
            messages.warning(request, f'Олдоогүй ID: {", ".join(map(str, not_found_ids))}')

        added_count = 0
        already_in_group = 0
        other_zone_users = []

        if is_teacher_olympiad:
            with transaction.atomic():
                for user in users:
                    if group.user_set.filter(id=user.id).exists():
                        already_in_group += 1
                        continue

                    Award.objects.get_or_create(
                        olympiad=olympiad,
                        contestant=user,
                        defaults={'place': '3.1 эрх (ID)'}
                    )

                    group.user_set.add(user)
                    added_count += 1

            if added_count > 0:
                messages.success(request, f'{added_count} багш нэмэгдлээ.')
            if already_in_group > 0:
                messages.info(request, f'{already_in_group} багш аль хэдийн группт байна.')
        else:
            round2_olympiads = Olympiad.objects.filter(next_round=olympiad)

            with transaction.atomic():
                for user in users:
                    if group.user_set.filter(id=user.id).exists():
                        already_in_group += 1
                        continue

                    # Өөр бүсийн сурагч бол анхааруулга (zone_id=5 үед шалгахгүй)
                    if zone.id != 5 and hasattr(user, 'data') and user.data and user.data.province_id not in zone_province_ids:
                        province_name = user.data.province.name if user.data.province else '?'
                        other_zone_users.append(f'{user.id} ({user.last_name} {user.first_name}, {province_name})')

                    scoresheet = ScoreSheet.objects.filter(
                        olympiad__in=round2_olympiads,
                        user=user,
                        is_official=True
                    ).first()

                    if scoresheet:
                        award_olympiad = scoresheet.olympiad
                    else:
                        award_olympiad = round2_olympiads.first() if round2_olympiads.exists() else olympiad

                    Award.objects.get_or_create(
                        olympiad=award_olympiad,
                        contestant=user,
                        defaults={'place': '3.1 эрх нэмэлтээр (ID)'}
                    )

                    group.user_set.add(user)
                    added_count += 1

            if added_count > 0:
                messages.success(request, f'{added_count} сурагч нэмэгдлээ.')
            if already_in_group > 0:
                messages.info(request, f'{already_in_group} сурагч аль хэдийн группт байна.')
            if other_zone_users:
                messages.warning(request, f'Өөр бүсийн {len(other_zone_users)} сурагч нэмэгдлээ: {", ".join(other_zone_users)}')

        return redirect('zone_olympiad_view', zone_id=zone_id, olympiad_id=olympiad_id)

    context = {
        'zone': zone,
        'olympiad': olympiad,
        'is_teacher_olympiad': is_teacher_olympiad,
    }
    return render(request, 'provinces/zone_add_students_by_id.html', context)


@login_required
def zone_generate_answer_sheet(request, zone_id, olympiad_id):
    """Excel загвар үүсгэж татах (Zone level)"""
    zone = get_object_or_404(Zone, id=zone_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_zone(request.user, zone):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_zones')

    group, created = ensure_olympiad_has_group(olympiad, group_name_template="Round3_Olympiad_{olympiad.id}")

    if zone.id == 5:
        contestants = group.user_set.all()
    else:
        zone_province_ids = Province.objects.filter(zone=zone).values_list('id', flat=True)
        contestants = group.user_set.filter(data__province_id__in=zone_province_ids)
    contestants = contestants.select_related('data__province', 'data__school').order_by(
        'data__province__name', 'data__school__name', 'last_name', 'first_name'
    )
    problems = olympiad.problem_set.all().order_by('order')

    answers_data = []
    for c in contestants:
        school_name = c.data.school.name if hasattr(c, 'data') and c.data.school else 'Тодорхойгүй'
        province_name = c.data.province.name if hasattr(c, 'data') and c.data.province else 'Тодорхойгүй'
        answers_data.append({
            'ID': c.id,
            'Овог': c.last_name,
            'Нэр': c.first_name,
            'Аймаг': province_name,
            'Сургууль': school_name,
            **{f'№{p.order}': '' for p in problems}
        })

    header = ['ID', 'Овог', 'Нэр', 'Аймаг', 'Сургууль'] + [f'№{p.order}' for p in problems]
    answers_df = pd.DataFrame(answers_data, columns=header)

    info_data = {
        'Түлхүүр': ['olympiad_id', 'olympiad_name', 'zone_id', 'zone_name', 'level_id', 'level_name'],
        'Утга': [
            olympiad.id,
            olympiad.name,
            zone.id,
            zone.name,
            olympiad.level.id if olympiad.level else '',
            olympiad.level.name if olympiad.level else ''
        ]
    }
    info_df = pd.DataFrame(info_data)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        answers_df.to_excel(writer, sheet_name='Хариулт', index=False)
        info_df.to_excel(writer, sheet_name='Мэдээлэл', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Хариулт']

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')

        for cell in worksheet["1:1"]:
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_align

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border

        for i, column_cells in enumerate(worksheet.columns):
            column_letter = column_cells[0].column_letter
            if i < 5:
                max_length = 0
                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 30)
            else:
                worksheet.column_dimensions[column_letter].width = 5

    output.seek(0)

    filename = f"zone_{zone.id}_olympiad_{olympiad.id}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def zone_import_answer_sheet(request, zone_id, olympiad_id):
    """Excel файл уншиж дүн оруулах (Zone level)"""
    zone = get_object_or_404(Zone, id=zone_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_zone(request.user, zone):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_zones')

    if request.method == 'POST':
        from .forms import UploadExcelForm
        form = UploadExcelForm(request.POST, request.FILES)

        if form.is_valid():
            excel_file = request.FILES['excel_file']

            try:
                xls = pd.ExcelFile(excel_file)

                if 'Хариулт' not in xls.sheet_names or 'Мэдээлэл' not in xls.sheet_names:
                    messages.error(request, 'Excel файлд "Хариулт" болон "Мэдээлэл" sheet байх ёстой.')
                    return redirect('zone_olympiad_view', zone_id=zone_id, olympiad_id=olympiad_id)

                info_df = pd.read_excel(excel_file, sheet_name='Мэдээлэл')
                info_dict = dict(zip(info_df['Түлхүүр'], info_df['Утга']))

                if int(info_dict.get('olympiad_id', 0)) != olympiad_id:
                    messages.error(request, 'Олимпиадын ID таарахгүй байна.')
                    return redirect('zone_olympiad_view', zone_id=zone_id, olympiad_id=olympiad_id)

                if int(info_dict.get('zone_id', 0)) != zone_id:
                    messages.error(request, 'Бүсийн ID таарахгүй байна.')
                    return redirect('zone_olympiad_view', zone_id=zone_id, olympiad_id=olympiad_id)

                answers_df = pd.read_excel(excel_file, sheet_name='Хариулт')
                problems = olympiad.problem_set.all().order_by('order')
                problem_dict = {f'№{p.order}': p for p in problems}

                created_count = 0
                updated_count = 0

                with transaction.atomic():
                    for idx, row in answers_df.iterrows():
                        user_id = row['ID']

                        try:
                            user = User.objects.get(id=user_id)

                            if olympiad.group and not olympiad.group.user_set.filter(id=user_id).exists():
                                continue

                            for col_name, problem in problem_dict.items():
                                if col_name in row and pd.notna(row[col_name]):
                                    answer_value = row[col_name]

                                    try:
                                        score = float(answer_value)
                                    except (ValueError, TypeError):
                                        score = 0

                                    result, created = Result.objects.update_or_create(
                                        contestant=user,
                                        olympiad=olympiad,
                                        problem=problem,
                                        defaults={
                                            'score': score,
                                            'state': 2
                                        }
                                    )

                                    if created:
                                        created_count += 1
                                    else:
                                        updated_count += 1

                        except User.DoesNotExist:
                            continue

                messages.success(request, f'Амжилттай. Шинээр үүссэн: {created_count}, Шинэчлэгдсэн: {updated_count}')
                return redirect('zone_olympiad_view', zone_id=zone_id, olympiad_id=olympiad_id)

            except Exception as e:
                messages.error(request, f'Алдаа гарлаа: {str(e)}')
                return redirect('zone_olympiad_view', zone_id=zone_id, olympiad_id=olympiad_id)

    return redirect('zone_olympiad_view', zone_id=zone_id, olympiad_id=olympiad_id)


@login_required
def zone_view_participants(request, zone_id, olympiad_id):
    """Олимпиадад эрх авсан багш, сурагчдын жагсаалт (Zone level)"""
    zone = get_object_or_404(Zone, id=zone_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_zone(request.user, zone):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_zones')

    if olympiad.round != 3:
        messages.error(request, 'Энэ олимпиад 3-р давааны олимпиад биш байна.')
        return redirect('zone_dashboard', zone_id=zone_id)

    group, created = ensure_olympiad_has_group(olympiad, group_name_template="Round3_Olympiad_{olympiad.id}")

    # POST: сонгосон сурагчдыг хасах
    if request.method == 'POST':
        remove_ids = request.POST.getlist('remove_users')
        if remove_ids:
            try:
                remove_ids = [int(uid) for uid in remove_ids]
            except ValueError:
                remove_ids = []
            if remove_ids:
                users_to_remove = User.objects.filter(id__in=remove_ids)
                group.user_set.remove(*users_to_remove)
                messages.success(request, f'{users_to_remove.count()} оролцогч хасагдлаа.')
        return redirect('zone_view_participants', zone_id=zone_id, olympiad_id=olympiad_id)

    # Round 2 олимпиадуудыг олох
    round2_olympiads = Olympiad.objects.filter(next_round=olympiad)
    if not round2_olympiads.exists():
        round2_olympiads = Olympiad.objects.filter(
            school_year=olympiad.school_year,
            level=olympiad.level,
            round=2,
        )

    # zone_id=5 бол бүх хэрэглэгч, бусад үед зөвхөн тухайн бүсийн
    if zone.id == 5:
        participants = group.user_set.all()
    else:
        zone_province_ids = Province.objects.filter(zone=zone).values_list('id', flat=True)
        participants = group.user_set.filter(data__province_id__in=zone_province_ids)
    participants = participants.select_related(
        'data__school', 'data__grade', 'data__province'
    ).order_by('data__province__name', 'last_name', 'first_name')

    # Round 2 оноог нэг query-гаар авах
    round2_scores = {}
    for ss in ScoreSheet.objects.filter(
        olympiad__in=round2_olympiads,
        user__in=participants,
        is_official=True
    ).values('user_id', 'total'):
        uid = ss['user_id']
        if uid not in round2_scores or (ss['total'] or 0) > round2_scores[uid]:
            round2_scores[uid] = ss['total'] or 0

    all_students = []
    for user in participants:
        user.province_name = user.data.province.name if hasattr(user, 'data') and user.data.province else 'Тодорхойгүй'
        user.round2_score = round2_scores.get(user.id, 0)
        user.school_name = user.data.school.name if hasattr(user, 'data') and user.data.school else '-'
        all_students.append(user)

    all_students.sort(key=lambda u: (u.province_name, -u.round2_score))

    context = {
        'zone': zone,
        'olympiad': olympiad,
        'students': all_students,
        'total_count': participants.count(),
    }
    return render(request, 'provinces/zone_participants.html', context)


@login_required
def zone_view_results(request, zone_id, olympiad_id):
    """Үр дүн харах (Zone level)"""
    zone = get_object_or_404(Zone, id=zone_id)
    olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_zone(request.user, zone):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_zones')

    group, created = ensure_olympiad_has_group(olympiad, group_name_template="Round3_Olympiad_{olympiad.id}")

    if zone.id == 5:
        group_user_ids = group.user_set.values_list('id', flat=True)
    else:
        zone_province_ids = Province.objects.filter(zone=zone).values_list('id', flat=True)
        group_user_ids = group.user_set.filter(
            data__province_id__in=zone_province_ids
        ).values_list('id', flat=True)

    results = Result.objects.filter(
        olympiad=olympiad,
        contestant_id__in=group_user_ids
    ).select_related('contestant', 'problem', 'contestant__data__school', 'contestant__data__province')

    students_scores = {}
    problems = olympiad.problem_set.all().order_by('order')

    for result in results:
        user_id = result.contestant.id
        if user_id not in students_scores:
            students_scores[user_id] = {
                'user': result.contestant,
                'school': result.contestant.data.school if hasattr(result.contestant, 'data') else None,
                'province': result.contestant.data.province if hasattr(result.contestant, 'data') else None,
                'scores': {},
                'total': 0
            }

        students_scores[user_id]['scores'][result.problem.order] = result.score
        students_scores[user_id]['total'] += result.score or 0

    students_list = sorted(
        students_scores.values(),
        key=lambda x: (-x['total'], x['user'].last_name)
    )

    context = {
        'zone': zone,
        'olympiad': olympiad,
        'students': students_list,
        'problems': problems,
    }
    return render(request, 'provinces/zone_olympiad_results.html', context)


@login_required
def zone_view_round2_results(request, zone_id, olympiad_id):
    """Round 2-ийн бүсийн дүнг харах (Round 3 олимпиадаас)"""
    zone = get_object_or_404(Zone, id=zone_id)
    round3_olympiad = get_object_or_404(Olympiad, id=olympiad_id)

    if not user_can_manage_zone(request.user, zone):
        messages.error(request, 'Хандах эрхгүй.')
        return redirect('my_managed_zones')

    round2_olympiads = Olympiad.objects.filter(next_round=round3_olympiad)
    if not round2_olympiads.exists():
        round2_olympiads = Olympiad.objects.filter(
            school_year=round3_olympiad.school_year,
            level=round3_olympiad.level,
            round=2,
        )

    if not round2_olympiads.exists():
        messages.warning(request, 'Round 2 олимпиад олдсонгүй.')
        return redirect('zone_olympiad_view', zone_id=zone_id, olympiad_id=olympiad_id)

    zone_provinces = Province.objects.filter(zone=zone).order_by('name')

    # Аймаг сонгох (GET parameter)
    selected_province_id = request.GET.get('province')
    selected_province = None

    if selected_province_id:
        try:
            selected_province = zone_provinces.get(id=int(selected_province_id))
        except (ValueError, Province.DoesNotExist):
            selected_province = None

    if selected_province:
        filter_province_ids = [selected_province.id]
    else:
        filter_province_ids = zone_provinces.values_list('id', flat=True)

    scoresheets = ScoreSheet.objects.filter(
        olympiad__in=round2_olympiads,
        user__data__province_id__in=filter_province_ids,
        is_official=True,
        total__gt=0,
    ).select_related('user', 'user__data__school', 'user__data__province', 'olympiad').order_by('-total', 'user__last_name')

    # 2.2 эрх авсан хэрэглэгчдийн ID-г нэг query-гаар авах
    qualified_user_ids = set(Award.objects.filter(
        olympiad__in=round2_olympiads,
        place__startswith='2.2 эрх',
        contestant_id__in=scoresheets.values_list('user_id', flat=True)
    ).values_list('contestant_id', flat=True))

    students_list = []
    for scoresheet in scoresheets:
        students_list.append({
            'user': scoresheet.user,
            'school': scoresheet.user.data.school if hasattr(scoresheet.user, 'data') else None,
            'province': scoresheet.user.data.province if hasattr(scoresheet.user, 'data') else None,
            'total': scoresheet.total or 0,
            'olympiad': scoresheet.olympiad,
            'is_qualified': scoresheet.user_id in qualified_user_ids,
        })

    context = {
        'zone': zone,
        'round3_olympiad': round3_olympiad,
        'round2_olympiads': round2_olympiads,
        'zone_provinces': zone_provinces,
        'selected_province': selected_province,
        'students': students_list,
    }
    return render(request, 'provinces/zone_round2_results.html', context)
