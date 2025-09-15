import pandas as pd
import random
from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from schools.models import School
from olympiad.models import Olympiad
# Excel-г загварчлахад ашиглах сангууд
from openpyxl.styles import Border, Side, Font, Alignment

class Command(BaseCommand):
    help = 'Сонгосон олимпиад, сургуулийн дагуу 2 sheet-тэй Excel загвар бэлдэнэ.'

    def add_arguments(self, parser):
        parser.add_argument('--olympiad-id', type=int, required=True, help='Бодлогын жагсаалт болон ангиллыг авах Олимпиадын ID')
        parser.add_argument('--school-id', type=int, required=True, help='Оролцогчдын жагсаалтыг авах Сургуулийн ID')
        parser.add_argument('--output-file', type=str, required=True, help='Хадгалах файлын зам ба нэр (жишээ нь: /tmp/answers.xlsx)')
        # --- ШИНЭ ПАРАМЕТР ---
        parser.add_argument(
            '--test',
            action='store_true',
            help='Импорт шалгах зорилгоор санамсаргүй хариултаар дүүргэсэн туршилтын файл үүсгэх'
        )

    def handle(self, *args, **options):
        olympiad_id = options['olympiad_id']
        school_id = options['school_id']
        output_file = options['output_file']
        is_test_mode = options['test'] # --test параметрийн утгыг авах

        try:
            olympiad = Olympiad.objects.select_related('level').get(pk=olympiad_id)
            school = School.objects.get(pk=school_id)
        except Olympiad.DoesNotExist:
            raise CommandError(f'ID={olympiad_id} бүхий олимпиад олдсонгүй.')
        except School.DoesNotExist:
            raise CommandError(f'ID={school_id} бүхий сургууль олдсонгүй.')

        if not school.group:
            raise CommandError(f"'{school.name}' сургуульд групп оноогоогүй тул оролцогчдын жагсаалтыг гаргах боломжгүй.")

        if not olympiad.level:
            raise CommandError(f"'{olympiad.name}' олимпиадад ангилал (level) оноогоогүй байна.")

        level = olympiad.level
        contestants = User.objects.filter(
            groups=school.group,
            data__level=level
        ).order_by('last_name', 'first_name')

        info_text = f"'{school.name}' сургуулийн '{level.name}' ангиллын"

        if not contestants.exists():
            self.stdout.write(self.style.WARNING(f'{info_text} оролцогч олдсонгүй.'))
            return

        self.stdout.write(self.style.NOTICE(f"'{olympiad.name}' олимпиадын {info_text} оролцогчдод зориулсан загварыг бэлдэж байна..."))

        problems = olympiad.problem_set.all().order_by('order')

        # --- 1-р Sheet: Хариулт ---
        answers_data = []
        for contestant in contestants:
            user_data = {
                'ID': contestant.id,
                'Овог': contestant.last_name,
                'Нэр': contestant.first_name,
            }
            for problem in problems:
                answer = ''
                # --- ТУРШИЛТЫН ФАЙЛ ҮҮСГЭХ ЛОГИК ---
                if is_test_mode:
                    # 1-ээс 1000-ийн хооронд санамсаргүй натурал тоо оноох
                    answer = random.randint(1, 1000)
                user_data[f'№{problem.order}'] = answer
            answers_data.append(user_data)

        header = ['ID', 'Овог', 'Нэр'] + [f'№{p.order}' for p in problems]
        answers_df = pd.DataFrame(answers_data, columns=header)

        # --- 2-р Sheet: Info ---
        info_data = {
            'Түлхүүр': ['olympiad_id', 'olympiad_name', 'school_id', 'school_name', 'level_id', 'level_name'],
            'Утга': [olympiad.id, olympiad.name, school.id, school.name, level.id, level.name]
        }
        info_df = pd.DataFrame(info_data)


        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                answers_df.to_excel(writer, sheet_name='Хариулт', index=False)
                info_df.to_excel(writer, sheet_name='Мэдээлэл', index=False)

                workbook = writer.book
                worksheet = writer.sheets['Хариулт']

                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                bold_font = Font(bold=True)
                center_align = Alignment(horizontal='center', vertical='center')

                for cell in worksheet["1:1"]:
                    cell.font = bold_font
                    cell.border = thin_border
                    cell.alignment = center_align

                for row in worksheet.iter_rows(min_row=2,
                                               max_row=worksheet.max_row,
                                               max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border
                # Баганын өргөнийг тохируулах
                for i, column_cells in enumerate(worksheet.columns):
                    column_letter = column_cells[0].column_letter
                    # Эхний 3 баганыг (ID, Овог, Нэр) автоматаар тохируулах
                    if i < 3:
                        length = max(len(str(cell.value)) for cell in column_cells)
                        worksheet.column_dimensions[column_letter].width = length + 2
                    # Бусад (хариултын) багануудыг тогтмол хэмжээтэй болгох
                    else:
                        # 1 инч нь ойролцоогоор 5 нэгж өргөнтэй тэнцүү
                        worksheet.column_dimensions[column_letter].width = 10

        except Exception as e:
            raise CommandError(f"Файл хадгалахад алдаа гарлаа: {e}")

        self.stdout.write(self.style.SUCCESS(
            f'Амжилттай үүслээ! "{output_file}" файлд 2 sheet-тэй хариултын хуудсыг хадгаллаа.'
        ))