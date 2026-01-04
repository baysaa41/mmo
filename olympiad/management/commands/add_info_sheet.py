import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from django.core.management.base import BaseCommand, CommandError


class Command(BaseCommand):
    help = 'Excel файлуудад Мэдээлэл sheet нэмэх'

    def add_arguments(self, parser):
        parser.add_argument(
            'folder',
            type=str,
            help='Excel файлууд байрлах фолдерын зам'
        )
        parser.add_argument(
            '--province-id',
            type=int,
            default=None,
            help='Бүх файлд нэмэх province_id (хэрэв заагаагүй бол хэрэглэгчээс асууна)'
        )

    def handle(self, *args, **options):
        folder_path = options['folder']
        default_province_id = options['province_id']

        if not os.path.isdir(folder_path):
            raise CommandError(f"Фолдер олдсонгүй: {folder_path}")

        self.stdout.write(self.style.NOTICE(f"📁 Фолдер: {folder_path}"))
        self.stdout.write("")

        # Бүх xlsx файлуудыг олох
        xlsx_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and not f.startswith('~')]

        if not xlsx_files:
            self.stdout.write(self.style.WARNING("⚠️ Excel файл олдсонгүй"))
            return

        self.stdout.write(f"📊 Нийт {len(xlsx_files)} файл олдлоо")
        self.stdout.write("")

        processed = 0
        skipped = 0
        added = 0

        for filename in sorted(xlsx_files):
            filepath = os.path.join(folder_path, filename)

            try:
                # Excel файл нээх
                wb = openpyxl.load_workbook(filepath)
                sheet_names = wb.sheetnames

                # Мэдээлэл sheet байгаа эсэхийг шалгах
                has_info_sheet = any('Мэдээлэл' in name or 'МЭДЭЭЛЭЛ' in name for name in sheet_names)

                if has_info_sheet:
                    self.stdout.write(f"  ⏭️  {filename}: Мэдээлэл sheet аль хэдийн байна")
                    skipped += 1
                    continue

                # Province ID олох
                province_id = default_province_id

                if province_id is None:
                    # Файлын нэрээс эсвэл хэрэглэгчээс province_id авах
                    self.stdout.write(f"\n📄 {filename}")
                    province_input = input("   Province ID оруулна уу (алгасах бол Enter): ").strip()

                    if province_input:
                        try:
                            province_id = int(province_input)
                        except ValueError:
                            self.stdout.write(self.style.WARNING(f"   ⚠️ Буруу province_id: {province_input}"))
                            skipped += 1
                            continue
                    else:
                        self.stdout.write(self.style.WARNING("   ⏭️  Алгасагдлаа"))
                        skipped += 1
                        continue

                # Мэдээлэл sheet үүсгэх
                ws_info = wb.create_sheet("Мэдээлэл", 0)  # Эхний байрлалд оруулах

                # Header
                ws_info['A1'] = 'key'
                ws_info['B1'] = 'value'

                # Province ID
                ws_info['A2'] = 'province_id'
                ws_info['B2'] = province_id

                # Файлын нэр
                file_display_name = os.path.splitext(filename)[0]
                ws_info['A3'] = 'Нэр'
                ws_info['B3'] = file_display_name

                # Хадгалах
                wb.save(filepath)
                wb.close()

                self.stdout.write(self.style.SUCCESS(
                    f"  ✅ {filename}: Мэдээлэл sheet нэмэгдлээ (province_id={province_id})"
                ))
                added += 1
                processed += 1

            except Exception as e:
                self.stdout.write(self.style.ERROR(
                    f"  ❌ {filename}: Алдаа гарлаа - {e}"
                ))
                continue

        # Тайлан
        self.stdout.write("")
        self.stdout.write("=" * 80)
        self.stdout.write(self.style.SUCCESS("📊 ТАЙЛАН"))
        self.stdout.write("=" * 80)
        self.stdout.write(f"📁 Фолдер: {os.path.abspath(folder_path)}")
        self.stdout.write(f"📄 Нийт файл: {len(xlsx_files)}")
        self.stdout.write(f"✅ Мэдээлэл нэмэгдсэн: {added}")
        self.stdout.write(f"⏭️  Алгасагдсан: {skipped}")
        self.stdout.write("=" * 80)
