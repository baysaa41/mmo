import os
import openpyxl
from olympiad.models import Result, Olympiad, Problem
from django.contrib.auth.models import User

# Define mappings based on sheet type.
problem_id_map = {
    'C (5-6)': [995, 996, 997, 998, 999, 1000, 1001, 1002, 1003, 1004],
    'D (7-8)': [1005, 1006, 1007, 1008, 1009, 1010, 1011, 1012, 1013, 1014],
    'E (9-10)': [1015, 1016, 1017, 1018, 1019, 1020, 1021, 1022, 1023, 1024, 1025, 1026],
    'F (11-12)': [1027, 1028, 1029, 1030, 1031, 1032, 1033, 1034, 1035, 1036, 1037, 1038]
}

olympiad_id_map = {
    'C (5-6)': 160,
    'D (7-8)': 161,
    'E (9-10)': 162,
    'F (11-12)': 163
}

def validate_score(value):
    """Helper function to validate that a score is a positive integer or None."""
    if value is None:
        return None
    elif isinstance(value, int) and value > 0:
        return value
    else:
        return None

def read_all_sheets_from_excel(directory_path):
    # Check if the directory exists
    if not os.path.isdir(directory_path):
        print(f"The directory {directory_path} does not exist.")
        return

    # Loop through each .xlsx file in the specified directory
    for filename in os.listdir(directory_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(directory_path, filename)
            workbook = openpyxl.load_workbook(file_path)

            # Loop through each specified sheet
            for sheet_name in ['C (5-6)', 'D (7-8)', 'E (9-10)', 'F (11-12)']:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    olympiad_id = olympiad_id_map[sheet_name]
                    problems = problem_id_map[sheet_name]

                    # Set headers based on sheet name
                    if sheet_name in ['C (5-6)', 'D (7-8)']:
                        columns = [
                            'number', 'student_id', 'surname', 'name', 'district_code',
                            'school', 'grade', 'score1', 'score2', 'score3', 'score4',
                            'score5', 'score6', 'score7', 'score8', 'score9', 'score10'
                        ]
                    else:  # For sheets E (9-10) and F (11-12)
                        columns = [
                            'number', 'student_id', 'surname', 'name', 'district_code',
                            'school', 'grade', 'score1', 'score2', 'score3', 'score4',
                            'score5', 'score6', 'score7', 'score8', 'score9', 'score10',
                            'score11', 'score12'
                        ]

                    # Read each row in the Excel sheet, skipping the header
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        # Create a dictionary for each record based on the column headers
                        record = {columns[i]: row[i] for i in range(len(columns))}
                        contestant_id = record['student_id']

                        # Validate contestant existence (if necessary)
                        try:
                            contestant = User.objects.get(id=contestant_id)
                        except User.DoesNotExist:
                            print(f"Contestant with ID {contestant_id} does not exist. Skipping row.")
                            continue

                        # Validate each score column and create or update Result objects
                        for i, problem_id in enumerate(problems):
                            score_column = columns[7 + i]  # Start from 'score1' index
                            answer = validate_score(record.get(score_column))

                            if answer is not None:
                                # Try to get the existing Result object
                                result, created = Result.objects.update_or_create(
                                    contestant=contestant,
                                    olympiad_id=olympiad_id,
                                    problem_id=problem_id,
                                    defaults={
                                        'answer': answer,
                                        'state': Result.States.not_submitted,  # Default state
                                        'is_active': True  # Set to True if active
                                    }
                                )
                                if created:
                                    print(f"Created result for contestant {contestant_id} on problem {problem_id}.")
                                else:
                                    print(f"Updated result for contestant {contestant_id} on problem {problem_id}.")

    print("All records have been processed.")

# Usage example
directory_path = '/home/deploy/2024-2025-first'
read_all_sheets_from_excel(directory_path)
