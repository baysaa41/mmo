import os
import openpyxl
from olympiad.models import Result, Olympiad, Problem
from django.contrib.auth.models import User

# Define mappings based on sheet type.
problem_id_map = {
    'T': [1054, 1055, 1056, 1057, 1058],
    'D (7-8)': [1039, 1040, 1041, 1042, 1043],
    'E (9-10)': [1044, 1045, 1046, 1047, 1048],
    'F (11-12)': [1049, 1050, 1051, 1052, 1053]
}

olympiad_id_map = {
    'T': 167,
    'D (7-8)': 164,
    'E (9-10)': 165,
    'F (11-12)': 166
}

def validate_score(value):
    """Helper function to validate that a score is a positive integer or None."""
    if value is None:
        return None
    elif isinstance(value, int) and value > 0:
        return value
    else:
        return None

def read_all_sheets_from_excel(directory_path):
    # Check if the directory exists
    if not os.path.isdir(directory_path):
        print(f"The directory {directory_path} does not exist.")
        return

    # Loop through each .xlsx file in the specified directory
    for filename in os.listdir(directory_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(directory_path, filename)
            workbook = openpyxl.load_workbook(file_path)

            # Loop through each specified sheet
            for sheet_name in ['T', 'D (7-8)', 'E (9-10)', 'F (11-12)']:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    olympiad_id = olympiad_id_map[sheet_name]
                    problems = problem_id_map[sheet_name]

                    columns = [
                        'number', 'surname', 'name', 'student_id',
                        'school', 'score1', 'score2', 'score3', 'score4',
                        'score5'
                    ]

                    # Read each row in the Excel sheet, skipping the header
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        # Create a dictionary for each record based on the column headers
                        record = {columns[i]: row[i] for i in range(len(columns))}
                        contestant_id = record['student_id']

                        # Validate contestant existence (if necessary)
                        try:
                            contestant = User.objects.get(id=contestant_id)
                        except User.DoesNotExist:
                            print(f"Contestant with ID {contestant_id} does not exist. Skipping row.")
                            continue

                        # Validate each score column and create or update Result objects
                        for i, problem_id in enumerate(problems):
                            score_column = columns[5 + i]  # Start from 'score1' index
                            score = validate_score(record.get(score_column))

                            if score is not None:
                                # Try to get the existing Result object
                                result, created = Result.objects.update_or_create(
                                    contestant=contestant,
                                    olympiad_id=olympiad_id,
                                    problem_id=problem_id,
                                    defaults={
                                        'score': score,
                                        'state': Result.States.not_submitted,  # Default state
                                        'is_active': True  # Set to True if active
                                    }
                                )
                                if created:
                                    print(f"Created result for contestant {contestant_id} on problem {problem_id}.")
                                else:
                                    print(f"Updated result for contestant {contestant_id} on problem {problem_id}.")

    print("All records have been processed.")

# Usage example
directory_path = '/home/deploy/iran'
read_all_sheets_from_excel(directory_path)