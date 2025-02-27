import os
import openpyxl
from olympiad.models import Result, Olympiad, Problem
from django.contrib.auth.models import User

# Define mappings based on sheet type.
problem_id_map = {
    'D': [1103,1104,1105,1106],
    'E': [1107,1108,1109,1110,1111,1112],
    'F': [1113,1114,1115,1116,1117,1118],
    'S': [1119,1120,1121,1122,1123,1124],
    'Т': [1125,1126,1127,1128,1129,1130],
}

#egmo added
olympiad_id_map = {
    'D': 176,
    'E': 177,
    'F': 178,
    'S': 179,
    'Т': 180,
}

def validate_score(value):
    """Helper function to validate that a score is a positive integer or None."""
    if value is None:
        return None
    elif isinstance(value, int) and value > 0:
        return value
    else:
        return None

def read_all_sheets_from_excel(directory_path):
    # Check if the directory exists
    if not os.path.isdir(directory_path):
        print(f"The directory {directory_path} does not exist.")
        return

    # Loop through each .xlsx file in the specified directory
    for filename in os.listdir(directory_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(directory_path, filename)
            workbook = openpyxl.load_workbook(file_path)

            # Loop through each specified sheet
            for sheet_name in ['D', 'E', 'F', 'S', 'Т']:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    olympiad_id = olympiad_id_map[sheet_name]
                    problems = problem_id_map[sheet_name]

                    # Set headers based on sheet name
                    if sheet_name in ['D']:
                        columns = [
                            'duureg', 'school', 'student_id', 'surname', 'name', 'score1', 'score2', 'score3', 'score4'
                        ]
                    else:  # For sheets E, F, S, T
                        columns = [
                            'duureg', 'school', 'student_id', 'surname', 'name', 'score1', 'score2', 'score3', 'score4',
                            'score5', 'score6'
                        ]

                    # Read each row in the Excel sheet, skipping the header
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        # Create a dictionary for each record based on the column headers
                        record = {columns[i]: row[i] for i in range(len(columns))}
                        contestant_id = record['student_id']

                        # Validate contestant existence (if necessary)
                        try:
                            contestant = User.objects.get(id=contestant_id)
                        except User.DoesNotExist:
                            print(f"Contestant with ID {contestant_id} does not exist. Skipping row.")
                            continue

                        # Validate each score column and create or update Result objects
                        for i, problem_id in enumerate(problems):
                            score_column = columns[5 + i]  # Start from 'score1' index
                            score = validate_score(record.get(score_column))

                            if score is not None:
                                # Try to get the existing Result object
                                result, created = Result.objects.update_or_create(
                                    contestant=contestant,
                                    olympiad_id=olympiad_id,
                                    problem_id=problem_id,
                                    defaults={
                                        'score': score,
                                        'state': Result.States.not_submitted,  # Default state
                                        'is_active': True  # Set to True if active
                                    }
                                )
                                if created:
                                    print(f"Created result for contestant {contestant_id} on problem {problem_id}.")
                                else:
                                    print(f"Updated result for contestant {contestant_id} on problem {problem_id}.")
                            else:
                                # Try to get the existing Result object
                                result, created = Result.objects.update_or_create(
                                    contestant=contestant,
                                    olympiad_id=olympiad_id,
                                    problem_id=problem_id,
                                    defaults={
                                        'score': 0,
                                        'state': Result.States.not_submitted,  # Default state
                                        'is_active': True  # Set to True if active
                                    }
                                )
                                if created:
                                    print(f"Created a result for contestant {contestant_id} on problem {problem_id}.")
                                else:
                                    print(f"Updated a result for contestant {contestant_id} on problem {problem_id}.")

    print("All records have been processed.")

# Usage example
directory_path = '/home/deploy/2024-2025-second'
read_all_sheets_from_excel(directory_path)