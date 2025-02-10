import os
import openpyxl
from olympiad.models import Result, Olympiad, Problem
from django.contrib.auth.models import User

# Define mappings based on sheet type.
problem_id_map = {
    'C (5-6)': [1059,1060,1061,1062],
    'D (7-8)': [1063,1064,1065,1066],
    'E (9-10)': [1067,1068,1069,1070,1071,1072],
    'F (11-12)': [1073,1074,1075,1076,1077,1078],
    'S (ББ)': [1079,1080,1081,1082,1083,1084],
    'Т (ДБ)': [1085,1086,1087,1088,1089,1090],
    'EGMO-1': [1091,1092,1093,1094,1095,1096],
    'EGMO-2': [1097,1098,1099,1101,1102,1103],
}

#egmo added
olympiad_id_map = {
    'C (5-6)': 168,
    'D (7-8)': 169,
    'E (9-10)': 170,
    'F (11-12)': 171,
    'S (ББ)': 172,
    'Т (ДБ)': 173,
    'EGMO-1': 174,
    'EGMO-2': 175,
}

def validate_score(value):
    """Helper function to validate that a score is a positive integer or None."""
    if value is None:
        return None
    elif isinstance(value, int) and value > 0:
        return value
    else:
        return None

def read_all_sheets_from_excel(directory_path):
    # Check if the directory exists
    if not os.path.isdir(directory_path):
        print(f"The directory {directory_path} does not exist.")
        return

    # Loop through each .xlsx file in the specified directory
    for filename in os.listdir(directory_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(directory_path, filename)
            workbook = openpyxl.load_workbook(file_path)

            # Loop through each specified sheet
            for sheet_name in ['C (5-6)', 'D (7-8)', 'E (9-10)', 'F (11-12)', 'S (ББ)', 'Т (ДБ)', 'EGMO-1', 'EGMO-2']:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    olympiad_id = olympiad_id_map[sheet_name]
                    problems = problem_id_map[sheet_name]

                    # Set headers based on sheet name
                    if sheet_name in ['C (5-6)', 'D (7-8)']:
                        columns = [
                            'number', 'student_id', 'surname', 'name', 'score1', 'score2', 'score3', 'score4'
                        ]
                    else:  # For sheets E (9-10) and F (11-12)
                        columns = [
                            'number', 'student_id', 'surname', 'name', 'score1', 'score2', 'score3', 'score4',
                            'score5', 'score6'
                        ]

                    # Read each row in the Excel sheet, skipping the header
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        # Create a dictionary for each record based on the column headers
                        record = {columns[i]: row[i] for i in range(len(columns))}
                        contestant_id = record['student_id']

                        # Validate contestant existence (if necessary)
                        try:
                            contestant = User.objects.get(id=contestant_id)
                        except User.DoesNotExist:
                            print(f"Contestant with ID {contestant_id} does not exist. Skipping row.")
                            continue

                        # Validate each score column and create or update Result objects
                        for i, problem_id in enumerate(problems):
                            score_column = columns[4 + i]  # Start from 'score1' index
                            score = validate_score(record.get(score_column))

                            if score is not None:
                                # Try to get the existing Result object
                                result, created = Result.objects.update_or_create(
                                    contestant=contestant,
                                    olympiad_id=olympiad_id,
                                    problem_id=problem_id,
                                    defaults={
                                        'score': score,
                                        'state': Result.States.not_submitted,  # Default state
                                        'is_active': True  # Set to True if active
                                    }
                                )
                                if created:
                                    print(f"Created result for contestant {contestant_id} on problem {problem_id}.")
                                else:
                                    print(f"Updated result for contestant {contestant_id} on problem {problem_id}.")
                            else:
                                # Try to get the existing Result object
                                result, created = Result.objects.update_or_create(
                                    contestant=contestant,
                                    olympiad_id=olympiad_id,
                                    problem_id=problem_id,
                                    defaults={
                                        'score': 0,
                                        'state': Result.States.not_submitted,  # Default state
                                        'is_active': True  # Set to True if active
                                    }
                                )
                                if created:
                                    print(f"Created 0 result for contestant {contestant_id} on problem {problem_id}.")
                                else:
                                    print(f"Updated 0 result for contestant {contestant_id} on problem {problem_id}.")

    print("All records have been processed.")

# Usage example
directory_path = '/home/deploy/2024-2025-second'
read_all_sheets_from_excel(directory_path)