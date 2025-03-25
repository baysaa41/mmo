import os
import openpyxl
from olympiad.models import Result, Olympiad, Problem
from django.contrib.auth.models import User

# Define mappings based on sheet type.
problem_id_map = {
    'D': [1131,1132,1133,1134],
    'E': [1135,1136,1137,1138,1139,1140],
    'F': [1141,1142,1143,1144,1145,1146],
    'S': [1147,1148,1149,1150,1151,1152],
    'T': [1153,1154,1155,1156,1157,1158],
    'IMO-1': [1159,1160,1161,1162,1163,1164],
}

#egmo added
olympiad_id_map = {
    'D': 181,
    'E': 182,
    'F': 183,
    'S': 184,
    'T': 185,
    'IMO-1': 186,
}

def validate_score(value):
    """Helper function to validate that a score is a positive integer or None."""
    if value is None:
        return None
    elif isinstance(value, int) and value > 0:
        return value
    else:
        return None

def read_all_sheets_from_excel(directory_path):
    # Check if the directory exists
    if not os.path.isdir(directory_path):
        print(f"The directory {directory_path} does not exist.")
        return

    # Loop through each .xlsx file in the specified directory
    for filename in os.listdir(directory_path):
        if filename.endswith('.xlsx'):
            print(filename)
            file_path = os.path.join(directory_path, filename)
            workbook = openpyxl.load_workbook(file_path)

            # Loop through each specified sheet
            for sheet_name in ['D', 'E', 'F', 'S', 'T', 'IMO-1']:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    olympiad_id = olympiad_id_map[sheet_name]
                    problems = problem_id_map[sheet_name]

                    # Set headers based on sheet name
                    if sheet_name in ['D']:
                        columns = [
                            'num', 'student_id', 'name', 'score1', 'score2', 'score3', 'score4'
                        ]
                    else:  # For sheets E, F, S, T, IMO-1
                        columns = [
                            'num', 'student_id', 'name', 'score1', 'score2', 'score3', 'score4',
                            'score5', 'score6'
                        ]

                    # Read each row in the Excel sheet, skipping the header
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        # Create a dictionary for each record based on the column headers
                        record = {columns[i]: row[i] for i in range(len(columns))}
                        contestant_id = record['student_id']

                        # Validate contestant existence (if necessary)
                        try:
                            contestant = User.objects.get(id=contestant_id)
                        except User.DoesNotExist:
                            print(f"Contestant with ID {contestant_id} does not exist. Skipping row.")
                            continue

                        # Validate each score column and create or update Result objects
                        for i, problem_id in enumerate(problems):
                            score_column = columns[3 + i]  # Start from 'score1' index
                            score = validate_score(record.get(score_column))

                            if score is not None:
                                # Try to get the existing Result object
                                result, created = Result.objects.update_or_create(
                                    contestant=contestant,
                                    olympiad_id=olympiad_id,
                                    problem_id=problem_id,
                                    defaults={
                                        'score': score,
                                        'state': Result.States.not_submitted,  # Default state
                                        'is_active': True  # Set to True if active
                                    }
                                )
                                if created:
                                    print(f"Created result for contestant {contestant_id} on problem {problem_id}.")
                                else:
                                    print(f"Updated result for contestant {contestant_id} on problem {problem_id}.")
                            else:
                                # Try to get the existing Result object
                                result, created = Result.objects.update_or_create(
                                    contestant=contestant,
                                    olympiad_id=olympiad_id,
                                    problem_id=problem_id,
                                    defaults={
                                        'score': 0,
                                        'state': Result.States.not_submitted,  # Default state
                                        'is_active': True  # Set to True if active
                                    }
                                )
                                if created:
                                    print(f"Created a result for contestant {contestant_id} on problem {problem_id}.")
                                else:
                                    print(f"Updated a result for contestant {contestant_id} on problem {problem_id}.")

    print("All records have been processed.")

# Usage example
directory_path = '/home/deploy/2024-2025-second'
read_all_sheets_from_excel(directory_path)